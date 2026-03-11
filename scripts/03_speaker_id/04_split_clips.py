#!/usr/bin/env python3
"""
YTAI Этап 4: Split Clips
Разбивка транскрипта по видео клипам.

Использование:
    python split_clips.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani"

Выход:
    02_Transcripts/02_04_ByClips/
    ├── {project}_split_clips_{timestamp}.xlsx
    ├── RYA-ZVE1-1146.srt
    ├── RYA-ZVE1-1147.srt
    └── ...
"""

import argparse
import sys
import json
from pathlib import Path
from datetime import datetime
from typing import List

sys.path.insert(0, str(Path(__file__).parent))

from utils import (
    get_project_paths,
    ensure_dirs,
    find_latest_file,
    cleanup_old_files,
    generate_timestamp,
    format_timestamp,
    format_srt_timestamp,
    setup_logging,
    print_header,
    get_video_clips,
    find_clip_for_timestamp,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================================
# Сохранение файлов
# ============================================================================

def save_clips_xlsx(
    segments: List[dict],
    clips: List[dict],
    output_path: Path,
    logger
) -> int:
    """
    Сохранить XLSX таблицу с сегментами по клипам.
    
    Колонки: file | timecode | speaker | text
    
    Returns:
        Количество смапленных сегментов
    """
    if not HAS_OPENPYXL:
        logger.warning("openpyxl не установлен, XLSX не создан")
        return 0
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript"
    
    # Заголовок
    headers = ["file", "timecode", "speaker", "text"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
    
    # Данные
    row = 2
    mapped = 0
    
    for seg in segments:
        global_ts = seg["start"]
        clip, local_ts = find_clip_for_timestamp(clips, global_ts)
        
        if clip is None:
            continue
        
        ws.cell(row=row, column=1, value=clip["file"])
        ws.cell(row=row, column=2, value=format_timestamp(local_ts))
        ws.cell(row=row, column=3, value=seg.get("speaker", "UNKNOWN"))
        ws.cell(row=row, column=4, value=seg.get("text", ""))
        
        row += 1
        mapped += 1
    
    # Ширина колонок
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 80
    
    wb.save(output_path)
    
    return mapped


def save_clip_srt(
    segments: List[dict],
    clip: dict,
    output_path: Path
) -> int:
    """
    Сохранить SRT для одного клипа с локальными таймкодами.
    
    Returns:
        Количество субтитров
    """
    lines = []
    idx = 1
    
    for seg in segments:
        global_ts = seg["start"]
        
        # Проверить что сегмент принадлежит этому клипу
        if not (clip["start"] <= global_ts < clip["end"]):
            continue
        
        # Локальные таймкоды
        local_start = global_ts - clip["start"]
        local_end = seg["end"] - clip["start"]
        local_end = min(local_end, clip["duration"])
        
        text = seg.get("text", "").strip()
        if not text:
            continue
        
        speaker = seg.get("speaker", "UNKNOWN")
        
        lines.append(str(idx))
        lines.append(f"{format_srt_timestamp(local_start)} --> {format_srt_timestamp(local_end)}")
        lines.append(f"[{speaker}] {text}")
        lines.append("")
        
        idx += 1
    
    if lines:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    
    return idx - 1


def save_all_clip_srts(
    segments: List[dict],
    clips: List[dict],
    output_dir: Path,
    logger
) -> int:
    """
    Сохранить SRT файлы для всех клипов.
    
    Returns:
        Количество созданных файлов
    """
    count = 0
    
    for clip in clips:
        # Имя SRT = имя клипа (без расширения)
        srt_name = Path(clip["file"]).stem + ".srt"
        srt_path = output_dir / srt_name
        
        num_subs = save_clip_srt(segments, clip, srt_path)
        
        if num_subs > 0:
            logger.debug(f"  {srt_name}: {num_subs} субтитров")
            count += 1
    
    return count


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="YTAI Этап 4: Разбивка транскрипта по видео клипам",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры:
    python split_clips.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani"
        """
    )
    
    parser.add_argument("--project", required=True, help="Путь к папке проекта")
    parser.add_argument("--named-json", type=Path, help="Конкретный JSON с именами")
    
    args = parser.parse_args()
    
    # Получить пути
    try:
        paths = get_project_paths(args.project)
    except FileNotFoundError as e:
        print(f"ОШИБКА: {e}", file=sys.stderr)
        sys.exit(1)
    
    ensure_dirs(paths)
    
    project_name = paths["project_name"]
    timestamp = generate_timestamp()
    
    logger = setup_logging(paths["logs_dir"], project_name, "split_clips")
    
    # Найти транскрипт с именами
    if args.named_json:
        named_path = args.named_json
    else:
        named_path = find_latest_file(
            paths["clean_dir"],
            f"{project_name}_apply_names_*.json"
        )
    
    if not named_path or not named_path.exists():
        logger.error("Транскрипт с именами не найден. Сначала запустите apply_names.py")
        sys.exit(1)
    
    # Проверить наличие видео
    if not paths["video_dir"].exists():
        logger.error(f"Папка с видео не найдена: {paths['video_dir']}")
        sys.exit(1)
    
    # ================================================================
    # Заголовок
    # ================================================================
    print_header("YTAI ЭТАП 4: РАЗБИВКА ПО КЛИПАМ")
    logger.info(f"Проект: {project_name}")
    logger.info(f"Транскрипт: {named_path.name}")
    logger.info("")
    
    # ================================================================
    # Загрузка данных
    # ================================================================
    logger.info("Загрузка транскрипта...")
    
    with open(named_path, encoding="utf-8") as f:
        data = json.load(f)
    
    segments = data.get("segments", [])
    speakers = data.get("speakers", [])
    
    logger.info(f"  Сегментов: {len(segments)}")
    logger.info(f"  Спикеров: {', '.join(speakers)}")
    
    # ================================================================
    # Анализ видео клипов
    # ================================================================
    logger.info("")
    logger.info("Анализ видео клипов...")
    
    clips = get_video_clips(paths["video_dir"], logger)
    
    if not clips:
        logger.error("Видео клипы не найдены!")
        sys.exit(1)
    
    # ================================================================
    # Сохранение
    # ================================================================
    logger.info("")
    logger.info("Сохранение файлов...")
    
    # Очистить старые файлы (кроме backup)
    old_count = cleanup_old_files(
        paths["clean_dir"],
        f"{project_name}_split_clips_*"
    )
    
    # Также очистить старые SRT
    old_srts = cleanup_old_files(
        paths["clean_dir"],
        "*.srt"
    )
    
    if old_count > 0 or old_srts > 0:
        logger.info(f"  Перемещено в backup: {old_count + old_srts} файлов")
    
    # XLSX
    if HAS_OPENPYXL:
        xlsx_path = paths["clean_dir"] / f"{project_name}_split_clips_{timestamp}.xlsx"
        mapped = save_clips_xlsx(segments, clips, xlsx_path, logger)
        logger.info(f"  XLSX: {xlsx_path.name} ({mapped} сегментов)")
    else:
        logger.warning("  XLSX пропущен (установите: pip install openpyxl)")
    
    # Per-clip SRT файлы
    logger.info("")
    logger.info("Создание SRT для каждого клипа...")
    
    srt_count = save_all_clip_srts(segments, clips, paths["clean_dir"], logger)
    logger.info(f"  Создано {srt_count} SRT файлов")
    
    # ================================================================
    # Итог
    # ================================================================
    print_header("ГОТОВО", char="=")
    logger.info(f"Выходная папка: {paths['clean_dir']}")
    logger.info("")
    logger.info(f"  Клипов: {len(clips)}")
    logger.info(f"  SRT файлов: {srt_count}")
    logger.info("")
    logger.info("Проверка в Premiere Pro:")
    logger.info("  1. Импортируйте видео клип (например RYA-ZVE1-1160.MP4)")
    logger.info("  2. Импортируйте соответствующий SRT (RYA-ZVE1-1160.srt)")
    logger.info("  3. Убедитесь что субтитры совпадают с речью по времени")
    logger.info("")
    logger.info("ОБРАБОТКА ЗАВЕРШЕНА!")


if __name__ == "__main__":
    main()
