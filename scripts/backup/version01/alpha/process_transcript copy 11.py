#!/usr/bin/env python3
"""
YTAI Step 2: Process Transcript
Speaker Naming (Individual LLM Analysis) + Split by Video Clips

This script takes the JSON output from Step 1 (transcribe_project.py) and:
1. Analyzes EACH speaker INDIVIDUALLY with LLM for accurate identification
2. Names speakers based on their actual speech content
3. Splits transcript by video clips using ffprobe durations
4. Generates per-clip SRT files for Premiere Pro

Usage:
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani"
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani" --skip-llm

Output:
    02_Transcripts/02_01_Runs/
        *_named.json, *_named.txt, *_named.srt, *_speakers.txt, *_by_clips.xlsx
    
    02_Transcripts/02_01_Runs/speaker_analysis/
        SPEAKER_XX_utterances.txt  (debug files for verification)
    
    02_Transcripts/02_02_Clean/
        <clip_name>.srt  (per-clip SRT files)

Requirements:
    pip install requests openpyxl
    ollama serve
    brew install ffmpeg
"""

import argparse
import os
import sys
import json
import re
import subprocess
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Set, Tuple

try:
    import requests
except ImportError:
    print("ERROR: requests not installed. Run: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# Configuration
# ============================================================================

VIDEO_DIR = "01_Raw/01_01_Video"
AUDIO_DIR = "01_Raw/01_02_Audio"
TRANSCRIPTS_RUNS_DIR = "02_Transcripts/02_01_Runs"
TRANSCRIPTS_CLEAN_DIR = "02_Transcripts/02_02_Clean"
LOGS_DIR = "08_Logs"

VIDEO_EXTS = {".mp4", ".mov", ".m4v", ".mts", ".avi", ".mkv",
              ".MP4", ".MOV", ".M4V", ".MTS", ".AVI", ".MKV"}

OLLAMA_API_URL = "http://localhost:11434/api/generate"
DEFAULT_LLM_MODEL = "qwen2.5:32b"
OLLAMA_TIMEOUT = 300  # 5 minutes per speaker


# ============================================================================
# Project Paths
# ============================================================================

def get_project_paths(project_dir: str) -> dict:
    """Get all relevant paths for a project."""
    project_root = Path(project_dir).expanduser().resolve()
    
    if not project_root.exists():
        raise FileNotFoundError(f"Project folder not found: {project_root}")
    
    return {
        "project_root": project_root,
        "project_name": project_root.name,
        "video_dir": project_root / VIDEO_DIR,
        "audio_dir": project_root / AUDIO_DIR,
        "transcripts_runs_dir": project_root / TRANSCRIPTS_RUNS_DIR,
        "transcripts_clean_dir": project_root / TRANSCRIPTS_CLEAN_DIR,
        "logs_dir": project_root / LOGS_DIR,
        "speaker_map_file": project_root / TRANSCRIPTS_CLEAN_DIR / "speaker_names.json",
        "project_profile_file": project_root / "project_profile.txt",
    }


def find_latest_transcript(transcripts_dir: Path, project_name: str) -> Optional[Path]:
    """Find the most recent transcript JSON file."""
    pattern = f"{project_name}_transcript_*.json"
    files = list(transcripts_dir.glob(pattern))
    files = [f for f in files if "_named" not in f.name]
    
    if not files:
        return None
    
    files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return files[0]


# ============================================================================
# Logging
# ============================================================================

def setup_logging(logs_dir: Path, project_name: str) -> logging.Logger:
    """Setup dual logging to file and console."""
    logs_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = logs_dir / f"{project_name}_process_{timestamp}.log"
    
    logger = logging.getLogger("process")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info(f"Log: {log_file}")
    
    return logger


# ============================================================================
# Timestamp Formatting
# ============================================================================

def format_timestamp(seconds: float) -> str:
    """Format as HH:MM:SS"""
    h, r = divmod(int(seconds), 3600)
    m, s = divmod(r, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def format_srt_timestamp(seconds: float) -> str:
    """Format for SRT: HH:MM:SS,mmm"""
    h, r = divmod(int(seconds), 3600)
    m, s = divmod(r, 60)
    ms = int((seconds - int(seconds)) * 1000)
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


# ============================================================================
# Video Clip Utilities
# ============================================================================

def natural_sort_key(s: str):
    """Sort strings with embedded numbers naturally."""
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]


def get_video_duration(video_path: Path) -> float:
    """Get video duration in seconds using ffprobe."""
    cmd = [
        "ffprobe", "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        str(video_path)
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return float(result.stdout.strip())
    except Exception:
        pass
    return 0.0


def get_video_clips(video_dir: Path, logger: logging.Logger) -> List[dict]:
    """Get list of video clips with their durations and offsets."""
    clips = []
    for ext in VIDEO_EXTS:
        for f in video_dir.glob(f"*{ext}"):
            if "Archive" not in str(f) and not f.name.startswith("."):
                clips.append(f)
    
    clips.sort(key=lambda p: natural_sort_key(p.name))
    
    if not clips:
        logger.warning(f"No video clips found in {video_dir}")
        return []
    
    logger.info(f"Found {len(clips)} video clips")
    
    result = []
    current_offset = 0.0
    
    for clip_path in clips:
        duration = get_video_duration(clip_path)
        if duration <= 0:
            logger.warning(f"Could not get duration for {clip_path.name}, skipping")
            continue
        
        result.append({
            "file": clip_path.name,
            "path": clip_path,
            "duration": duration,
            "start": current_offset,
            "end": current_offset + duration
        })
        
        logger.debug(f"  {clip_path.name}: {format_timestamp(duration)} ({current_offset:.1f}s - {current_offset + duration:.1f}s)")
        current_offset += duration
    
    logger.info(f"Total duration: {format_timestamp(current_offset)}")
    return result


def find_clip_for_timestamp(clips: List[dict], timestamp: float) -> Tuple[Optional[dict], float]:
    """Find which clip contains the given timestamp."""
    for clip in clips:
        if clip["start"] <= timestamp < clip["end"]:
            return clip, timestamp - clip["start"]
    
    if clips and timestamp >= clips[-1]["end"] - 0.1:
        clip = clips[-1]
        return clip, timestamp - clip["start"]
    
    return None, 0.0


# ============================================================================
# Ollama LLM
# ============================================================================

def check_ollama_server(logger: logging.Logger) -> bool:
    """Check if Ollama server is running."""
    try:
        response = requests.get("http://localhost:11434/api/tags", timeout=5)
        if response.status_code == 200:
            models = response.json().get("models", [])
            model_names = [m.get("name", "") for m in models]
            logger.debug(f"Available models: {model_names}")
            return True
        return False
    except requests.exceptions.ConnectionError:
        return False
    except Exception as e:
        logger.debug(f"Ollama check error: {e}")
        return False


def call_ollama_api(prompt: str, model: str, logger: logging.Logger) -> Optional[str]:
    """Call Ollama API."""
    try:
        response = requests.post(
            OLLAMA_API_URL,
            json={
                "model": model,
                "prompt": prompt,
                "stream": False,
                "options": {"temperature": 0.3, "num_predict": 500}
            },
            timeout=OLLAMA_TIMEOUT
        )
        
        if response.status_code == 200:
            return response.json().get("response", "")
        else:
            logger.warning(f"Ollama API error: {response.status_code}")
            return None
            
    except requests.exceptions.Timeout:
        logger.warning(f"Ollama timeout ({OLLAMA_TIMEOUT}s)")
        return None
    except Exception as e:
        logger.warning(f"Ollama API error: {e}")
        return None


# ============================================================================
# Individual Speaker Analysis (MAIN METHOD)
# ============================================================================

def analyze_single_speaker(speaker_id: str, utterances: List[str], 
                           model: str, logger: logging.Logger) -> Tuple[str, str]:
    """
    Analyze ONE speaker with LLM to determine their name/role.
    
    Sends up to 200 utterances from ONE speaker to LLM.
    Samples from beginning, middle, and end for full coverage.
    
    Returns: (name, context)
    """
    total_count = len(utterances)
    
    # Берём до 200 реплик (хороший баланс скорости и покрытия)
    # Распределяем: начало, середина, конец — чтобы поймать имя где бы оно ни было
    max_sample = 200
    
    if total_count <= max_sample:
        sample = utterances
    else:
        # 70 из начала, 60 из середины, 70 из конца
        sample = (
            utterances[:70] +
            utterances[total_count//2 - 30 : total_count//2 + 30] +
            utterances[-70:]
        )
    
    utterances_text = "\n".join([f'{i+1}. "{text}"' for i, text in enumerate(sample)])
    
    prompt = f"""Analyze these utterances from ONE person in a YouTube interview about business in UAE/Saudi Arabia.

YOUR TASK: Determine WHO this person is - find their NAME or determine their ROLE.

UTTERANCES FROM THIS SPEAKER ({total_count} total, showing {len(sample)} from start/middle/end):

{utterances_text}

ANALYSIS INSTRUCTIONS:
1. SEARCH FOR ACTUAL NAME in the text:
   - "My name is Hadi" → Name is "Hadi"
   - "I am Roman" → Name is "Roman"  
   - "This is Ahmed speaking" → Name is "Ahmed"

2. IF NO NAME FOUND, determine ROLE by speech patterns:
   - Asks questions like "Could you tell us?", "What do you think?" → "Host" (interviewer)
   - Talks about OWN business: "my company", "we started", "I own" → "Guest" (interviewee)
   - Technical terms about coffee/equipment, demonstrations → "Barista"
   - Only short responses: "yeah", "okay", "right" → "Minor"

3. IMPORTANT CLUES:
   - Person asking questions is the HOST/INTERVIEWER
   - Person explaining their business at length is the GUEST
   - The channel host is usually named "Roman"

RESPOND WITH ONLY JSON (no other text):
{{
  "name": "Hadi",
  "found_actual_name": true,
  "role": "guest",
  "reasoning": "Said 'My name is Hadi' and talks about his coffee business"
}}

Rules for "name" field:
- Use actual name if found (Hadi, Roman, Ahmed, etc.)
- If no name, use role: Host, Guest, Barista, Employee, Minor
- NO parentheses, NO descriptions - just the name/role
- Maximum 2 words"""

    response = call_ollama_api(prompt, model, logger)
    
    if not response:
        return f"Speaker", "LLM no response"
    
    # Парсим JSON
    try:
        json_match = re.search(r'\{[^{}]*\}', response, re.DOTALL)
        if json_match:
            data = json.loads(json_match.group())
            
            name = data.get("name", "Speaker")
            found_name = data.get("found_actual_name", False)
            role = data.get("role", "unknown")
            reasoning = data.get("reasoning", "")
            
            # Очистка имени
            name = re.sub(r'[/\\()\[\]"]', '', name).strip()
            words = name.split()
            if len(words) > 2:
                name = ' '.join(words[:2])
            
            name = name.title()
            
            context = f"{role}: {reasoning}" if reasoning else role
            
            logger.info(f"  → {name} (found_name={found_name}, role={role})")
            
            return name, context
            
    except (json.JSONDecodeError, KeyError) as e:
        logger.warning(f"  Parse error: {e}")
        logger.debug(f"  Response was: {response[:200]}")
    
    return "Speaker", "Parse error"


def identify_speakers_individual(segments: List[dict], speakers: Set[str], 
                                  model: str, logger: logging.Logger,
                                  debug_dir: Optional[Path] = None) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Identify speakers by analyzing EACH ONE INDIVIDUALLY.
    
    This is the main speaker identification function.
    Each speaker is sent to LLM separately for accurate analysis.
    """
    if not check_ollama_server(logger):
        logger.error("Ollama server not running!")
        logger.error("Start with: ollama serve")
        return {}, {}
    
    # Собираем реплики по спикерам
    speaker_utterances = {s: [] for s in speakers}
    
    for seg in segments:
        speaker = seg.get("speaker", "UNKNOWN")
        text = seg.get("text", "").strip()
        if speaker in speakers and text:
            speaker_utterances[speaker].append(text)
    
    # Сортируем по количеству (начинаем с активных)
    sorted_speakers = sorted(speakers, key=lambda s: len(speaker_utterances[s]), reverse=True)
    
    logger.info(f"Analyzing {len(speakers)} speakers individually with {model}...")
    
    # Создаём debug директорию
    if debug_dir:
        debug_dir.mkdir(parents=True, exist_ok=True)
    
    speaker_names = {}
    speaker_context = {}
    used_names = set()
    
    for i, speaker in enumerate(sorted_speakers, 1):
        utterances = speaker_utterances[speaker]
        count = len(utterances)
        
        logger.info(f"[{i}/{len(speakers)}] {speaker} ({count} utterances)")
        
        # Сохраняем debug файл
        if debug_dir:
            debug_file = debug_dir / f"{speaker}_utterances.txt"
            with open(debug_file, "w", encoding="utf-8") as f:
                f.write(f"# {speaker} - {count} utterances\n\n")
                for j, text in enumerate(utterances, 1):
                    f.write(f"{j}. {text}\n\n")
        
        # Пропускаем спикеров с <5 репликами
        if count < 5:
            speaker_names[speaker] = "Minor"
            speaker_context[speaker] = f"Only {count} utterances"
            logger.info(f"  → Minor (too few utterances)")
            continue
        
        # Анализируем спикера
        name, context = analyze_single_speaker(speaker, utterances, model, logger)
        
        # Проверяем дубликаты
        name_lower = name.lower()
        if name_lower in used_names:
            # Добавляем номер
            counter = 2
            while f"{name_lower} {counter}" in used_names:
                counter += 1
            name = f"{name} {counter}"
            logger.info(f"  Renamed to avoid duplicate: {name}")
        
        used_names.add(name.lower())
        speaker_names[speaker] = name
        speaker_context[speaker] = context
    
    return speaker_names, speaker_context


# ============================================================================
# Speaker Map Management
# ============================================================================

def load_speaker_map(path: Path, logger: logging.Logger) -> Dict[str, str]:
    """Load speaker mapping from JSON cache."""
    if not path.exists():
        return {}
    
    try:
        with open(path, encoding="utf-8") as f:
            names = json.load(f)
        
        # Проверяем что это не пустой шаблон
        if all(k == v for k, v in names.items()):
            logger.info(f"Speaker map is empty template: {path}")
            return {}
        
        logger.info(f"Loaded speaker map from cache: {path}")
        return names
    except Exception as e:
        logger.warning(f"Could not load speaker map: {e}")
        return {}


def apply_speaker_names(segments: List[dict], names: Dict[str, str]) -> List[dict]:
    """Replace SPEAKER_XX with actual names."""
    result = []
    for seg in segments:
        new_seg = seg.copy()
        speaker = seg.get("speaker", "UNKNOWN")
        
        if "original_speaker_id" not in new_seg:
            new_seg["original_speaker_id"] = speaker
        
        if speaker in names and names[speaker] != speaker:
            new_seg["speaker"] = names[speaker]
        
        result.append(new_seg)
    return result


# ============================================================================
# Output Writers
# ============================================================================

def save_named_json(segments: List[dict], speakers: Set[str], project_name: str,
                    output_path: Path, metadata: dict, speaker_map: Dict[str, str]):
    """Save JSON with named speakers."""
    data = {
        "project_name": project_name,
        "generated": datetime.now().isoformat(),
        "language": metadata.get("language"),
        "num_speakers": len(speakers),
        "speakers": sorted(list(speakers)),
        "speaker_map": speaker_map,
        "total_segments": len(segments),
        "segments": segments
    }
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def save_named_txt(segments: List[dict], speakers: Set[str], project_name: str,
                   output_path: Path, metadata: dict):
    """Save human-readable transcript."""
    lines = [
        "=" * 60,
        "PROJECT TRANSCRIPTION (Named Speakers)",
        "=" * 60,
        "",
        f"Project: {project_name}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Speakers ({len(speakers)}): {', '.join(sorted(speakers))}",
        "",
        "-" * 60,
        ""
    ]
    
    for seg in segments:
        ts = format_timestamp(seg["start"])
        speaker = seg.get("speaker", "UNKNOWN")
        text = seg.get("text", "")
        lines.append(f"[{ts}] {speaker}:")
        lines.append(f"  {text}")
        lines.append("")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def save_full_srt(segments: List[dict], output_path: Path) -> int:
    """Save full SRT file with global timecodes."""
    lines = []
    idx = 1
    
    for seg in segments:
        text = seg.get("text", "").strip()
        if not text:
            continue
        
        speaker = seg.get("speaker", "UNKNOWN")
        
        lines.append(str(idx))
        lines.append(f"{format_srt_timestamp(seg['start'])} --> {format_srt_timestamp(seg['end'])}")
        lines.append(f"[{speaker}] {text}")
        lines.append("")
        idx += 1
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    
    return idx - 1


def save_clips_xlsx(segments: List[dict], clips: List[dict], output_path: Path, logger: logging.Logger):
    """Save XLSX table with segments split by clips."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript"
    
    headers = ["file", "timecode", "speaker", "text"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    row_num = 2
    mapped = 0
    
    for seg in segments:
        clip, local_ts = find_clip_for_timestamp(clips, seg["start"])
        if clip is None:
            continue
        
        ws.cell(row=row_num, column=1, value=clip["file"])
        ws.cell(row=row_num, column=2, value=format_timestamp(local_ts))
        ws.cell(row=row_num, column=3, value=seg.get("speaker", "UNKNOWN"))
        ws.cell(row=row_num, column=4, value=seg.get("text", ""))
        
        row_num += 1
        mapped += 1
    
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 80
    
    wb.save(output_path)
    logger.info(f"Mapped {mapped}/{len(segments)} segments to clips")


def save_clip_srt(segments: List[dict], clip: dict, output_path: Path) -> int:
    """Save SRT file for a single clip with local timecodes."""
    lines = []
    idx = 1
    
    for seg in segments:
        if not (clip["start"] <= seg["start"] < clip["end"]):
            continue
        
        local_start = seg["start"] - clip["start"]
        local_end = min(seg["end"] - clip["start"], clip["duration"])
        
        text = seg.get("text", "").strip()
        if not text:
            continue
        
        speaker = seg.get("speaker", "UNKNOWN")
        
        lines.append(str(idx))
        lines.append(f"{format_srt_timestamp(local_start)} --> {format_srt_timestamp(local_end)}")
        lines.append(f"[{speaker}] {text}")
        lines.append("")
        idx += 1
    
    if lines:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    
    return idx - 1


def save_speakers_txt(speaker_map: Dict[str, str], speaker_context: Dict[str, str],
                      segments: List[dict], output_path: Path):
    """Save speakers verification file with sample utterances."""
    
    # Собираем данные
    speaker_data = {}
    for seg in segments:
        original_id = seg.get("original_speaker_id", seg.get("speaker", "UNKNOWN"))
        current_name = seg.get("speaker", "UNKNOWN")
        text = seg.get("text", "").strip()
        
        if original_id not in speaker_data:
            speaker_data[original_id] = {
                "name": current_name,
                "count": 0,
                "utterances": []
            }
        
        speaker_data[original_id]["count"] += 1
        if text and len(speaker_data[original_id]["utterances"]) < 10:
            speaker_data[original_id]["utterances"].append({
                "time": seg.get("start", 0),
                "text": text[:200]
            })
    
    lines = [
        "=" * 70,
        "SPEAKERS VERIFICATION FILE",
        "=" * 70,
        "",
        "Review sample utterances to verify speaker identification.",
        "If wrong, edit speaker_names.json and re-run with --skip-llm",
        "",
        "=" * 70,
        "",
    ]
    
    sorted_speakers = sorted(speaker_data.keys(), 
                            key=lambda x: speaker_data[x]["count"], 
                            reverse=True)
    
    for original_id in sorted_speakers:
        data = speaker_data[original_id]
        name = speaker_map.get(original_id, data["name"])
        context = speaker_context.get(original_id, "")
        
        lines.append(f"{'─' * 70}")
        lines.append(f"ID: {original_id}")
        lines.append(f"ASSIGNED NAME: {name}")
        lines.append(f"Segments: {data['count']}")
        if context:
            lines.append(f"Context: {context}")
        lines.append("")
        lines.append("SAMPLE UTTERANCES:")
        
        for utt in data["utterances"][:5]:
            time_str = format_timestamp(utt["time"])
            lines.append(f"  [{time_str}] \"{utt['text']}\"")
        
        if data["count"] > 5:
            lines.append(f"  ... and {data['count'] - 5} more")
        lines.append("")
    
    lines.extend([
        "=" * 70,
        "TO FIX WRONG NAMES:",
        "=" * 70,
        "",
        "1. Edit speaker_names.json:",
        '   {"SPEAKER_05": "Hadi", "SPEAKER_02": "Roman"}',
        "",
        "2. Re-run: python process_transcript.py --project ... --skip-llm",
        "",
    ])
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="YTAI Step 2: Process Transcript (Individual Speaker Analysis)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python process_transcript.py --project "/Volumes/RYA Blue/PROJECT"
    python process_transcript.py --project "/Volumes/RYA Blue/PROJECT" --skip-llm
    python process_transcript.py --project "/Volumes/RYA Blue/PROJECT" --force
        """
    )
    
    parser.add_argument("--project", required=True, help="Project folder path")
    parser.add_argument("--json", type=Path, help="Specific transcript JSON (default: latest)")
    parser.add_argument("--skip-llm", action="store_true", 
                        help="Skip LLM, use cached speaker_names.json")
    parser.add_argument("--force", action="store_true",
                        help="Force re-analyze (ignore cache)")
    parser.add_argument("--llm-model", default=DEFAULT_LLM_MODEL,
                        help=f"Ollama model (default: {DEFAULT_LLM_MODEL})")
    
    args = parser.parse_args()
    
    # Setup
    try:
        paths = get_project_paths(args.project)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    
    paths["transcripts_runs_dir"].mkdir(parents=True, exist_ok=True)
    paths["transcripts_clean_dir"].mkdir(parents=True, exist_ok=True)
    paths["logs_dir"].mkdir(parents=True, exist_ok=True)
    
    logger = setup_logging(paths["logs_dir"], paths["project_name"])
    
    # Find transcript
    json_path = args.json or find_latest_transcript(paths["transcripts_runs_dir"], paths["project_name"])
    
    if not json_path or not json_path.exists():
        logger.error("No transcript JSON found. Run transcribe_project.py first.")
        sys.exit(1)
    
    # Load transcript
    logger.info("=" * 60)
    logger.info("YTAI STEP 2: PROCESS TRANSCRIPT")
    logger.info("=" * 60)
    logger.info(f"Project: {paths['project_name']}")
    logger.info(f"Input: {json_path}")
    logger.info(f"LLM Model: {args.llm_model}")
    logger.info("")
    
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    
    segments = data.get("segments", [])
    if not segments:
        logger.error("No segments in transcript")
        sys.exit(1)
    
    speakers = set(data.get("speakers", []))
    if not speakers:
        speakers = set(seg.get("speaker", "UNKNOWN") for seg in segments)
        speakers.discard("UNKNOWN")
    
    metadata = {"language": data.get("language"), "model": data.get("model")}
    
    logger.info(f"Loaded: {len(segments)} segments, {len(speakers)} speakers")
    logger.info(f"Speakers: {', '.join(sorted(speakers))}")
    logger.info("")
    
    # Phase 1: Speaker Naming
    logger.info("PHASE 1: Speaker Identification")
    logger.info("-" * 40)
    
    speaker_map = {}
    speaker_context = {}
    debug_dir = paths["transcripts_runs_dir"] / "speaker_analysis"
    
    if args.skip_llm:
        speaker_map = load_speaker_map(paths["speaker_map_file"], logger)
        if not speaker_map:
            logger.warning("No cached speaker_names.json found, using original IDs")
    elif args.force:
        logger.info("Force mode - re-analyzing all speakers")
        speaker_map, speaker_context = identify_speakers_individual(
            segments, speakers, args.llm_model, logger, debug_dir
        )
    else:
        # Пробуем кеш, иначе анализируем
        speaker_map = load_speaker_map(paths["speaker_map_file"], logger)
        if not speaker_map:
            logger.info("No cache, analyzing speakers individually...")
            speaker_map, speaker_context = identify_speakers_individual(
                segments, speakers, args.llm_model, logger, debug_dir
            )
    
    # Apply names
    if speaker_map:
        segments = apply_speaker_names(segments, speaker_map)
        
        # Save cache
        with open(paths["speaker_map_file"], "w", encoding="utf-8") as f:
            json.dump(speaker_map, f, indent=2, ensure_ascii=False)
        logger.info(f"Saved: {paths['speaker_map_file']}")
    
    final_speakers = set(seg.get("speaker", "UNKNOWN") for seg in segments)
    final_speakers.discard("UNKNOWN")
    logger.info(f"Final speakers: {', '.join(sorted(final_speakers))}")
    logger.info("")
    
    # Phase 2: Video Clips
    logger.info("PHASE 2: Analyzing Video Clips")
    logger.info("-" * 40)
    
    if not paths["video_dir"].exists():
        logger.error(f"Video directory not found: {paths['video_dir']}")
        sys.exit(1)
    
    clips = get_video_clips(paths["video_dir"], logger)
    if not clips:
        logger.error("No video clips found!")
        sys.exit(1)
    
    logger.info("")
    
    # Phase 3: Save Outputs
    logger.info("PHASE 3: Saving Results")
    logger.info("-" * 40)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = paths["transcripts_runs_dir"] / f"{paths['project_name']}_process_{timestamp}"
    
    # JSON
    json_out = base.with_name(base.name + "_named.json")
    save_named_json(segments, final_speakers, paths["project_name"], json_out, metadata, speaker_map)
    logger.info(f"JSON: {json_out}")
    
    # TXT
    txt_out = base.with_name(base.name + "_named.txt")
    save_named_txt(segments, final_speakers, paths["project_name"], txt_out, metadata)
    logger.info(f"TXT: {txt_out}")
    
    # XLSX
    xlsx_out = base.with_name(base.name + "_by_clips.xlsx")
    save_clips_xlsx(segments, clips, xlsx_out, logger)
    logger.info(f"XLSX: {xlsx_out}")
    
    # Full SRT
    srt_out = base.with_name(base.name + "_named.srt")
    srt_count = save_full_srt(segments, srt_out)
    logger.info(f"SRT: {srt_out} ({srt_count} subtitles)")
    
    # Speakers verification
    speakers_out = base.with_name(base.name + "_speakers.txt")
    save_speakers_txt(speaker_map, speaker_context, segments, speakers_out)
    logger.info(f"Speakers: {speakers_out}")
    
    # Per-clip SRTs
    logger.info("")
    logger.info("Generating per-clip SRT files...")
    
    clip_count = 0
    for clip in clips:
        srt_name = Path(clip["file"]).stem + ".srt"
        srt_path = paths["transcripts_clean_dir"] / srt_name
        
        num_subs = save_clip_srt(segments, clip, srt_path)
        if num_subs > 0:
            logger.info(f"  {srt_name}: {num_subs} subtitles")
            clip_count += 1
    
    logger.info(f"Generated {clip_count} SRT files")
    
    # Summary
    logger.info("")
    logger.info("=" * 60)
    logger.info("COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Speakers: {', '.join(sorted(final_speakers))}")
    logger.info(f"Segments: {len(segments)}")
    logger.info(f"Clips: {len(clips)}")


if __name__ == "__main__":
    main()
