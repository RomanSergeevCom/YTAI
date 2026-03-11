#!/usr/bin/env python3
"""
YTAI Step 2: Process Transcript
Speaker Naming + Split by Video Clips

This script takes the JSON output from Step 1 (transcribe_project.py) and:
1. Names speakers via Ollama LLM (llama3.3:70b)
2. Splits transcript by video clips using ffprobe durations
3. Generates per-clip SRT files for Premiere Pro

Usage:
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani"
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani" --skip-llm

Output:
    02_Transcripts/02_01_Runs/
    â”œâ”€â”€ *_named.json (with speaker names)
    â”œâ”€â”€ *_named.txt
    â””â”€â”€ *_by_clips.xlsx (table: file | timecode | speaker | text)

    02_Transcripts/02_02_Clean/
    â”œâ”€â”€ RYA-ZVE1-1146.srt
    â”œâ”€â”€ RYA-ZVE1-1147.srt
    â””â”€â”€ ...

Requirements:
    pip install requests openpyxl
    
    # For LLM naming:
    export OLLAMA_MODELS=~/YTAI/models/ollama
    OLLAMA_MAX_VRAM=20g ollama serve
    
    # ffprobe (part of ffmpeg):
    brew install ffmpeg
"""

import argparse
import os
import sys
import json
import re
import subprocess
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Set, Tuple

try:
    import requests
except ImportError:
    print("ERROR: requests not installed. Run: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# Configuration
# ============================================================================

# Project structure
VIDEO_DIR = "01_Raw/01_01_Video"
AUDIO_DIR = "01_Raw/01_02_Audio"
TRANSCRIPTS_RUNS_DIR = "02_Transcripts/02_01_Runs"
TRANSCRIPTS_CLEAN_DIR = "02_Transcripts/02_02_Clean"
LOGS_DIR = "08_Logs"

VIDEO_EXTS = {".mp4", ".mov", ".m4v", ".mts", ".avi", ".mkv",
              ".MP4", ".MOV", ".M4V", ".MTS", ".AVI", ".MKV"}

# Ollama settings
OLLAMA_API_URL = "http://localhost:11434/api/generate"
DEFAULT_LLM_MODEL = "qwen2.5:32b"
OLLAMA_TIMEOUT = 300  # 5 minutes (достаточно для 32B модели)


# ============================================================================
# Project Paths
# ============================================================================

def get_project_paths(project_dir: str) -> dict:
    """Get all relevant paths for a project."""
    project_root = Path(project_dir).expanduser().resolve()
    
    if not project_root.exists():
        raise FileNotFoundError(f"Project folder not found: {project_root}")
    
    return {
        "project_root": project_root,
        "project_name": project_root.name,
        "video_dir": project_root / VIDEO_DIR,
        "audio_dir": project_root / AUDIO_DIR,
        "transcripts_runs_dir": project_root / TRANSCRIPTS_RUNS_DIR,
        "transcripts_clean_dir": project_root / TRANSCRIPTS_CLEAN_DIR,
        "logs_dir": project_root / LOGS_DIR,
        "speaker_map_file": project_root / "speaker_names.json"
    }


def find_latest_transcript(transcripts_dir: Path, project_name: str) -> Optional[Path]:
    """Find the most recent transcript JSON file."""
    pattern = f"{project_name}_transcript_*.json"
    files = list(transcripts_dir.glob(pattern))
    
    # Exclude already processed files (*_named.json)
    files = [f for f in files if "_named" not in f.name]
    
    if not files:
        return None
    
    # Sort by modification time, newest first
    files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return files[0]


# ============================================================================
# Logging
# ============================================================================

def setup_logging(logs_dir: Path, project_name: str) -> logging.Logger:
    """Setup dual logging."""
    logs_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = logs_dir / f"{project_name}_process_{timestamp}.log"
    
    logger = logging.getLogger("process")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info(f"Log: {log_file}")
    
    return logger


# ============================================================================
# Timestamp Formatting
# ============================================================================

def format_timestamp(seconds: float) -> str:
    """Format as HH:MM:SS"""
    h, r = divmod(int(seconds), 3600)
    m, s = divmod(r, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def format_srt_timestamp(seconds: float) -> str:
    """Format for SRT: HH:MM:SS,mmm"""
    h, r = divmod(int(seconds), 3600)
    m, s = divmod(r, 60)
    ms = int((seconds - int(seconds)) * 1000)
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


# ============================================================================
# Video Clip Utilities
# ============================================================================

def natural_sort_key(s: str):
    """Sort strings with embedded numbers naturally."""
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]


def get_video_duration(video_path: Path) -> float:
    """Get video duration in seconds using ffprobe."""
    cmd = [
        "ffprobe",
        "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        str(video_path)
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return float(result.stdout.strip())
    except Exception:
        pass
    
    return 0.0


def get_video_clips(video_dir: Path, logger: logging.Logger) -> List[dict]:
    """
    Get list of video clips with their durations and offsets.
    
    Returns list of dicts:
    [
        {"file": "RYA-ZVE1-1146.MP4", "path": Path, "duration": 150.5, "start": 0.0, "end": 150.5},
        {"file": "RYA-ZVE1-1147.MP4", "path": Path, "duration": 120.0, "start": 150.5, "end": 270.5},
        ...
    ]
    """
    # Find all video files (exclude Archive folder)
    clips = []
    for ext in VIDEO_EXTS:
        for f in video_dir.glob(f"*{ext}"):
            # Skip files in Archive folder
            if "Archive" in str(f):
                continue
            if not f.name.startswith("."):
                clips.append(f)
    
    # Sort naturally
    clips.sort(key=lambda p: natural_sort_key(p.name))
    
    if not clips:
        logger.warning(f"No video clips found in {video_dir}")
        return []
    
    logger.info(f"Found {len(clips)} video clips")
    
    # Get durations and calculate offsets
    result = []
    current_offset = 0.0
    
    for clip_path in clips:
        duration = get_video_duration(clip_path)
        
        if duration <= 0:
            logger.warning(f"Could not get duration for {clip_path.name}, skipping")
            continue
        
        result.append({
            "file": clip_path.name,
            "path": clip_path,
            "duration": duration,
            "start": current_offset,
            "end": current_offset + duration
        })
        
        logger.debug(f"  {clip_path.name}: {format_timestamp(duration)} ({current_offset:.1f}s - {current_offset + duration:.1f}s)")
        current_offset += duration
    
    logger.info(f"Total duration: {format_timestamp(current_offset)}")
    return result


def find_clip_for_timestamp(clips: List[dict], timestamp: float) -> Tuple[Optional[dict], float]:
    """
    Find which clip contains the given timestamp.
    Returns (clip_info, local_timestamp) or (None, 0) if not found.
    """
    for clip in clips:
        if clip["start"] <= timestamp < clip["end"]:
            local_ts = timestamp - clip["start"]
            return clip, local_ts
    
    # Edge case: timestamp at very end
    if clips and timestamp >= clips[-1]["end"] - 0.1:
        clip = clips[-1]
        local_ts = timestamp - clip["start"]
        return clip, local_ts
    
    return None, 0.0


# ============================================================================
# Ollama LLM for Speaker Naming
# ============================================================================

def check_ollama_server(logger: logging.Logger) -> bool:
    """Check if Ollama server is running."""
    try:
        response = requests.get("http://localhost:11434/api/tags", timeout=5)
        if response.status_code == 200:
            models = response.json().get("models", [])
            model_names = [m.get("name", "") for m in models]
            logger.debug(f"Available models: {model_names}")
            return True
        return False
    except requests.exceptions.ConnectionError:
        return False
    except Exception as e:
        logger.debug(f"Ollama check error: {e}")
        return False


def call_ollama_api(prompt: str, model: str, logger: logging.Logger) -> Optional[str]:
    """Call Ollama API with specified model."""
    try:
        logger.debug(f"Calling Ollama API with {model}...")
        
        response = requests.post(
            OLLAMA_API_URL,
            json={
                "model": model,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.3,
                    "num_predict": 500,
                }
            },
            timeout=OLLAMA_TIMEOUT
        )
        
        if response.status_code == 200:
            result = response.json()
            return result.get("response", "")
        else:
            logger.warning(f"Ollama API error: {response.status_code}")
            return None
            
    except requests.exceptions.Timeout:
        logger.warning(f"Ollama timeout ({OLLAMA_TIMEOUT}s)")
        return None
    except requests.exceptions.ConnectionError:
        logger.warning("Ollama server not running")
        return None
    except Exception as e:
        logger.warning(f"Ollama API error: {e}")
        return None


def extract_speaker_names_llm(segments: List[dict], speakers: Set[str], model: str, logger: logging.Logger) -> Dict[str, str]:
    """Use Ollama to identify speaker names/roles from context."""
    if not check_ollama_server(logger):
        logger.warning("Ollama server not running!")
        logger.warning("Start with: ollama serve")
        return {}
    
    # Build transcript samples for each speaker
    # Take more lines and longer excerpts for better context
    speaker_lines = {s: [] for s in speakers}
    
    for seg in segments:
        speaker = seg.get("speaker", "UNKNOWN")
        text = seg.get("text", "").strip()
        # Skip very short lines (like "Okay", "Yes", "Hmm")
        if text and speaker in speakers and len(text) > 20:
            speaker_lines[speaker].append(text)
    
    # Build sample: up to 10 lines per speaker, 200 chars each
    # Mix of first lines (intro) and middle lines (substance)
    transcript_parts = []
    for speaker in sorted(speakers):
        all_lines = speaker_lines.get(speaker, [])
        
        if not all_lines:
            # Fallback: include even short lines if nothing else
            all_lines = [seg.get("text", "") for seg in segments 
                        if seg.get("speaker") == speaker and seg.get("text", "").strip()][:5]
        
        # Take first 5 + middle 5 for variety
        sample_lines = all_lines[:5]
        if len(all_lines) > 10:
            mid = len(all_lines) // 2
            sample_lines += all_lines[mid:mid+5]
        elif len(all_lines) > 5:
            sample_lines += all_lines[5:10]
        
        if sample_lines:
            transcript_parts.append(f"\n{speaker} says:")
            for line in sample_lines:
                transcript_parts.append(f'  "{line[:200]}"')
    
    transcript = "\n".join(transcript_parts)
    speakers_list = ", ".join(sorted(speakers))
    
    prompt = f"""Analyze this interview transcript and identify WHO each speaker is.

SPEAKERS TO IDENTIFY: {speakers_list}

TRANSCRIPT SAMPLES:
{transcript}

RULES:
1. Look for self-introductions: "My name is...", "I am...", "I'm..."
2. Look for how others address them: "Thank you Ahmed", "As Roman mentioned"
3. If a name is found, verify it sounds like a real name (not a transcription error)
4. If NO name is found, assign a DESCRIPTIVE ROLE based on what they talk about:
   - "Interviewer" or "Host" - person asking questions
   - "Guest" - person being interviewed
   - "Coffee Shop Owner" - if they talk about their coffee business
   - "Business Owner" - if they talk about their business
   - "Expert" - subject matter expert
   - Use specific roles when possible (e.g. "Restaurant Owner", "Lawyer", "Accountant")
5. If someone speaks Arabic/another language, still identify their role from context
6. Each speaker MUST have a UNIQUE label - no duplicates allowed

CRITICAL: 
- Provide a mapping for EVERY speaker: {speakers_list}
- Each label must be UNIQUE (not "Guest" for two different people)
- If multiple guests, use "Guest 1", "Guest 2" or descriptive roles
- NEVER leave any speaker as SPEAKER_XX

OUTPUT FORMAT - JSON only, no explanation:
{{"SPEAKER_00": "Name or Role", "SPEAKER_01": "Name or Role"}}

Provide UNIQUE mapping for ALL {len(speakers)} speakers."""

    logger.info(f"Asking {model} to identify {len(speakers)} speakers...")
    
    response = call_ollama_api(prompt, model, logger)
    
    if not response:
        return {}
    
    logger.debug(f"LLM response: {response[:500]}")
    
    # Extract JSON
    try:
        json_match = re.search(r'\{[^{}]+\}', response, re.DOTALL)
        if json_match:
            raw_names = json.loads(json_match.group())
            names = validate_speaker_names(raw_names, speakers, logger)
            logger.info(f"Speaker names: {names}")
            return names
        
        logger.warning("Could not parse LLM response as JSON")
        return {}
        
    except json.JSONDecodeError as e:
        logger.warning(f"Invalid JSON: {e}")
        return {}


def validate_speaker_names(raw_names: Dict[str, str], speakers: Set[str], logger: logging.Logger) -> Dict[str, str]:
    """Validate and clean speaker names from LLM."""
    validated = {}
    used_names = set()  # Track used names to ensure uniqueness
    
    invalid_patterns = [
        r'^unknown$',
        r'^speaker[\s_]*\d*$',  # Reject SPEAKER_XX format
        r'^person\s*\d*$',
        r'^\?+$',
        r'^n/?a$',
        r'^not\s+(specified|identified|known)$',
        r'^unidentified$',
    ]
    
    for speaker in speakers:
        name = raw_names.get(speaker, "").strip()
        
        is_valid = True
        
        if len(name) < 2:
            is_valid = False
        
        if len(name) > 30:  # Allow slightly longer descriptive roles
            is_valid = False
            logger.debug(f"Name too long for {speaker}: '{name}'")
        
        for pattern in invalid_patterns:
            if re.match(pattern, name.lower().replace("_", " ")):
                is_valid = False
                logger.debug(f"Invalid pattern for {speaker}: '{name}'")
                break
        
        if is_valid and not re.match(r'^[\w\s\-\'\.]+$', name, re.UNICODE):
            is_valid = False
            logger.debug(f"Invalid chars in name for {speaker}: '{name}'")
        
        # Check for uniqueness
        if is_valid and name.lower() in used_names:
            # Make unique by adding number
            base_name = name
            counter = 2
            while f"{base_name} {counter}".lower() in used_names:
                counter += 1
            name = f"{base_name} {counter}"
            logger.debug(f"Made unique: {speaker} -> '{name}'")
        
        if is_valid:
            validated[speaker] = name.title().strip()
            used_names.add(name.lower())
        else:
            # Generate descriptive fallback instead of keeping SPEAKER_XX
            fallback = f"Speaker {len(validated) + 1}"
            validated[speaker] = fallback
            used_names.add(fallback.lower())
            logger.debug(f"Fallback for {speaker}: '{fallback}'")
    
    return validated


def load_speaker_map(path: Path, logger: logging.Logger) -> Dict[str, str]:
    """Load manual speaker mapping from JSON."""
    if not path.exists():
        return {}
    
    try:
        with open(path) as f:
            names = json.load(f)
        
        # Check if it's just the template (values == keys)
        if all(k == v for k, v in names.items()):
            logger.info(f"Speaker map is template (not edited): {path}")
            return {}
        
        logger.info(f"Loaded speaker map from {path}")
        return names
    except Exception as e:
        logger.warning(f"Could not load speaker map: {e}")
        return {}


def apply_speaker_names(segments: List[dict], names: Dict[str, str]) -> List[dict]:
    """Replace SPEAKER_XX with actual names."""
    result = []
    for seg in segments:
        new_seg = seg.copy()
        speaker = seg.get("speaker", "UNKNOWN")
        if speaker in names and names[speaker] != speaker:
            new_seg["speaker_id"] = speaker
            new_seg["speaker"] = names[speaker]
        result.append(new_seg)
    return result


# ============================================================================
# Output Writers
# ============================================================================

def save_named_txt(segments: List[dict], speakers: Set[str], 
                   project_name: str, output_path: Path, metadata: dict):
    """Save human-readable transcript with named speakers."""
    lines = [
        "=" * 60,
        "PROJECT TRANSCRIPTION (Named Speakers)",
        "=" * 60,
        "",
        f"Project: {project_name}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Language: {metadata.get('language', 'unknown')}",
        f"Speakers ({len(speakers)}): {', '.join(sorted(speakers))}",
        "",
        "-" * 60,
        "TRANSCRIPT",
        "-" * 60,
        ""
    ]
    
    for seg in segments:
        ts = format_timestamp(seg["start"])
        speaker = seg.get("speaker", "UNKNOWN")
        text = seg.get("text", "")
        lines.append(f"[{ts}] {speaker}:")
        lines.append(f"  {text}")
        lines.append("")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def save_named_json(segments: List[dict], speakers: Set[str],
                    project_name: str, output_path: Path, 
                    metadata: dict, speaker_map: Dict[str, str]):
    """Save JSON with named speakers."""
    data = {
        "project_name": project_name,
        "generated": datetime.now().isoformat(),
        "language": metadata.get("language"),
        "num_speakers": len(speakers),
        "speakers": sorted(list(speakers)),
        "speaker_map": speaker_map,
        "total_segments": len(segments),
        "segments": segments
    }
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def save_clips_xlsx(segments: List[dict], clips: List[dict], output_path: Path, logger: logging.Logger):
    """
    Save XLSX table with segments split by clips.
    
    Columns: file | timecode | speaker | text
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript"
    
    # Header
    headers = ["file", "timecode", "speaker", "text"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
    
    # Data rows
    row_num = 2
    segments_mapped = 0
    
    for seg in segments:
        global_ts = seg["start"]
        clip, local_ts = find_clip_for_timestamp(clips, global_ts)
        
        if clip is None:
            logger.debug(f"No clip found for timestamp {global_ts:.1f}s")
            continue
        
        ws.cell(row=row_num, column=1, value=clip["file"])
        ws.cell(row=row_num, column=2, value=format_timestamp(local_ts))
        ws.cell(row=row_num, column=3, value=seg.get("speaker", "UNKNOWN"))
        ws.cell(row=row_num, column=4, value=seg.get("text", ""))
        
        row_num += 1
        segments_mapped += 1
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 80
    
    wb.save(output_path)
    logger.info(f"Mapped {segments_mapped}/{len(segments)} segments to clips")


def save_clip_srt(segments: List[dict], clip: dict, output_path: Path) -> int:
    """
    Save SRT file for a single clip with local timecodes.
    Returns number of subtitles written.
    """
    lines = []
    idx = 1
    
    for seg in segments:
        global_ts = seg["start"]
        
        # Check if segment belongs to this clip
        if not (clip["start"] <= global_ts < clip["end"]):
            continue
        
        # Calculate local timestamps
        local_start = global_ts - clip["start"]
        local_end = seg["end"] - clip["start"]
        
        # Clamp end to clip duration
        local_end = min(local_end, clip["duration"])
        
        text = seg.get("text", "").strip()
        if not text:
            continue
        
        speaker = seg.get("speaker", "UNKNOWN")
        
        lines.append(str(idx))
        lines.append(f"{format_srt_timestamp(local_start)} --> {format_srt_timestamp(local_end)}")
        lines.append(f"[{speaker}] {text}")
        lines.append("")
        
        idx += 1
    
    if lines:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    
    return idx - 1


def save_full_srt(segments: List[dict], output_path: Path) -> int:
    """
    Save SRT file for full audio with global timecodes (named speakers).
    Returns number of subtitles written.
    """
    lines = []
    idx = 1
    
    for seg in segments:
        text = seg.get("text", "").strip()
        if not text:
            continue
        
        start = seg["start"]
        end = seg["end"]
        speaker = seg.get("speaker", "UNKNOWN")
        
        lines.append(str(idx))
        lines.append(f"{format_srt_timestamp(start)} --> {format_srt_timestamp(end)}")
        lines.append(f"[{speaker}] {text}")
        lines.append("")
        
        idx += 1
    
    if lines:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    
    return idx - 1


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="YTAI Step 2: Process Transcript (naming + split by clips)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani"
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani" --skip-llm
        """
    )
    
    parser.add_argument("--project", required=True,
                        help="Project folder path")
    parser.add_argument("--json", type=Path,
                        help="Specific transcript JSON (default: latest)")
    parser.add_argument("--skip-llm", action="store_true",
                        help="Skip LLM speaker naming (use manual map or IDs)")
    parser.add_argument("--llm-model", default=DEFAULT_LLM_MODEL,
                        help=f"Ollama model for speaker naming (default: {DEFAULT_LLM_MODEL})")
    
    args = parser.parse_args()
    
    # Get project paths
    try:
        paths = get_project_paths(args.project)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    
    # Setup directories
    paths["transcripts_runs_dir"].mkdir(parents=True, exist_ok=True)
    paths["transcripts_clean_dir"].mkdir(parents=True, exist_ok=True)
    paths["logs_dir"].mkdir(parents=True, exist_ok=True)
    
    # Setup logging
    logger = setup_logging(paths["logs_dir"], paths["project_name"])
    
    # Find transcript JSON
    if args.json:
        json_path = args.json
    else:
        json_path = find_latest_transcript(paths["transcripts_runs_dir"], paths["project_name"])
    
    if not json_path or not json_path.exists():
        logger.error(f"No transcript JSON found. Run transcribe_project.py first.")
        sys.exit(1)
    
    # Load transcript
    logger.info("=" * 60)
    logger.info("YTAI STEP 2: PROCESS TRANSCRIPT")
    logger.info("=" * 60)
    logger.info(f"Project: {paths['project_name']}")
    logger.info(f"Input: {json_path}")
    if not args.skip_llm:
        logger.info(f"LLM Model: {args.llm_model}")
    logger.info("")
    
    with open(json_path) as f:
        data = json.load(f)
    
    segments = data.get("segments", [])
    original_speakers = set(data.get("speakers", []))
    metadata = {
        "language": data.get("language"),
        "model": data.get("model")
    }
    
    logger.info(f"Loaded: {len(segments)} segments, {len(original_speakers)} speakers")
    logger.info(f"Speakers: {', '.join(sorted(original_speakers))}")
    logger.info("")
    
    # Phase 1: Speaker naming
    logger.info("PHASE 1: Speaker Naming")
    logger.info("-" * 40)
    
    speaker_map = {}
    
    # Try manual map first
    speaker_map = load_speaker_map(paths["speaker_map_file"], logger)
    
    # Try LLM if needed
    if not args.skip_llm and not speaker_map:
        llm_names = extract_speaker_names_llm(segments, original_speakers, args.llm_model, logger)
        speaker_map = llm_names
    
    # Apply names
    if speaker_map:
        segments = apply_speaker_names(segments, speaker_map)
        
        # Update speaker_names.json with LLM results
        with open(paths["speaker_map_file"], "w") as f:
            json.dump(speaker_map, f, indent=2)
        logger.info(f"Updated: {paths['speaker_map_file']}")
    else:
        logger.info("No speaker naming applied (using original IDs)")
    
    # Get final speakers
    final_speakers = set(seg.get("speaker", "UNKNOWN") for seg in segments)
    final_speakers.discard("UNKNOWN")
    logger.info(f"Final speakers: {', '.join(sorted(final_speakers))}")
    logger.info("")
    
    # Phase 2: Get video clips
    logger.info("PHASE 2: Analyzing Video Clips")
    logger.info("-" * 40)
    
    clips = get_video_clips(paths["video_dir"], logger)
    
    if not clips:
        logger.error("No video clips found!")
        sys.exit(1)
    
    logger.info("")
    
    # Phase 3: Save outputs
    logger.info("PHASE 3: Saving Results")
    logger.info("-" * 40)
    
    # Generate output base name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_base = paths["transcripts_runs_dir"] / f"{paths['project_name']}_transcript_{timestamp}"
    
    # Save named JSON
    named_json = output_base.with_name(output_base.name + "_named.json")
    save_named_json(segments, final_speakers, paths["project_name"], 
                    named_json, metadata, speaker_map)
    logger.info(f"JSON: {named_json}")
    
    # Save named TXT
    named_txt = output_base.with_name(output_base.name + "_named.txt")
    save_named_txt(segments, final_speakers, paths["project_name"], named_txt, metadata)
    logger.info(f"TXT: {named_txt}")
    
    # Save XLSX by clips
    xlsx_path = output_base.with_name(output_base.name + "_by_clips.xlsx")
    save_clips_xlsx(segments, clips, xlsx_path, logger)
    logger.info(f"XLSX: {xlsx_path}")
    
    # Save full SRT (named speakers, global timecodes)
    full_srt_path = output_base.with_name(output_base.name + "_named.srt")
    full_srt_count = save_full_srt(segments, full_srt_path)
    logger.info(f"SRT (full): {full_srt_path} ({full_srt_count} subtitles)")
    
    # Save per-clip SRT files
    logger.info("")
    logger.info("Generating per-clip SRT files...")
    
    srt_count = 0
    for clip in clips:
        # SRT filename: same as video but .srt
        srt_name = Path(clip["file"]).stem + ".srt"
        srt_path = paths["transcripts_clean_dir"] / srt_name
        
        num_subs = save_clip_srt(segments, clip, srt_path)
        
        if num_subs > 0:
            logger.info(f"  {srt_name}: {num_subs} subtitles")
            srt_count += 1
    
    logger.info(f"Generated {srt_count} SRT files in {paths['transcripts_clean_dir']}")
    
    # Summary
    logger.info("")
    logger.info("=" * 60)
    logger.info("STEP 2 COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Speakers: {', '.join(sorted(final_speakers))}")
    logger.info(f"Segments: {len(segments)}")
    logger.info(f"Clips: {len(clips)}")
    logger.info(f"SRT files: {srt_count}")
    logger.info("")
    logger.info("Output files:")
    logger.info(f"  {named_json}")
    logger.info(f"  {named_txt}")
    logger.info(f"  {full_srt_path}")
    logger.info(f"  {xlsx_path}")
    logger.info(f"  {paths['transcripts_clean_dir']}/*.srt ({srt_count} clips)")


if __name__ == "__main__":
    main()
