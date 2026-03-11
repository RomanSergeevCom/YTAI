#!/usr/bin/env python3
"""
YTAI Step 2: Process Transcript
Speaker Naming + Split by Video Clips

This script takes the JSON output from Step 1 (transcribe_project.py) and:
1. Analyzes and optionally merges speakers (dialogue detection)
2. Names speakers via Ollama LLM (default: qwen2.5:32b)
3. Splits transcript by video clips using ffprobe durations
4. Generates per-clip SRT files for Premiere Pro

Usage:
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani"
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani" --skip-llm
    python process_transcript.py --project "/Volumes/RYA Blue/YTCG37_Hadi_Dawani" --no-merge

Output:
    02_Transcripts/02_01_Runs/
        <project>_process_transcript_<timestamp>_named.json
        <project>_process_transcript_<timestamp>_named.txt
        <project>_process_transcript_<timestamp>_named.srt
        <project>_process_transcript_<timestamp>_speakers.txt
        <project>_process_transcript_<timestamp>_by_clips.xlsx

    02_Transcripts/02_02_Clean/
        <clip_name>.srt   - Per-clip SRT files with local timecodes

Requirements:
    pip install requests openpyxl
    ollama serve
    brew install ffmpeg
"""

import argparse
import os
import sys
import json
import re
import subprocess
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Set, Tuple

try:
    import requests
except ImportError:
    print("ERROR: requests not installed. Run: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# Configuration
# ============================================================================

# Project structure
VIDEO_DIR = "01_Raw/01_01_Video"
AUDIO_DIR = "01_Raw/01_02_Audio"
TRANSCRIPTS_RUNS_DIR = "02_Transcripts/02_01_Runs"
TRANSCRIPTS_CLEAN_DIR = "02_Transcripts/02_02_Clean"
LOGS_DIR = "08_Logs"

VIDEO_EXTS = {".mp4", ".mov", ".m4v", ".mts", ".avi", ".mkv",
              ".MP4", ".MOV", ".M4V", ".MTS", ".AVI", ".MKV"}

# Ollama settings
OLLAMA_API_URL = "http://localhost:11434/api/generate"
DEFAULT_LLM_MODEL = "qwen2.5:32b"
OLLAMA_TIMEOUT = 600  # 10 minutes (больше контекста = дольше обработка)

# Speaker filtering
MIN_SPEAKER_SEGMENTS = 10  # Ignore speakers with fewer segments (likely noise)


# ============================================================================
# Project Paths
# ============================================================================

def get_project_paths(project_dir: str) -> dict:
    """Get all relevant paths for a project."""
    project_root = Path(project_dir).expanduser().resolve()
    
    if not project_root.exists():
        raise FileNotFoundError(f"Project folder not found: {project_root}")
    
    return {
        "project_root": project_root,
        "project_name": project_root.name,
        "video_dir": project_root / VIDEO_DIR,
        "audio_dir": project_root / AUDIO_DIR,
        "transcripts_runs_dir": project_root / TRANSCRIPTS_RUNS_DIR,
        "transcripts_clean_dir": project_root / TRANSCRIPTS_CLEAN_DIR,
        "logs_dir": project_root / LOGS_DIR,
        "speaker_map_file": project_root / TRANSCRIPTS_CLEAN_DIR / "speaker_names.json",
        "project_profile_file": project_root / "project_profile.txt"
    }


def find_latest_transcript(transcripts_dir: Path, project_name: str) -> Optional[Path]:
    """
    Find the most recent transcript JSON file from Step 1 (transcribe_project.py).
    
    Looks for: <project>_transcript_<timestamp>.json
    Excludes: *_named.json (already processed by this script)
              *_process_transcript_* (output of this script)
    """
    pattern = f"{project_name}_transcript_*.json"
    files = list(transcripts_dir.glob(pattern))
    
    # Exclude already processed files
    files = [f for f in files if "_named" not in f.name and "_process_transcript_" not in f.name]
    
    if not files:
        return None
    
    # Sort by modification time, newest first
    files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return files[0]


# ============================================================================
# Logging
# ============================================================================

def setup_logging(logs_dir: Path, project_name: str) -> logging.Logger:
    """Setup dual logging."""
    logs_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = logs_dir / f"{project_name}_process_transcript_{timestamp}.log"
    
    logger = logging.getLogger("process")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    logger.info(f"Log: {log_file}")
    
    return logger


# ============================================================================
# Timestamp Formatting
# ============================================================================

def format_timestamp(seconds: float) -> str:
    """Format as HH:MM:SS"""
    h, r = divmod(int(seconds), 3600)
    m, s = divmod(r, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def format_srt_timestamp(seconds: float) -> str:
    """Format for SRT: HH:MM:SS,mmm"""
    h, r = divmod(int(seconds), 3600)
    m, s = divmod(r, 60)
    ms = int((seconds - int(seconds)) * 1000)
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


# ============================================================================
# Video Clip Utilities
# ============================================================================

def natural_sort_key(s: str):
    """Sort strings with embedded numbers naturally."""
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]


def get_video_duration(video_path: Path) -> float:
    """Get video duration in seconds using ffprobe."""
    cmd = [
        "ffprobe",
        "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        str(video_path)
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return float(result.stdout.strip())
    except Exception:
        pass
    
    return 0.0


def get_video_clips(video_dir: Path, logger: logging.Logger) -> List[dict]:
    """
    Get list of video clips with their durations and offsets.
    
    Returns list of dicts:
    [
        {"file": "RYA-ZVE1-1146.MP4", "path": Path, "duration": 150.5, "start": 0.0, "end": 150.5},
        {"file": "RYA-ZVE1-1147.MP4", "path": Path, "duration": 120.0, "start": 150.5, "end": 270.5},
        ...
    ]
    """
    # Find all video files (exclude Archive folder)
    clips = []
    for ext in VIDEO_EXTS:
        for f in video_dir.glob(f"*{ext}"):
            # Skip files in Archive folder
            if "Archive" in str(f):
                continue
            if not f.name.startswith("."):
                clips.append(f)
    
    # Sort naturally
    clips.sort(key=lambda p: natural_sort_key(p.name))
    
    if not clips:
        logger.warning(f"No video clips found in {video_dir}")
        return []
    
    logger.info(f"Found {len(clips)} video clips")
    
    # Get durations and calculate offsets
    result = []
    current_offset = 0.0
    
    for clip_path in clips:
        duration = get_video_duration(clip_path)
        
        if duration <= 0:
            logger.warning(f"Could not get duration for {clip_path.name}, skipping")
            continue
        
        result.append({
            "file": clip_path.name,
            "path": clip_path,
            "duration": duration,
            "start": current_offset,
            "end": current_offset + duration
        })
        
        logger.debug(f"  {clip_path.name}: {format_timestamp(duration)} ({current_offset:.1f}s - {current_offset + duration:.1f}s)")
        current_offset += duration
    
    logger.info(f"Total duration: {format_timestamp(current_offset)}")
    return result


def find_clip_for_timestamp(clips: List[dict], timestamp: float) -> Tuple[Optional[dict], float]:
    """
    Find which clip contains the given timestamp.
    Returns (clip_info, local_timestamp) or (None, 0) if not found.
    """
    for clip in clips:
        if clip["start"] <= timestamp < clip["end"]:
            local_ts = timestamp - clip["start"]
            return clip, local_ts
    
    # Edge case: timestamp at very end
    if clips and timestamp >= clips[-1]["end"] - 0.1:
        clip = clips[-1]
        local_ts = timestamp - clip["start"]
        return clip, local_ts
    
    return None, 0.0


# ============================================================================
# Ollama LLM for Speaker Naming
# ============================================================================

def check_ollama_server(logger: logging.Logger) -> bool:
    """Check if Ollama server is running."""
    try:
        response = requests.get("http://localhost:11434/api/tags", timeout=5)
        if response.status_code == 200:
            models = response.json().get("models", [])
            model_names = [m.get("name", "") for m in models]
            logger.debug(f"Available models: {model_names}")
            return True
        return False
    except requests.exceptions.ConnectionError:
        return False
    except Exception as e:
        logger.debug(f"Ollama check error: {e}")
        return False


def call_ollama_api(prompt: str, model: str, logger: logging.Logger) -> Optional[str]:
    """Call Ollama API with specified model."""
    try:
        logger.debug(f"Calling Ollama API with {model}...")
        
        response = requests.post(
            OLLAMA_API_URL,
            json={
                "model": model,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.3,
                    "num_predict": 500,
                }
            },
            timeout=OLLAMA_TIMEOUT
        )
        
        if response.status_code == 200:
            result = response.json()
            return result.get("response", "")
        else:
            logger.warning(f"Ollama API error: {response.status_code}")
            return None
            
    except requests.exceptions.Timeout:
        logger.warning(f"Ollama timeout ({OLLAMA_TIMEOUT}s)")
        return None
    except requests.exceptions.ConnectionError:
        logger.warning("Ollama server not running")
        return None
    except Exception as e:
        logger.warning(f"Ollama API error: {e}")
        return None


def extract_speaker_names_llm(segments: List[dict], speakers: Set[str], model: str, logger: logging.Logger, project_profile: str = "") -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Use Ollama to identify speaker names/roles from context.
    
    Simple approach:
    1. Collect ALL utterances for each speaker
    2. Send to LLM with full context
    3. LLM determines who is who based on WHAT they say
    4. Return mapping: SPEAKER_XX -> Name/Role
    
    Returns: (speaker_names_map, speaker_context_map)
    """
    if not check_ollama_server(logger):
        logger.warning("Ollama server not running!")
        logger.warning("Start with: ollama serve")
        return {}, {}
    
    # Collect ALL utterances for each speaker
    speaker_utterances = {s: [] for s in speakers}
    speaker_counts = {s: 0 for s in speakers}
    
    for seg in segments:
        speaker = seg.get("speaker", "UNKNOWN")
        text = seg.get("text", "").strip()
        if speaker in speakers and text:
            speaker_counts[speaker] += 1
            speaker_utterances[speaker].append(text)
    
    # Sort speakers by segment count
    sorted_speakers = sorted(speakers, key=lambda s: speaker_counts[s], reverse=True)
    
    # Build transcript for LLM - give SUBSTANTIAL context for each speaker
    # More lines = better identification
    transcript_parts = []
    
    for speaker in sorted_speakers:
        utterances = speaker_utterances.get(speaker, [])
        count = speaker_counts[speaker]
        
        if not utterances:
            continue
        
        # Take up to 30 utterances spread across the interview
        # First 10, middle 10, last 10
        sample = []
        if len(utterances) <= 30:
            sample = utterances
        else:
            sample.extend(utterances[:10])  # Beginning
            mid = len(utterances) // 2
            sample.extend(utterances[mid-5:mid+5])  # Middle
            sample.extend(utterances[-10:])  # End
        
        # Remove duplicates
        seen = set()
        unique = [u for u in sample if not (u in seen or seen.add(u))]
        
        transcript_parts.append(f"\n=== {speaker} ({count} utterances total) ===")
        for i, text in enumerate(unique[:30], 1):
            # Truncate very long utterances
            transcript_parts.append(f'{i}. "{text[:400]}"')
    
    transcript = "\n".join(transcript_parts)
    
    # Build profile section if available
    profile_section = ""
    if project_profile:
        profile_section = f"""
KNOWN PARTICIPANTS (from project profile):
{project_profile}

Match speakers to these known participants based on what they say.
"""
    
    # Simple, direct prompt
    prompt = f"""You are analyzing an interview transcript to identify WHO each speaker is.

This is a YouTube interview about business in UAE/Saudi Arabia.
{profile_section}
SPEAKERS TO IDENTIFY:
{', '.join(sorted_speakers)}

Below are sample utterances from each speaker. Based on WHAT they say, determine:
- Their name (if mentioned or inferable)
- Their role (Interviewer, Guest, Business Owner, etc.)

{transcript}

RULES:
1. The person ASKING questions about the other person's business/life = INTERVIEWER
2. The person EXPLAINING their business/experience = GUEST  
3. Look for self-introductions: "My name is...", "I am...", "I own..."
4. Look for how others address them: "Thank you Hadi", "So Roman..."
5. Each speaker must have a UNIQUE name/role
6. Use actual names when found (Hadi, Roman, Ahmed, etc.)

OUTPUT FORMAT - JSON only:
{{
  "names": {{
    "SPEAKER_00": "Name or Role",
    "SPEAKER_01": "Name or Role"
  }},
  "context": {{
    "SPEAKER_00": "Brief description of who they are",
    "SPEAKER_01": "Brief description"
  }}
}}"""

    logger.info(f"Asking {model} to identify {len(speakers)} speakers...")
    logger.debug(f"Prompt length: {len(prompt)} chars")
    
    response = call_ollama_api(prompt, model, logger)
    
    if not response:
        return {}, {}
    
    logger.debug(f"LLM response: {response[:500]}")
    
    # Extract JSON - try new format first, then fallback to old
    try:
        # Try to find JSON with nested structure
        json_match = re.search(r'\{[\s\S]*"names"[\s\S]*\}', response)
        if json_match:
            data = json.loads(json_match.group())
            if "names" in data:
                raw_names = data.get("names", {})
                raw_context = data.get("context", {})
                names = validate_speaker_names(raw_names, speakers, logger)
                
                # Map context to SPEAKER_XX IDs
                # LLM might return context with names as keys instead of IDs
                mapped_context = {}
                for speaker_id, speaker_name in names.items():
                    # Try to find context by ID first, then by name
                    if speaker_id in raw_context:
                        mapped_context[speaker_id] = raw_context[speaker_id]
                    elif speaker_name in raw_context:
                        mapped_context[speaker_id] = raw_context[speaker_name]
                    else:
                        # Try case-insensitive match
                        for ctx_key, ctx_value in raw_context.items():
                            if ctx_key.lower() == speaker_name.lower() or ctx_key.lower() == speaker_id.lower():
                                mapped_context[speaker_id] = ctx_value
                                break
                
                logger.info(f"Speaker names: {names}")
                return names, mapped_context
        
        # Fallback: old format (flat dict)
        json_match = re.search(r'\{[^{}]+\}', response, re.DOTALL)
        if json_match:
            raw_names = json.loads(json_match.group())
            names = validate_speaker_names(raw_names, speakers, logger)
            logger.info(f"Speaker names: {names}")
            return names, {}
        
        logger.warning("Could not parse LLM response as JSON")
        return {}, {}
        
    except json.JSONDecodeError as e:
        logger.warning(f"Invalid JSON: {e}")
        return {}, {}


def analyze_and_merge_speakers(segments: List[dict], speakers: Set[str], model: str, logger: logging.Logger, min_segments: int = MIN_SPEAKER_SEGMENTS) -> Dict[str, str]:
    """
    Use LLM to analyze if some speakers should be merged (same person, different voice segments).
    Returns a mapping: {old_speaker_id: new_speaker_id} for merging.
    
    IMPORTANT: Speakers who alternate (A->B->A->B) are DIFFERENT people (dialogue).
    Only merge speakers who appear in separate parts of the recording without alternation.
    """
    if not check_ollama_server(logger):
        return {}
    
    # Count segments per speaker
    speaker_counts = {s: 0 for s in speakers}
    speaker_samples = {s: [] for s in speakers}
    
    for seg in segments:
        speaker = seg.get("speaker", "UNKNOWN")
        text = seg.get("text", "").strip()
        if speaker in speakers:
            speaker_counts[speaker] += 1
            if text and len(text) > 30 and len(speaker_samples[speaker]) < 5:
                speaker_samples[speaker].append(text[:200])
    
    # Filter out minor speakers (noise)
    major_speakers = {s for s in speakers if speaker_counts[s] >= min_segments}
    minor_speakers = speakers - major_speakers
    
    if minor_speakers:
        logger.info(f"Filtering out {len(minor_speakers)} minor speakers (<{min_segments} segments): {minor_speakers}")
    
    if len(major_speakers) <= 2:
        # No need to merge if only 1-2 speakers
        merge_map = {s: s for s in speakers}
        for minor in minor_speakers:
            merge_map[minor] = "MINOR_SPEAKER"
        return merge_map
    
    # === DIALOGUE ANALYSIS ===
    # Count how often speakers alternate (A->B means A and B are in dialogue = DIFFERENT people)
    alternation_counts = {}  # {(speaker_a, speaker_b): count}
    
    prev_speaker = None
    for seg in segments:
        curr_speaker = seg.get("speaker", "UNKNOWN")
        if curr_speaker in major_speakers and prev_speaker in major_speakers:
            if curr_speaker != prev_speaker:
                # They alternate - likely different people
                pair = tuple(sorted([prev_speaker, curr_speaker]))
                alternation_counts[pair] = alternation_counts.get(pair, 0) + 1
        prev_speaker = curr_speaker
    
    # Build dialogue pairs info for LLM
    dialogue_info = []
    for pair, count in sorted(alternation_counts.items(), key=lambda x: -x[1]):
        if count >= 5:  # Significant alternation
            dialogue_info.append(f"{pair[0]} <-> {pair[1]}: {count} alternations (likely DIFFERENT people)")
    
    dialogue_section = ""
    if dialogue_info:
        dialogue_section = f"""
DIALOGUE ANALYSIS (speakers who alternate are DIFFERENT people):
{chr(10).join(dialogue_info)}

CRITICAL: If two speakers alternate frequently (A->B->A->B pattern), they are having a CONVERSATION.
This means they are DIFFERENT PEOPLE - DO NOT merge them!
"""
    
    # Build prompt for LLM
    speaker_info = []
    for speaker in sorted(major_speakers, key=lambda s: speaker_counts[s], reverse=True):
        samples = speaker_samples.get(speaker, [])
        sample_text = " | ".join(samples[:3]) if samples else "No samples"
        speaker_info.append(f"{speaker} ({speaker_counts[speaker]} segments): \"{sample_text[:300]}\"")
    
    prompt = f"""Analyze these speakers from an interview transcript and determine if any are the SAME PERSON.

Sometimes voice recognition splits one person into multiple speaker IDs due to:
- Change in tone or volume
- Background noise  
- Speaking into/away from microphone

SPEAKERS TO ANALYZE:
{chr(10).join(speaker_info)}
{dialogue_section}
RULES FOR MERGING:
1. NEVER merge speakers who alternate frequently - they are having a conversation (different people!)
2. Only merge if speakers appear in SEPARATE PARTS of recording without talking to each other
3. The HOST asks questions, the GUEST answers - these are DIFFERENT people
4. Same topic does NOT mean same person - interviewer and guest discuss same topics!
5. Different languages (English vs Arabic) = DIFFERENT people

WHEN TO MERGE:
- Speaker appears only at beginning, another only at end (same person, voice changed)
- Very similar speech patterns AND no alternation between them

OUTPUT FORMAT - JSON only:
{{
  "merge_groups": [
    ["SPEAKER_00", "SPEAKER_02"]
  ],
  "reasoning": "Brief explanation"
}}

If NO speakers should be merged (most common case), return:
{{"merge_groups": [], "reasoning": "Speakers alternate in dialogue - different people"}}
"""

    logger.info(f"Asking {model} to analyze speaker merging...")
    
    response = call_ollama_api(prompt, model, logger)
    
    if not response:
        return {s: s for s in speakers}
    
    logger.debug(f"Merge analysis response: {response[:500]}")
    
    # Parse response
    try:
        json_match = re.search(r'\{[\s\S]*"merge_groups"[\s\S]*\}', response)
        if json_match:
            data = json.loads(json_match.group())
            merge_groups = data.get("merge_groups", [])
            reasoning = data.get("reasoning", "")
            
            # Build merge map
            merge_map = {s: s for s in speakers}
            
            # SAFETY CHECK: Don't merge speakers who alternate frequently
            valid_merge_groups = []
            for group in merge_groups:
                if len(group) >= 2:
                    # Check if any pair in group alternates frequently
                    should_merge = True
                    for i, s1 in enumerate(group):
                        for s2 in group[i+1:]:
                            pair = tuple(sorted([s1, s2]))
                            if alternation_counts.get(pair, 0) >= 5:
                                logger.warning(f"Blocking merge of {s1} and {s2} - they alternate {alternation_counts[pair]} times (dialogue)")
                                should_merge = False
                                break
                        if not should_merge:
                            break
                    
                    if should_merge:
                        valid_merge_groups.append(group)
            
            if valid_merge_groups:
                logger.info(f"Approved merges: {valid_merge_groups}")
                logger.info(f"Reasoning: {reasoning}")
                
                for group in valid_merge_groups:
                    primary = max(group, key=lambda s: speaker_counts.get(s, 0))
                    for speaker in group:
                        if speaker in merge_map:
                            merge_map[speaker] = primary
            else:
                if merge_groups:
                    logger.info("All proposed merges blocked (speakers alternate in dialogue)")
                else:
                    logger.info("No merges suggested")
            
            # Map minor speakers
            for minor in minor_speakers:
                merge_map[minor] = "MINOR_SPEAKER"
            
            return merge_map
    
    except (json.JSONDecodeError, KeyError) as e:
        logger.debug(f"Could not parse merge response: {e}")
    
    return {s: s for s in speakers}


def validate_speaker_names(raw_names: Dict[str, str], speakers: Set[str], logger: logging.Logger) -> Dict[str, str]:
    """Validate and clean speaker names from LLM."""
    validated = {}
    used_names = set()  # Track used names to ensure uniqueness
    
    invalid_patterns = [
        r'^unknown$',
        r'^speaker[\s_]*\d*$',  # Reject SPEAKER_XX format
        r'^person\s*\d*$',
        r'^\?+$',
        r'^n/?a$',
        r'^not\s+(specified|identified|known)$',
        r'^unidentified$',
    ]
    
    for speaker in speakers:
        name = raw_names.get(speaker, "").strip()
        
        is_valid = True
        
        if len(name) < 2:
            is_valid = False
        
        if len(name) > 30:  # Allow slightly longer descriptive roles
            is_valid = False
            logger.debug(f"Name too long for {speaker}: '{name}'")
        
        for pattern in invalid_patterns:
            if re.match(pattern, name.lower().replace("_", " ")):
                is_valid = False
                logger.debug(f"Invalid pattern for {speaker}: '{name}'")
                break
        
        if is_valid and not re.match(r'^[\w\s\-\'\.]+$', name, re.UNICODE):
            is_valid = False
            logger.debug(f"Invalid chars in name for {speaker}: '{name}'")
        
        # Check for uniqueness
        if is_valid and name.lower() in used_names:
            # Make unique by adding number
            base_name = name
            counter = 2
            while f"{base_name} {counter}".lower() in used_names:
                counter += 1
            name = f"{base_name} {counter}"
            logger.debug(f"Made unique: {speaker} -> '{name}'")
        
        if is_valid:
            validated[speaker] = name.title().strip()
            used_names.add(name.lower())
        else:
            # Generate descriptive fallback instead of keeping SPEAKER_XX
            fallback = f"Speaker {len(validated) + 1}"
            validated[speaker] = fallback
            used_names.add(fallback.lower())
            logger.debug(f"Fallback for {speaker}: '{fallback}'")
    
    return validated


def load_speaker_map(path: Path, logger: logging.Logger) -> Dict[str, str]:
    """Load manual speaker mapping from JSON."""
    if not path.exists():
        return {}
    
    try:
        with open(path) as f:
            names = json.load(f)
        
        # Check if it's just the template (values == keys)
        if all(k == v for k, v in names.items()):
            logger.info(f"Speaker map is template (not edited): {path}")
            return {}
        
        logger.info(f"Loaded speaker map from {path}")
        return names
    except Exception as e:
        logger.warning(f"Could not load speaker map: {e}")
        return {}


def load_project_profile(path: Path, logger: logging.Logger) -> str:
    """
    Load project profile with participant info.
    
    Expected format (project_profile.txt):
    ---
    Interview about coffee business in Saudi Arabia.
    
    Participants:
    - Roman (Host/Interviewer) - asks questions
    - Hadi Dawani (Main Guest) - entrepreneur, owns coffee chain
    - Ahmed (Coffee Shop Owner) - owns local coffee shop, speaks Arabic
    ---
    """
    if not path.exists():
        logger.debug(f"No project profile found: {path}")
        return ""
    
    try:
        with open(path, encoding="utf-8") as f:
            content = f.read().strip()
        
        if content:
            logger.info(f"Loaded project profile: {path}")
            return content
        return ""
    except Exception as e:
        logger.warning(f"Could not load project profile: {e}")
        return ""


def create_project_profile_template(path: Path, project_name: str) -> None:
    """Create a template project_profile.txt if it does not exist."""
    if path.exists():
        return
    
    template = f"""# Project Profile: {project_name}
# Fill in this information to help identify speakers accurately.

## Interview Topic
[Describe what this interview is about]

## Participants
# List all people in this interview with their roles:
# - Name (Role) - brief description, language they speak

- Roman (Host) - interviewer, asks questions in English
- [Guest Name] (Main Guest) - [their role/business], speaks [language]
- [Other Person] (Role) - [description]

## Notes
# Any additional context that might help identify speakers
# e.g. "The coffee shop owner speaks mostly Arabic"
# e.g. "There's background conversation from customers"
"""
    
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(template)
    except Exception as e:
        # Не критично - просто не создаём шаблон
        pass


def apply_speaker_names(segments: List[dict], names: Dict[str, str]) -> List[dict]:
    """Replace SPEAKER_XX with actual names, preserving original ID for debugging."""
    result = []
    for seg in segments:
        new_seg = seg.copy()
        speaker = seg.get("speaker", "UNKNOWN")
        
        # Always save original speaker ID for debugging (if not already saved)
        if "original_speaker_id" not in new_seg:
            new_seg["original_speaker_id"] = speaker
        
        if speaker in names and names[speaker] != speaker:
            new_seg["speaker"] = names[speaker]
        
        result.append(new_seg)
    return result


def apply_speaker_merge(segments: List[dict], merge_map: Dict[str, str], logger: logging.Logger) -> List[dict]:
    """
    Apply speaker merging to segments.
    Replaces speaker IDs according to merge_map.
    Preserves original_speaker_id for debugging.
    """
    result = []
    merge_count = 0
    
    for seg in segments:
        new_seg = seg.copy()
        speaker = seg.get("speaker", "UNKNOWN")
        
        # Save original ID before any modifications
        if "original_speaker_id" not in new_seg:
            new_seg["original_speaker_id"] = speaker
        
        if speaker in merge_map:
            new_speaker = merge_map[speaker]
            if new_speaker != speaker:
                if new_speaker == "MINOR_SPEAKER":
                    new_seg["speaker"] = "Minor Speaker"
                else:
                    new_seg["speaker"] = new_speaker
                merge_count += 1
        
        result.append(new_seg)
    
    if merge_count > 0:
        logger.info(f"Merged {merge_count} segments")
    
    return result


# ============================================================================
# Output Writers
# ============================================================================

def save_named_txt(segments: List[dict], speakers: Set[str], 
                   project_name: str, output_path: Path, metadata: dict):
    """Save human-readable transcript with named speakers."""
    lines = [
        "=" * 60,
        "PROJECT TRANSCRIPTION (Named Speakers)",
        "=" * 60,
        "",
        f"Project: {project_name}",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Language: {metadata.get('language', 'unknown')}",
        f"Speakers ({len(speakers)}): {', '.join(sorted(speakers))}",
        "",
        "-" * 60,
        "TRANSCRIPT",
        "-" * 60,
        ""
    ]
    
    for seg in segments:
        ts = format_timestamp(seg["start"])
        speaker = seg.get("speaker", "UNKNOWN")
        text = seg.get("text", "")
        lines.append(f"[{ts}] {speaker}:")
        lines.append(f"  {text}")
        lines.append("")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def save_named_json(segments: List[dict], speakers: Set[str],
                    project_name: str, output_path: Path, 
                    metadata: dict, speaker_map: Dict[str, str]):
    """Save JSON with named speakers."""
    data = {
        "project_name": project_name,
        "generated": datetime.now().isoformat(),
        "language": metadata.get("language"),
        "num_speakers": len(speakers),
        "speakers": sorted(list(speakers)),
        "speaker_map": speaker_map,
        "total_segments": len(segments),
        "segments": segments
    }
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def save_clips_xlsx(segments: List[dict], clips: List[dict], output_path: Path, logger: logging.Logger):
    """
    Save XLSX table with segments split by clips.
    
    Columns: file | timecode | speaker | original_id | text
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript"
    
    # Header
    headers = ["file", "timecode", "speaker", "original_id", "text"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
    
    # Data rows
    row_num = 2
    segments_mapped = 0
    
    for seg in segments:
        global_ts = seg["start"]
        clip, local_ts = find_clip_for_timestamp(clips, global_ts)
        
        if clip is None:
            logger.debug(f"No clip found for timestamp {global_ts:.1f}s")
            continue
        
        ws.cell(row=row_num, column=1, value=clip["file"])
        ws.cell(row=row_num, column=2, value=format_timestamp(local_ts))
        ws.cell(row=row_num, column=3, value=seg.get("speaker", "UNKNOWN"))
        ws.cell(row=row_num, column=4, value=seg.get("original_speaker_id", seg.get("speaker", "")))
        ws.cell(row=row_num, column=5, value=seg.get("text", ""))
        
        row_num += 1
        segments_mapped += 1
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 80
    
    wb.save(output_path)
    logger.info(f"Mapped {segments_mapped}/{len(segments)} segments to clips")


def save_clip_srt(segments: List[dict], clip: dict, output_path: Path) -> int:
    """
    Save SRT file for a single clip with local timecodes.
    Returns number of subtitles written.
    """
    lines = []
    idx = 1
    
    for seg in segments:
        global_ts = seg["start"]
        
        # Check if segment belongs to this clip
        if not (clip["start"] <= global_ts < clip["end"]):
            continue
        
        # Calculate local timestamps
        local_start = global_ts - clip["start"]
        local_end = seg["end"] - clip["start"]
        
        # Clamp end to clip duration
        local_end = min(local_end, clip["duration"])
        
        text = seg.get("text", "").strip()
        if not text:
            continue
        
        speaker = seg.get("speaker", "UNKNOWN")
        
        lines.append(str(idx))
        lines.append(f"{format_srt_timestamp(local_start)} --> {format_srt_timestamp(local_end)}")
        lines.append(f"[{speaker}] {text}")
        lines.append("")
        
        idx += 1
    
    if lines:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    
    return idx - 1


def save_full_srt(segments: List[dict], output_path: Path) -> int:
    """
    Save SRT file for full audio with global timecodes (named speakers).
    Returns number of subtitles written.
    """
    lines = []
    idx = 1
    
    for seg in segments:
        text = seg.get("text", "").strip()
        if not text:
            continue
        
        start = seg["start"]
        end = seg["end"]
        speaker = seg.get("speaker", "UNKNOWN")
        
        lines.append(str(idx))
        lines.append(f"{format_srt_timestamp(start)} --> {format_srt_timestamp(end)}")
        lines.append(f"[{speaker}] {text}")
        lines.append("")
        
        idx += 1
    
    if lines:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
    
    return idx - 1


def save_speakers_txt(speaker_map: Dict[str, str], speaker_context: Dict[str, str], 
                      segments: List[dict], output_path: Path) -> None:
    """
    Save speakers info file with ID, name, and context.
    """
    # Count segments per speaker
    speaker_counts = {}
    for seg in segments:
        spk = seg.get("speaker", "UNKNOWN")
        speaker_counts[spk] = speaker_counts.get(spk, 0) + 1
    
    lines = [
        "=" * 60,
        "SPEAKERS INFO",
        "=" * 60,
        "",
    ]
    
    for original_id in sorted(speaker_map.keys()):
        name = speaker_map.get(original_id, original_id)
        context = speaker_context.get(original_id, "No context available")
        count = speaker_counts.get(name, speaker_counts.get(original_id, 0))
        
        lines.append(f"ID: {original_id}")
        lines.append(f"Name: {name}")
        lines.append(f"Segments: {count}")
        lines.append(f"Context: {context}")
        lines.append("-" * 40)
        lines.append("")
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ============================================================================
# Main
# ============================================================================

def main():
    """
    Main entry point for transcript processing.
    
    Workflow:
    1. Parse arguments and load project paths
    2. Load transcript JSON from Step 1
    3. Phase 1a: Analyze and merge speakers (dialogue detection)
    4. Phase 1b: Name speakers via LLM
    5. Phase 2: Analyze video clips for timing
    6. Phase 3: Save outputs (JSON, TXT, SRT, XLSX)
    """
    parser = argparse.ArgumentParser(
        description="YTAI Step 2: Process Transcript (naming + split by clips)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Examples:\n"
               "    python process_transcript.py --project \"/Volumes/RYA Blue/PROJECT\"\n"
               "    python process_transcript.py --project \"/Volumes/RYA Blue/PROJECT\" --skip-llm\n"
               "    python process_transcript.py --project \"/Volumes/RYA Blue/PROJECT\" --no-merge"
    )
    
    parser.add_argument("--project", required=True,
                        help="Project folder path")
    parser.add_argument("--json", type=Path,
                        help="Specific transcript JSON (default: latest)")
    parser.add_argument("--skip-llm", action="store_true",
                        help="Skip LLM speaker naming (use manual map or IDs)")
    parser.add_argument("--llm-model", default=DEFAULT_LLM_MODEL,
                        help=f"Ollama model for speaker naming (default: {DEFAULT_LLM_MODEL})")
    parser.add_argument("--no-merge", action="store_true",
                        help="Skip LLM speaker merge analysis")
    parser.add_argument("--min-segments", type=int, default=MIN_SPEAKER_SEGMENTS,
                        help=f"Minimum segments to consider speaker (default: {MIN_SPEAKER_SEGMENTS})")
    
    args = parser.parse_args()
    
    # Get project paths
    try:
        paths = get_project_paths(args.project)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    
    # Setup directories
    paths["transcripts_runs_dir"].mkdir(parents=True, exist_ok=True)
    paths["transcripts_clean_dir"].mkdir(parents=True, exist_ok=True)
    paths["logs_dir"].mkdir(parents=True, exist_ok=True)
    
    # Setup logging
    logger = setup_logging(paths["logs_dir"], paths["project_name"])
    
    # Find transcript JSON
    if args.json:
        json_path = args.json
    else:
        json_path = find_latest_transcript(paths["transcripts_runs_dir"], paths["project_name"])
    
    if not json_path or not json_path.exists():
        logger.error(f"No transcript JSON found. Run transcribe_project.py first.")
        sys.exit(1)
    
    # Load transcript
    logger.info("=" * 60)
    logger.info("YTAI STEP 2: PROCESS TRANSCRIPT")
    logger.info("=" * 60)
    logger.info(f"Project: {paths['project_name']}")
    logger.info(f"Input: {json_path}")
    if not args.skip_llm:
        logger.info(f"LLM Model: {args.llm_model}")
    logger.info("")
    
    # Load and validate JSON
    try:
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON file: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Could not read file: {e}")
        sys.exit(1)
    
    segments = data.get("segments", [])
    if not segments:
        logger.error("No segments found in transcript JSON")
        sys.exit(1)
    
    original_speakers = set(data.get("speakers", []))
    if not original_speakers:
        # Extract speakers from segments
        original_speakers = set(seg.get("speaker", "UNKNOWN") for seg in segments)
        original_speakers.discard("UNKNOWN")
    
    metadata = {
        "language": data.get("language"),
        "model": data.get("model")
    }
    
    logger.info(f"Loaded: {len(segments)} segments, {len(original_speakers)} speakers")
    logger.info(f"Speakers: {', '.join(sorted(original_speakers))}")
    logger.info("")
    
    # Check video directory exists
    if not paths["video_dir"].exists():
        logger.error(f"Video directory not found: {paths['video_dir']}")
        logger.error("Make sure video clips are in 01_Raw/01_01_Video/")
        sys.exit(1)
    
    # Phase 1a: Analyze and merge speakers
    logger.info("PHASE 1a: Speaker Analysis & Merge")
    logger.info("-" * 40)
    
    merge_map = {}
    if not args.no_merge and not args.skip_llm:
        merge_map = analyze_and_merge_speakers(segments, original_speakers, args.llm_model, logger, args.min_segments)
        
        # Apply merge
        if merge_map:
            segments = apply_speaker_merge(segments, merge_map, logger)
            
            # Update speakers set after merge
            merged_speakers = set(seg.get("speaker", "UNKNOWN") for seg in segments)
            merged_speakers.discard("UNKNOWN")
            merged_speakers.discard("Minor Speaker")
            
            if len(merged_speakers) < len(original_speakers):
                logger.info(f"Speakers after merge: {len(merged_speakers)} (was {len(original_speakers)})")
                original_speakers = merged_speakers
    else:
        logger.info("Speaker merge analysis skipped")
    
    logger.info("")
    
    # Phase 1b: Speaker naming
    logger.info("PHASE 1b: Speaker Naming")
    logger.info("-" * 40)
    
    # Load project profile if available
    project_profile = load_project_profile(paths["project_profile_file"], logger)
    if not project_profile:
        # Create template for future use
        create_project_profile_template(paths["project_profile_file"], paths["project_name"])
        logger.info(f"Created profile template: {paths['project_profile_file']}")
        logger.info("TIP: Fill in project_profile.txt to improve speaker identification")
    
    speaker_map = {}
    speaker_context = {}
    
    # Try manual map first
    speaker_map = load_speaker_map(paths["speaker_map_file"], logger)
    
    # Try LLM if needed
    if not args.skip_llm and not speaker_map:
        speaker_map, speaker_context = extract_speaker_names_llm(
            segments, original_speakers, args.llm_model, logger, project_profile
        )
    
    # Apply names
    if speaker_map:
        segments = apply_speaker_names(segments, speaker_map)
        
        # Update speaker_names.json with LLM results
        with open(paths["speaker_map_file"], "w") as f:
            json.dump(speaker_map, f, indent=2)
        logger.info(f"Updated: {paths['speaker_map_file']}")
    else:
        logger.info("No speaker naming applied (using original IDs)")
    
    # Get final speakers
    final_speakers = set(seg.get("speaker", "UNKNOWN") for seg in segments)
    final_speakers.discard("UNKNOWN")
    final_speakers.discard("Minor Speaker")
    logger.info(f"Final speakers: {', '.join(sorted(final_speakers))}")
    logger.info("")
    
    # Phase 2: Get video clips
    logger.info("PHASE 2: Analyzing Video Clips")
    logger.info("-" * 40)
    
    clips = get_video_clips(paths["video_dir"], logger)
    
    if not clips:
        logger.error("No video clips found!")
        sys.exit(1)
    
    logger.info("")
    
    # Phase 3: Save outputs
    logger.info("PHASE 3: Saving Results")
    logger.info("-" * 40)
    
    # Generate output base name (includes script name for clarity)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_base = paths["transcripts_runs_dir"] / f"{paths['project_name']}_process_transcript_{timestamp}"
    
    # Save named JSON
    named_json = output_base.with_name(output_base.name + "_named.json")
    save_named_json(segments, final_speakers, paths["project_name"], 
                    named_json, metadata, speaker_map)
    logger.info(f"JSON: {named_json}")
    
    # Save named TXT
    named_txt = output_base.with_name(output_base.name + "_named.txt")
    save_named_txt(segments, final_speakers, paths["project_name"], named_txt, metadata)
    logger.info(f"TXT: {named_txt}")
    
    # Save XLSX by clips
    xlsx_path = output_base.with_name(output_base.name + "_by_clips.xlsx")
    save_clips_xlsx(segments, clips, xlsx_path, logger)
    logger.info(f"XLSX: {xlsx_path}")
    
    # Save full SRT (named speakers, global timecodes)
    full_srt_path = output_base.with_name(output_base.name + "_named.srt")
    full_srt_count = save_full_srt(segments, full_srt_path)
    logger.info(f"SRT (full): {full_srt_path} ({full_srt_count} subtitles)")
    
    # Save speakers info file
    speakers_txt_path = output_base.with_name(output_base.name + "_speakers.txt")
    save_speakers_txt(speaker_map, speaker_context, segments, speakers_txt_path)
    logger.info(f"Speakers: {speakers_txt_path}")
    
    # Save per-clip SRT files
    logger.info("")
    logger.info("Generating per-clip SRT files...")
    
    srt_count = 0
    for clip in clips:
        # SRT filename: same as video but .srt
        srt_name = Path(clip["file"]).stem + ".srt"
        srt_path = paths["transcripts_clean_dir"] / srt_name
        
        num_subs = save_clip_srt(segments, clip, srt_path)
        
        if num_subs > 0:
            logger.info(f"  {srt_name}: {num_subs} subtitles")
            srt_count += 1
    
    logger.info(f"Generated {srt_count} SRT files in {paths['transcripts_clean_dir']}")
    
    # Summary
    logger.info("")
    logger.info("=" * 60)
    logger.info("STEP 2 COMPLETE")
    logger.info("=" * 60)
    logger.info(f"Speakers: {', '.join(sorted(final_speakers))}")
    logger.info(f"Segments: {len(segments)}")
    logger.info(f"Clips: {len(clips)}")
    logger.info(f"SRT files: {srt_count}")
    logger.info("")
    logger.info("Output files:")
    logger.info(f"  {named_json}")
    logger.info(f"  {named_txt}")
    logger.info(f"  {full_srt_path}")
    logger.info(f"  {speakers_txt_path}")
    logger.info(f"  {xlsx_path}")
    logger.info(f"  {paths['transcripts_clean_dir']}/*.srt ({srt_count} clips)")


if __name__ == "__main__":
    main()
