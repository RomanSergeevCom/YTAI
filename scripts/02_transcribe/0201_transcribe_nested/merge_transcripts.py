"""Cross-scene transcript merger for nested multi-scene projects.

Reads per-scene transcript JSON files from 01_Media/Source/Transcription/ and
combines them into:
  1. {project}_transcript.json — standard pipeline format with clips[].segments[]
  2. {project}_transcript.xlsx — 3-tab Excel (Transcript, Summary, Media)

The merged transcript follows the same schema as per-scene transcripts from
transcribe_project.py, with an added "scene" field on each clip. This makes it
directly compatible with the brief creation pipeline (0501_brief).
"""
import json
import re
from collections import Counter
from pathlib import Path

_CODE_RE = re.compile(r'^(YT[A-Z]{2,4}\d+)_')

def _project_code(name: str) -> str:
    m = _CODE_RE.match(name)
    return m.group(1) if m else name


def merge_transcripts(project: Path, scene_names: list) -> Path:
    """Merge per-scene transcripts into a single {project}_transcript.json.

    Produces standard pipeline format: clips[].segments[] with full metadata.
    Each clip gets a "scene" field added. Speakers and stats are aggregated.

    Output: {project}/01_Media/Source/Transcription/{project.name}_transcript.json

    Args:
        project: Project root path.
        scene_names: List of scene names to include in the merge.

    Returns:
        Path to the written transcript file.
    """
    tr_dir = project / "01_Media" / "Source" / "Transcription"
    scenes_dir = tr_dir / "scenes"
    if not scenes_dir.exists():
        scenes_dir = tr_dir  # fallback: flat layout

    all_clips = []
    all_speakers = {}
    totals = {
        "total_duration": 0.0,
        "total_words": 0,
        "total_segments": 0,
        "low_confidence_segments": 0,
    }
    conf_weighted_sum = 0.0
    conf_weighted_count = 0.0
    languages = []

    for scene in sorted(scene_names):
        tp = scenes_dir / f"{scene}_transcript.json"
        if not tp.exists():
            tp = tr_dir / f"{scene}_transcript.json"  # fallback
        if not tp.exists():
            continue
        with open(tp) as f:
            data = json.load(f)

        # Merge clips — add scene field
        for clip in data.get("clips", []):
            clip["scene"] = scene
            all_clips.append(clip)

        # Merge speakers (first occurrence wins)
        for sp_id, sp_info in data.get("speakers", {}).items():
            if sp_id not in all_speakers:
                all_speakers[sp_id] = sp_info

        # Aggregate stats
        stats = data.get("stats", {})
        for key in totals:
            totals[key] += stats.get(key, 0)

        # Weighted confidence
        scene_dur = stats.get("total_duration", 0)
        scene_conf = stats.get("avg_confidence", 0)
        if scene_dur > 0:
            conf_weighted_sum += scene_conf * scene_dur
            conf_weighted_count += scene_dur

        # Collect language
        lang = data.get("language")
        if lang:
            languages.append(lang)

    # Determine dominant language
    lang_counts = Counter(languages)
    main_lang = lang_counts.most_common(1)[0][0] if lang_counts else "unknown"

    # Weighted average confidence
    avg_conf = round(conf_weighted_sum / conf_weighted_count, 4) if conf_weighted_count > 0 else 0.0

    merged = {
        "version": "3.0",
        "project": project.name,
        "language": main_lang,
        "speakers": all_speakers,
        "stats": {
            **totals,
            "num_speakers": len(all_speakers),
            "avg_confidence": avg_conf,
        },
        "structure": {
            "type": "nested",
            "transcription_dir": "Transcription",
            "scenes": sorted(scene_names),
        },
        "clips": all_clips,
    }

    code = _project_code(project.name)
    out_path = tr_dir / f"{code}_transcript.json"
    with open(out_path, "w") as f:
        json.dump(merged, f, indent=2, ensure_ascii=False)

    # Generate compact transcript for Claude Desktop Assembly → Setup/
    try:
        setup_dir = project / "01_Media" / "Source" / "Setup"
        setup_dir.mkdir(parents=True, exist_ok=True)
        assembly_script = Path(__file__).resolve().parent.parent.parent / "05_editing" / "0502_assembly" / "generate_transcript_assembly.py"
        if assembly_script.exists():
            import importlib.util
            spec = importlib.util.spec_from_file_location("gen_assembly", assembly_script)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            compact = mod.generate_assembly_transcript(merged)
            assembly_out = setup_dir / f"{code}_Claude4_assembly.json"
            with open(assembly_out, "w", encoding="utf-8") as af:
                json.dump(compact, af, indent=2, ensure_ascii=False)
    except Exception:
        pass  # Non-critical — can be generated separately

    return out_path


def _find_scene_transcript(tr_dir: Path, scene: str) -> Path | None:
    """Find scene transcript in scenes/ subdir or flat layout."""
    for candidate in [tr_dir / "scenes" / f"{scene}_transcript.json",
                       tr_dir / f"{scene}_transcript.json"]:
        if candidate.exists():
            return candidate
    return None


def _fmt_duration(seconds: float) -> str:
    """Format seconds as M:SS or H:MM:SS."""
    s = int(seconds)
    if s >= 3600:
        return f"{s // 3600}:{(s % 3600) // 60:02d}:{s % 60:02d}"
    return f"{s // 60}:{s % 60:02d}"


def generate_merged_xlsx(project: Path, scene_names: list) -> Path:
    """Generate a merged XLSX with 3 tabs from per-scene transcript JSONs.

    Sheet 1 (Transcript): all segments from all scenes with Scene column.
    Sheet 2 (Summary): per-scene stats table.
    Sheet 3 (Media): all clips with metadata.

    Output: {project}/01_Media/Source/Transcription/{project.name}_transcript.xlsx

    Args:
        project: Project root path.
        scene_names: List of scene names to include.

    Returns:
        Path to the written XLSX file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tr_dir = project / "01_Media" / "Source" / "Transcription"

    # Styles
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    clip_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    clip_font = Font(bold=True, size=11, color="1F3864")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    total_font = Font(bold=True, size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    wb = openpyxl.Workbook()

    # ── Sheet 1: Transcript ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transcript"
    headers = ["Scene", "Timecode", "#", "Clip", "Start", "End",
               "Duration", "Words", "Speaker", "Text", "Confidence"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    seg_n = 0

    for scene in sorted(scene_names):
        tp = _find_scene_transcript(tr_dir, scene)
        if not tp:
            continue
        data = json.load(open(tp))
        speaker_map = data.get("speakers", {})
        offsets = {}
        for c in data.get("clips", []):
            offsets[c["clip_id"]] = c.get("offset", 0.0)

        for clip in data.get("clips", []):
            cid = clip["clip_id"]
            co = offsets.get(cid, 0.0)
            clip_dur = clip.get("duration", 0)
            segments = clip.get("segments", [])
            clip_words = sum(len(s.get("text", "").split()) for s in segments)

            # Clip separator row
            sep = f"{scene} / {cid}  |  {_fmt_duration(clip_dur)}  |  {clip_words} words  |  {len(segments)} segments"
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = clip_fill
            ws.cell(row=row, column=1, value=sep).font = clip_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            row += 1

            if not segments:
                seg_n += 1
                ws.cell(row=row, column=1, value=scene)
                ws.cell(row=row, column=2, value=_fmt_duration(co))
                ws.cell(row=row, column=3, value=seg_n)
                ws.cell(row=row, column=4, value=cid)
                ws.cell(row=row, column=10, value="[no speech]")
                row += 1
                continue

            for seg in segments:
                seg_n += 1
                gs = co + seg.get("start", 0)
                sp_id = seg.get("speaker", seg.get("speaker_id", ""))
                sp_name = speaker_map.get(sp_id, {}).get("name", sp_id) if isinstance(speaker_map, dict) else sp_id

                ws.cell(row=row, column=1, value=scene)
                ws.cell(row=row, column=2, value=_fmt_duration(gs))
                ws.cell(row=row, column=3, value=seg_n)
                ws.cell(row=row, column=4, value=cid)
                ws.cell(row=row, column=5, value=round(seg.get("start", 0), 2))
                ws.cell(row=row, column=6, value=round(seg.get("end", 0), 2))
                ws.cell(row=row, column=7, value=round(seg.get("end", 0) - seg.get("start", 0), 2))
                ws.cell(row=row, column=8, value=len(seg.get("text", "").split()))
                ws.cell(row=row, column=9, value=sp_name)
                tc = ws.cell(row=row, column=10, value=seg.get("text", ""))
                tc.alignment = wrap
                ws.cell(row=row, column=11, value=round(seg.get("confidence", seg.get("avg_confidence", 0)), 3))
                if seg.get("low_confidence"):
                    tc.fill = yellow_fill
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).border = thin_border
                row += 1

    ws.freeze_panes = "A2"
    widths = {"A": 18, "B": 10, "C": 5, "D": 10, "E": 8, "F": 8,
              "G": 9, "H": 7, "I": 14, "J": 80, "K": 11}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    ws2.cell(row=1, column=1, value="Project").font = total_font
    ws2.cell(row=1, column=2, value=project.name)
    ws2.cell(row=2, column=1, value="Scenes").font = total_font
    ws2.cell(row=2, column=2, value=len(scene_names))

    sum_headers = ["Scene", "Clips", "Segments", "Words", "Duration",
                   "Speakers", "Language", "Avg Confidence", "Low Conf Segs"]
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    srow = 5
    sum_totals = {"clips": 0, "segments": 0, "words": 0, "duration": 0.0, "low_conf": 0}

    for scene in sorted(scene_names):
        tp = _find_scene_transcript(tr_dir, scene)
        if not tp:
            continue
        data = json.load(open(tp))
        stats = data.get("stats", {})
        n_clips = len(data.get("clips", []))
        n_segs = stats.get("total_segments", 0)
        n_words = stats.get("total_words", 0)
        dur = stats.get("total_duration", 0)
        n_speakers = stats.get("num_speakers", 0)
        lang = data.get("language", "?")
        avg_conf = stats.get("avg_confidence", 0)
        low_conf = stats.get("low_confidence_segments", 0)

        ws2.cell(row=srow, column=1, value=scene)
        ws2.cell(row=srow, column=2, value=n_clips)
        ws2.cell(row=srow, column=3, value=n_segs)
        ws2.cell(row=srow, column=4, value=n_words)
        ws2.cell(row=srow, column=5, value=_fmt_duration(dur))
        ws2.cell(row=srow, column=6, value=n_speakers)
        ws2.cell(row=srow, column=7, value=lang)
        ws2.cell(row=srow, column=8, value=round(avg_conf, 3))
        ws2.cell(row=srow, column=9, value=low_conf)

        sum_totals["clips"] += n_clips
        sum_totals["segments"] += n_segs
        sum_totals["words"] += n_words
        sum_totals["duration"] += dur
        sum_totals["low_conf"] += low_conf
        srow += 1

    # Total row
    ws2.cell(row=srow, column=1, value="TOTAL").font = total_font
    ws2.cell(row=srow, column=2, value=sum_totals["clips"]).font = total_font
    ws2.cell(row=srow, column=3, value=sum_totals["segments"]).font = total_font
    ws2.cell(row=srow, column=4, value=sum_totals["words"]).font = total_font
    ws2.cell(row=srow, column=5, value=_fmt_duration(sum_totals["duration"])).font = total_font
    ws2.cell(row=srow, column=9, value=sum_totals["low_conf"]).font = total_font

    for col in range(1, len(sum_headers) + 1):
        ws2.column_dimensions[chr(64 + col)].width = 16

    # ── Sheet 3: Media ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Media")
    media_headers = ["Scene", "Clip", "File", "Duration", "Resolution",
                     "FPS", "Frames", "Video Codec", "Bitrate (Mbps)"]
    for col, h in enumerate(media_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    mrow = 2
    for scene in sorted(scene_names):
        tp = _find_scene_transcript(tr_dir, scene)
        if not tp:
            continue
        data = json.load(open(tp))
        for clip in data.get("clips", []):
            cid = clip["clip_id"]
            mi = clip.get("media", {})
            ws3.cell(row=mrow, column=1, value=scene)
            ws3.cell(row=mrow, column=2, value=cid)
            ws3.cell(row=mrow, column=3, value=clip.get("filename", ""))
            ws3.cell(row=mrow, column=4, value=_fmt_duration(clip.get("duration", 0)))
            w = mi.get("width", 0)
            h = mi.get("height", 0)
            ws3.cell(row=mrow, column=5, value=f"{w}\u00d7{h}" if w else "")
            ws3.cell(row=mrow, column=6, value=mi.get("fps", ""))
            ws3.cell(row=mrow, column=7, value=mi.get("nb_frames", ""))
            ws3.cell(row=mrow, column=8, value=mi.get("video_codec", ""))
            ws3.cell(row=mrow, column=9, value=mi.get("video_bitrate_mbps", ""))
            for c in range(1, len(media_headers) + 1):
                ws3.cell(row=mrow, column=c).border = thin_border
            mrow += 1

    for col in range(1, len(media_headers) + 1):
        ws3.column_dimensions[chr(64 + col)].width = 14
    ws3.column_dimensions["A"].width = 20
    ws3.freeze_panes = "A2"

    # Save
    out_path = tr_dir / f"{project.name}_transcript.xlsx"
    wb.save(str(out_path))
    return out_path
