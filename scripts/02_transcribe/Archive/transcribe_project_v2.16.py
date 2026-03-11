#!/usr/bin/env python3
"""
YTAI Transcribe v2.16
Полная транскрипция: word-level timestamps, Premiere Transcript JSON,
глобальная диаризация с маппингом на клипы.

Usage:
    python 01_transcribe_project.py --project "/path/to/Interview" -n 2
    python 01_transcribe_project.py --project "/path/to/video.MP4" -n 2 -y
"""

import argparse
import gzip
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path

import warnings
import numpy as np

# Suppress noisy third-party warnings (torchcodec, torchaudio, whisper FP16)
# Use module-level filter for pyannote's io.py which emits torchcodec warning at import
warnings.filterwarnings("ignore", category=UserWarning, module=r"pyannote\.audio\.core\.io")
warnings.filterwarnings("ignore", category=UserWarning, module=r"speechbrain")
warnings.filterwarnings("ignore", category=UserWarning, module=r"whisper")

# ─── Constants ───────────────────────────────────────────────────────────────

VERSION = "2.16"

VIDEO_EXTENSIONS = {'.mp4', '.mov', '.mkv', '.avi', '.mts', '.m4v'}

SPEED_COEFFICIENTS = {
    "extract_audio": 0.03,
    "diarization":   0.15,
    "whisper":       0.70,
    "speaker_map":   0.001,
    "generate":      0.002,
    "prproj":        0.001,
}

PUNCT_PATTERN = re.compile(r'^(.*?)([.?!,;:\u2026\-"\']+)$')
EOS_CHARS = set('.?!')

LOW_CONFIDENCE_THRESHOLD = 0.7
PAUSE_SPLIT_THRESHOLD = 1.5  # seconds — split segment if pause > this
LONG_SEGMENT_PAUSE = 1.0     # seconds — split >30s segments at pauses > this

TICKS_PER_FRAME_25FPS = 10160640000
TICKS_PER_SECOND = TICKS_PER_FRAME_25FPS * 25  # 254016000000

TEMPLATE_PATH = Path.home() / "YTAI" / "templates" / "premiere_template.prproj"

LANG_MAP = {
    "en": "en-us", "ru": "ru-ru", "ar": "ar-sa", "fr": "fr-fr",
    "de": "de-de", "es": "es-es", "it": "it-it", "pt": "pt-br",
    "ja": "ja-jp", "ko": "ko-kr", "zh": "zh-cn", "hi": "hi-in",
}

# ─── Colors ──────────────────────────────────────────────────────────────────

class C:
    BOLD       = "\033[1m"
    DIM        = "\033[2m"
    ITALIC     = "\033[3m"
    GREEN      = "\033[92m"
    YELLOW     = "\033[93m"
    RED        = "\033[91m"
    CYAN       = "\033[96m"
    BLUE       = "\033[94m"
    MAGENTA    = "\033[95m"
    WHITE      = "\033[97m"
    BG_CYAN    = "\033[46m"
    BG_GREEN   = "\033[42m"
    BG_BLUE    = "\033[44m"
    RESET      = "\033[0m"
    CLEAR_LINE = "\033[2K\r"

# ─── ANSI helpers ────────────────────────────────────────────────────────────

_ANSI_RE = re.compile(r'\033\[[0-9;]*m')

def vlen(s):
    """Visual length of string (excluding ANSI escape codes)."""
    return len(_ANSI_RE.sub('', s))

def _trow(content, W=62, border="│"):
    """Table row: pad content to W visual chars, add borders."""
    pad = W - vlen(content)
    return f"  {C.DIM}{border}{C.RESET}{content}{' ' * max(pad, 0)}{C.DIM}{border}{C.RESET}"

def compute_drift(st, est_times):
    """Compute drift factor: ratio of actual/estimated for completed stages.

    Returns value clamped to [0.3, 3.0]. Used to adjust remaining ETAs.
    """
    est_done = sum(est_times.get(k, 0) for k in st if est_times.get(k, 0) > 1)
    act_done = sum(st[k] for k in st if est_times.get(k, 0) > 1)
    return min(max(act_done / est_done, 0.3), 3.0) if est_done > 1 else 1.0

# ─── Utilities ───────────────────────────────────────────────────────────────

def now_iso():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def now_stamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def fmt_time(seconds):
    """Format seconds as human-readable string."""
    if seconds < 60:
        return f"{seconds:.0f}s"
    m, s = divmod(int(seconds), 60)
    if m < 60:
        return f"{m}m {s:02d}s"
    h, m = divmod(m, 60)
    return f"{h}h {m:02d}m {s:02d}s"

def fmt_duration(seconds):
    """Format duration as MM:SS."""
    m, s = divmod(int(seconds), 60)
    return f"{m}:{s:02d}"

def fmt_timecode(seconds):
    """Format seconds as HH:MM:SS for display."""
    h, rem = divmod(int(seconds), 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"

def fmt_srt_time(seconds):
    """Format seconds as SRT timecode HH:MM:SS,mmm."""
    h, rem = divmod(seconds, 3600)
    m, rem = divmod(rem, 60)
    s, ms = divmod(rem, 1)
    return f"{int(h):02d}:{int(m):02d}:{int(s):02d},{int(ms * 1000):03d}"

def set_terminal_title(text):
    sys.stdout.write(f"\033]0;{text}\007")
    sys.stdout.flush()

def print_environment(args, device):
    """Print environment info block at startup."""
    import torch
    W = 62
    py_ver = sys.version.split()[0]
    torch_ver = torch.__version__
    device_desc = device.upper()
    if device == "mps":
        device_desc += " (Apple Silicon)"
    elif device == "cuda":
        try:
            device_desc += f" ({torch.cuda.get_device_name(0)})"
        except Exception:
            pass
    try:
        import pyannote.audio
        pa_ver = pyannote.audio.__version__
    except Exception:
        pa_ver = "?"
    os_name = platform.system()
    os_ver = platform.mac_ver()[0] if os_name == "Darwin" else platform.release()
    os_str = f"macOS {os_ver}" if os_name == "Darwin" else f"{os_name} {os_ver}"
    arch = platform.machine()
    venv_name = Path(sys.prefix).name

    print(f"\n  {C.DIM}╭{'─' * W}╮{C.RESET}")
    print(_trow(f"  {C.BOLD}Environment{C.RESET}", W))
    print(_trow(f"  {C.DIM}{'─' * 58}{C.RESET}", W))
    print(_trow(f"  Python:     {py_ver}", W))
    print(_trow(f"  Torch:      {torch_ver} · {device_desc}", W))
    print(_trow(f"  Whisper:    {args.model}", W))
    print(_trow(f"  Pyannote:   {pa_ver}", W))
    print(_trow(f"  Platform:   {os_str} · {arch}", W))
    print(_trow(f"  venv:       {venv_name}", W))
    print(f"  {C.DIM}╰{'─' * W}╯{C.RESET}")

def print_stage(stage_num, total, text):
    print(f"\n  {C.BG_CYAN}{C.WHITE}{C.BOLD} Stage {stage_num}/{total} {C.RESET} {C.BOLD}{text}{C.RESET}")


def print_stage_header(num, total, text, stage_key, est_times, st, str_set):
    """Extended stage header: shows plan estimate + what comes after.

    Example output:
      Stage 3/5  Whisper transcription (word_timestamps)
      ╸ Plan ~11m 24s  ·  After: 4→5→5b (~4s)
    """
    drift = compute_drift(st, est_times)
    plan_raw = est_times.get(stage_key, 0)
    plan_adj = plan_raw * drift

    # Build "After:" list — remaining stages after current
    stage_order = ["1", "2", "3", "4", "5", "5b"]
    try:
        idx = stage_order.index(stage_key)
    except ValueError:
        idx = len(stage_order)
    after_keys = [k for k in stage_order[idx + 1:] if k in str_set]
    if after_keys:
        after_est = sum(est_times.get(k, 0) for k in after_keys) * drift
        after_chain = "→".join(after_keys)
        after_str = f"After: {after_chain} (~{fmt_time(after_est)})"
    else:
        after_str = "Last stage"

    # Line 1: stage header
    print(f"\n  {C.BG_CYAN}{C.WHITE}{C.BOLD} Stage {num}/{total} {C.RESET} {C.BOLD}{text}{C.RESET}")
    # Line 2: plan + after info
    plan_str = f"~{fmt_time(plan_adj)}" if plan_adj > 0 else "—"
    print(f"  {C.DIM}╸ Plan {plan_str}  ·  {after_str}{C.RESET}")

def print_done(text, elapsed):
    print(f"  {C.GREEN}{C.BOLD}✓{C.RESET} {text}  {C.DIM}{fmt_time(elapsed)}{C.RESET}")

def print_warn(text):
    print(f"  {C.YELLOW}⚠{C.RESET} {text}")

def print_error(text):
    print(f"  {C.RED}✗{C.RESET} {text}")


def progress_bar(current, total, width=30, label="", eta_str="", extra=""):
    """Render a Unicode progress bar string (no trailing newline)."""
    if total <= 0:
        frac = 0.0
    else:
        frac = min(current / total, 1.0)
    filled = int(width * frac)
    remainder = width * frac - filled
    BLOCKS = " ▏▎▍▌▋▊▉█"
    bar = "█" * filled
    if filled < width:
        partial_idx = int(remainder * 8)
        bar += BLOCKS[partial_idx]
        bar += "░" * (width - filled - 1)
    pct = frac * 100
    # Color gradient: green when nearly done, yellow mid-way, cyan early
    if pct >= 80:
        bar_color = C.GREEN
    elif pct >= 40:
        bar_color = C.YELLOW
    else:
        bar_color = C.CYAN
    parts = [f"  {label} " if label else "  "]
    parts.append(f"{bar_color}{bar}{C.RESET}")
    parts.append(f" {C.BOLD}{pct:5.1f}%{C.RESET}")
    if eta_str:
        parts.append(f"  {C.CYAN}~{eta_str} left{C.RESET}")
    if extra:
        parts.append(f"  {C.DIM}{extra}{C.RESET}")
    return "".join(parts)


def summary_bar(value, max_value, width=20):
    """Small inline bar for the summary table."""
    if max_value <= 0:
        return "░" * width
    frac = min(value / max_value, 1.0)
    filled = int(width * frac)
    return "█" * filled + "░" * (width - filled)


STAGE_LABELS = [
    ("1", "Extract audio"), ("2", "Diarization"), ("3", "Whisper"),
    ("4", "Speaker map"), ("5", "Generate"),
    ("5b", "Premiere proj"),
]

def print_status_table(st, est_times, str_set, current_stage=None, ctx=None):
    """Print a comprehensive pipeline dashboard after each stage.

    Uses _trow() for guaranteed border alignment on every row.
    ETA shows '~Xm left' instead of ambiguous 'ETA ~Xm'.
    """
    if not est_times:
        return
    all_stages = list(STAGE_LABELS)
    n_total = len(all_stages)
    n_done = sum(1 for k, _ in all_stages if k in st)

    total_actual = sum(st.values())

    # Drift factor: how actual compares to estimates for completed stages
    drift = compute_drift(st, est_times)

    # Remaining time adjusted by drift
    remaining_keys = [k for k, _ in all_stages if k not in st and k in str_set]
    est_remaining = sum(est_times.get(k, 0) for k in remaining_keys)
    eta = est_remaining * drift

    W = 62  # inner visual width

    # ── Header ──
    proj = ctx.get("project_name", "") if ctx else ""
    n_clips = len(ctx.get("video_files", [])) if ctx else 0
    dur_str = fmt_duration(ctx.get("total_duration", 0)) if ctx else ""

    print(f"\n  {C.DIM}┌{'─' * W}┐{C.RESET}")

    # Line 1: project name
    print(_trow(f"  {C.BOLD}{proj}{C.RESET}", W))

    # Line 2: left = clips · duration, right = progress · elapsed · ~left
    left_parts = []
    if n_clips:
        left_parts.append(f"{n_clips} clips")
    if dur_str:
        left_parts.append(dur_str)
    left_str = " · ".join(left_parts)

    done_str = f"{n_done}/{n_total} done"
    elapsed_str = fmt_time(total_actual)
    if remaining_keys:
        left_time = f"~{fmt_time(max(eta, 0))} left"
        right_str = f"{done_str} · {elapsed_str} · {C.CYAN}{left_time}{C.RESET}"
    else:
        right_str = f"{done_str} · {elapsed_str} · {C.GREEN}Done{C.RESET}"

    gap = W - 4 - len(left_str) - vlen(right_str)  # 4 = leading "  " + trailing "  "
    line2 = f"  {left_str}{' ' * max(gap, 2)}{right_str}  "
    print(_trow(line2, W))

    # ── Column headers ──
    print(f"  {C.DIM}├{'─' * W}┤{C.RESET}")
    hdr = f"  {C.DIM}   {'Stage':<16s}  {'Plan':>7s}   {'Actual':>7s}   {'Delta':>7s}{C.RESET}"
    print(_trow(hdr, W))
    print(_trow(f"  {'─' * 58}", W))

    # ── Per-stage rows ──
    for k, label in all_stages:
        e = est_times.get(k, 0)
        plan_str = fmt_time(e)

        if k in st:
            # Completed
            t = st[k]
            delta = t - e
            d_color = C.RED if delta > e * 0.3 else (C.GREEN if delta < -e * 0.3 else C.DIM)
            d_sign = "+" if delta >= 0 else "−"
            delta_txt = f"{d_sign}{fmt_time(abs(delta))}"
            content = (f"  {C.GREEN}✓{C.RESET}  {label:<16s}  {plan_str:>7s}"
                       f"   → {fmt_time(t):>6s}"
                       f"   {d_color}{delta_txt}{C.RESET}")
        elif k == current_stage:
            # Running (about to start)
            content = (f"  {C.CYAN}▶{C.RESET}  {C.BOLD}{label:<16s}{C.RESET}  {plan_str:>7s}"
                       f"   {C.CYAN}running...{C.RESET}")
        elif k in str_set:
            # Pending
            base = f"  {C.DIM}·  {label:<16s}  {plan_str:>7s}{C.RESET}"
            if abs(drift - 1.0) > 0.15:
                adj_str = f"~{fmt_time(e * drift)}"
                gap = W - vlen(base) - len(adj_str) - 2
                content = f"{base}{' ' * max(gap, 2)}{C.DIM}{adj_str}{C.RESET}"
            else:
                content = base
        else:
            # Skipped
            content = f"  {C.DIM}─  {label:<16s}     skip{C.RESET}"

        print(_trow(content, W))

    # ── Footer: progress bar ──
    pct = n_done / n_total * 100 if n_total > 0 else 0
    bar_w = 30
    bar_filled = int(bar_w * pct / 100)
    bar_color = C.GREEN if pct >= 80 else (C.YELLOW if pct >= 40 else C.CYAN)
    bar_str = f"{bar_color}{'█' * bar_filled}{C.RESET}{C.DIM}{'░' * (bar_w - bar_filled)}{C.RESET}"

    if remaining_keys:
        footer_info = f"{fmt_time(total_actual)} / ~{fmt_time(max(eta, 0))} left"
    else:
        footer_info = f"{C.GREEN}{C.BOLD}Done{C.RESET} in {fmt_time(total_actual)}"

    footer = f"  {bar_str}  {C.BOLD}{pct:3.0f}%{C.RESET}   {footer_info}"
    print(f"  {C.DIM}├{'─' * W}┤{C.RESET}")
    print(_trow(footer, W))
    print(f"  {C.DIM}└{'─' * W}┘{C.RESET}")


class Spinner:
    """Animated spinner with elapsed time for long-running operations.

    If plan_est > 0, shows plan alongside elapsed and delta on stop:
      ⠋ Diarizing 54:19  1m 36s / plan ~2m 26s
      ✓ Diarization complete (54:19)  1m 40s  (plan ~2m 26s, −46s)
    """
    FRAMES = "⠋⠙⠹⠸⠼⠴⠦⠧⠇⠏"

    def __init__(self, text="Working", plan_est=0):
        self.text = text
        self.plan_est = plan_est
        self._stop = threading.Event()
        self._thread = None
        self._start_time = 0

    def start(self):
        self._start_time = time.time()
        self._stop.clear()
        self._thread = threading.Thread(target=self._spin, daemon=True)
        self._thread.start()

    def _spin(self):
        i = 0
        while not self._stop.is_set():
            elapsed = time.time() - self._start_time
            frame = self.FRAMES[i % len(self.FRAMES)]
            plan_part = f" / plan ~{fmt_time(self.plan_est)}" if self.plan_est > 0 else ""
            line = (f"  {C.CYAN}{frame}{C.RESET} {self.text}  "
                    f"{C.DIM}{fmt_time(elapsed)}{plan_part}{C.RESET}")
            sys.stdout.write(f"{C.CLEAR_LINE}{line}")
            sys.stdout.flush()
            i += 1
            self._stop.wait(0.1)

    def stop(self, final_text=None):
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=1)
        elapsed = time.time() - self._start_time
        if final_text:
            if self.plan_est > 0:
                delta = elapsed - self.plan_est
                d_sign = "+" if delta >= 0 else "−"
                delta_str = f"  (plan ~{fmt_time(self.plan_est)}, {d_sign}{fmt_time(abs(delta))})"
            else:
                delta_str = ""
            sys.stdout.write(f"{C.CLEAR_LINE}  {C.GREEN}{C.BOLD}✓{C.RESET} {final_text}  "
                             f"{C.DIM}{fmt_time(elapsed)}{delta_str}{C.RESET}\n")
        else:
            sys.stdout.write(f"{C.CLEAR_LINE}")
        sys.stdout.flush()
        return elapsed


class PipelineProgress:
    """Track overall pipeline progress across stages."""
    STAGE_WEIGHTS = {
        "1": 0.03, "2": 0.15, "3": 0.70, "4": 0.001, "5": 0.002,
        "5b": 0.001,
    }

    def __init__(self):
        self.total_weight = sum(self.STAGE_WEIGHTS.values())
        self.completed_weight = 0.0
        self.pipeline_start = time.time()

    def stage_started(self, stage_key):
        """Print pipeline-level progress line before a stage begins."""
        pct = self.completed_weight / self.total_weight * 100
        elapsed = time.time() - self.pipeline_start
        bar = progress_bar(self.completed_weight, self.total_weight, width=40,
                           label=f"{C.MAGENTA}Pipeline{C.RESET}",
                           extra=fmt_time(elapsed))
        print(bar)

    def stage_completed(self, stage_key):
        """Mark a stage as done and accumulate weight."""
        self.completed_weight += self.STAGE_WEIGHTS.get(stage_key, 0)

    def final_bar(self):
        """Print 100% completed bar."""
        elapsed = time.time() - self.pipeline_start
        bar = progress_bar(1, 1, width=40,
                           label=f"{C.MAGENTA}Pipeline{C.RESET}",
                           extra=f"Total: {fmt_time(elapsed)}")
        print(f"\n{bar}")


def get_duration(video_path):
    """Get video/audio duration in seconds via ffprobe."""
    cmd = [
        "ffprobe", "-v", "quiet", "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1", str(video_path)
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    try:
        return float(result.stdout.strip())
    except (ValueError, AttributeError):
        print_error(f"Cannot read duration: {video_path}")
        sys.exit(1)

def get_media_info(video_path):
    """Get full media metadata via ffprobe JSON output."""
    cmd = ["ffprobe", "-v", "quiet", "-print_format", "json",
           "-show_format", "-show_streams", str(video_path)]
    result = subprocess.run(cmd, capture_output=True, text=True)
    try:
        probe = json.loads(result.stdout)
    except json.JSONDecodeError:
        return {"duration": get_duration(video_path)}
    vs = next((s for s in probe.get("streams", []) if s.get("codec_type") == "video"), {})
    as_ = next((s for s in probe.get("streams", []) if s.get("codec_type") == "audio"), {})
    ds = next((s for s in probe.get("streams", []) if s.get("codec_type") == "data"), {})
    fmt = probe.get("format", {})
    # Parse FPS from fraction string like "25/1"
    fps_str = vs.get("r_frame_rate", "0/1")
    try:
        num, den = fps_str.split("/")
        fps = round(int(num) / int(den), 2) if int(den) else 0
    except (ValueError, ZeroDivisionError):
        fps = 0
    return {
        "duration": float(fmt.get("duration", 0)),
        "file_size": int(fmt.get("size", 0)),
        "overall_bitrate": int(fmt.get("bit_rate", 0)) if fmt.get("bit_rate") else 0,
        "creation_time": fmt.get("tags", {}).get("creation_time"),
        "width": int(vs.get("width", 0)),
        "height": int(vs.get("height", 0)),
        "fps": fps,
        "fps_raw": fps_str,
        "nb_frames": int(vs.get("nb_frames", 0)) if vs.get("nb_frames") else 0,
        "video_codec": vs.get("codec_name", ""),
        "video_profile": vs.get("profile", ""),
        "video_bitrate": int(vs.get("bit_rate", 0)) if vs.get("bit_rate") else 0,
        "pix_fmt": vs.get("pix_fmt", ""),
        "bits_per_raw_sample": vs.get("bits_per_raw_sample", ""),
        "color_range": vs.get("color_range", ""),
        "audio_codec": as_.get("codec_name", ""),
        "audio_channels": int(as_.get("channels", 0)) if as_.get("channels") else 0,
        "audio_sample_rate": int(as_.get("sample_rate", 0)) if as_.get("sample_rate") else 0,
        "audio_bits_per_sample": int(as_.get("bits_per_sample", 0)) if as_.get("bits_per_sample") else 0,
        "timecode": ds.get("tags", {}).get("timecode", ""),
    }


def find_videos(folder):
    """Find video files recursively, skip hidden dirs/files."""
    return sorted([
        f for f in folder.rglob('*')
        if f.is_file()
        and f.suffix.lower() in VIDEO_EXTENSIONS
        and not f.name.startswith('.')
        and not any(p.startswith('.') for p in f.relative_to(folder).parts[:-1])
    ])


def get_clip_dir(ctx, clip_id):
    """Get per_clip directory for a clip. Subfolder-aware for multi-camera."""
    groups = ctx.get("subfolder_groups")
    if groups:
        for grp_name, grp in groups.items():
            if clip_id in grp["clip_ids"]:
                td = grp["source_dir"] / f"{grp_name}_transcription"
                return td / "per_clip" / clip_id
    return ctx["transcription_dir"] / "per_clip" / clip_id


def premiere_lang(whisper_lang):
    """Convert Whisper language code to Premiere format."""
    if whisper_lang and whisper_lang in LANG_MAP:
        return LANG_MAP[whisper_lang]
    return "??-??"


# ─── Argparse ────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="YTAI Transcribe v2.16 — Word-level transcription with speaker diarization"
    )
    parser.add_argument("--project", required=True, help="Folder with videos or single video file")
    parser.add_argument("-n", type=int, default=None, help="Number of speakers (auto-detect if omitted)")
    parser.add_argument("-m", "--model", default="large-v3", help="Whisper model (default: large-v3)")
    parser.add_argument("-y", action="store_true", help="Skip confirmations")
    parser.add_argument("--language", default=None, help="Language (default: auto-detect)")
    parser.add_argument("--resume", action="store_true", help="Resume from last completed stage")
    parser.add_argument("--stages", default=None, help="Only run specific stages (e.g. 3,4,5)")
    parser.add_argument("--no-prproj", action="store_true", help="Skip .prproj generation")
    parser.add_argument("--dry-run", action="store_true", help="Show what would run, then exit")
    args = parser.parse_args()

    # Input validation
    if args.n is not None and args.n < 1:
        parser.error("-n must be a positive integer")
    if args.stages:
        valid_stages = {"1", "2", "3", "4", "5", "5b"}
        given = set(args.stages.split(","))
        invalid = given - valid_stages
        if invalid:
            parser.error(f"Invalid stages: {', '.join(sorted(invalid))}. Valid: {', '.join(sorted(valid_stages))}")
    if args.language:
        try:
            import whisper.tokenizer
            supported = whisper.tokenizer.LANGUAGES
            if args.language not in supported and args.language not in supported.values():
                parser.error(f"Unsupported language: {args.language}. See whisper docs for valid codes.")
        except ImportError:
            pass  # Validated at preflight
    return args


# ─── Input Detection ─────────────────────────────────────────────────────────

def detect_input(args):
    """Determine input type and set up paths."""
    input_path = Path(args.project).resolve()

    if input_path.is_file():
        if input_path.suffix.lower() not in VIDEO_EXTENSIONS:
            print_error(f"Not a video file: {input_path.suffix}")
            sys.exit(1)
        video_files = [input_path]
        project_name = input_path.stem
        work_dir = input_path.parent
        subfolder_groups = None
    elif input_path.is_dir():
        # Check for subfolders with videos (multi-camera mode)
        subfolder_groups = {}
        for item in sorted(input_path.iterdir()):
            if item.is_dir() and not item.name.startswith('.'):
                sub_vids = sorted([
                    f for f in item.iterdir()
                    if f.is_file() and f.suffix.lower() in VIDEO_EXTENSIONS
                    and not f.name.startswith('.')
                ])
                if sub_vids:
                    subfolder_groups[item.name] = {
                        "source_dir": item,
                        "clips": sub_vids,
                        "clip_ids": [f.stem for f in sub_vids],
                    }

        if subfolder_groups:
            # Multi-camera mode: collect all videos from subfolders
            video_files = []
            for grp in subfolder_groups.values():
                video_files.extend(grp["clips"])
            video_files.sort()
        else:
            # Flat mode (backward compat): search root only
            video_files = find_videos(input_path)
            subfolder_groups = None

        if not video_files:
            print_error(f"No video files found in {input_path}")
            sys.exit(1)
        # Check for duplicate clip names (different extensions, same stem)
        stems = [f.stem for f in video_files]
        dupes = [s for s in set(stems) if stems.count(s) > 1]
        if dupes:
            print_error(f"Duplicate clip names: {', '.join(sorted(dupes))}")
            sys.exit(1)
        project_name = input_path.name
        work_dir = input_path
    else:
        print_error(f"Path not found: {input_path}")
        sys.exit(1)

    transcription_dir = work_dir / f"{project_name}_transcription"

    return {
        "input_path": input_path,
        "input_type": "file" if input_path.is_file() else "folder",
        "video_files": video_files,
        "project_name": project_name,
        "work_dir": work_dir,
        "transcription_dir": transcription_dir,
        "subfolder_groups": subfolder_groups,
    }


# ─── Preflight ───────────────────────────────────────────────────────────────

def preflight(ctx, args):
    """Run all preflight checks. Returns dict with loaded resources."""
    print(f"\n{'═' * 60}")
    print(f"  {C.BG_BLUE}{C.WHITE}{C.BOLD} Preflight Checks {C.RESET}")

    # 1. ffmpeg
    if shutil.which("ffmpeg") is None:
        print_error("ffmpeg not found"); sys.exit(1)
    if shutil.which("ffprobe") is None:
        print_error("ffprobe not found"); sys.exit(1)
    print(f"  {C.GREEN}✓{C.RESET} ffmpeg")

    # Dry-run fast path: skip ML model loading, only scan media info
    if args.dry_run:
        n_files = len(ctx["video_files"])
        print(f"  {C.GREEN}✓{C.RESET} Input: {ctx['input_type']} — {n_files} file{'s' if n_files > 1 else ''}")
        total_duration = 0.0; clip_durations = {}; clip_media_info = {}
        for vf in ctx["video_files"]:
            info = get_media_info(vf)
            dur = info["duration"]; clip_durations[vf.stem] = dur
            clip_media_info[vf.stem] = info; total_duration += dur
        ctx["clip_durations"] = clip_durations
        ctx["clip_media_info"] = clip_media_info
        ctx["total_duration"] = total_duration
        td = total_duration
        ctx["stage_estimates"] = {
            "1": td * SPEED_COEFFICIENTS["extract_audio"],
            "2": td * SPEED_COEFFICIENTS["diarization"],
            "3": td * SPEED_COEFFICIENTS["whisper"],
            "4": td * SPEED_COEFFICIENTS["speaker_map"],
            "5": td * SPEED_COEFFICIENTS["generate"],
            "5b": td * SPEED_COEFFICIENTS["prproj"],
        }
        first_info = clip_media_info.get(ctx["video_files"][0].stem, {})
        res_str = f"{first_info.get('width', '?')}×{first_info.get('height', '?')}"
        fps_str = f"{first_info.get('fps', '?')}fps"
        codec_str = first_info.get("video_codec", "?").upper()
        print(f"  {C.GREEN}✓{C.RESET} Total duration: {fmt_duration(total_duration)} ({total_duration:.1f}s)")
        print(f"  {C.GREEN}✓{C.RESET} Media: {res_str} {fps_str} {codec_str}")
        return {"has_template": False}

    # 2. Python packages
    try:
        import whisper
        import pyannote.audio
        import torch
        import soundfile
        import openpyxl
    except ImportError as e:
        print_error(f"Missing package: {e}"); sys.exit(1)
    print(f"  {C.GREEN}✓{C.RESET} Python packages")

    # 3. Whisper model
    print(f"  {C.DIM}Loading Whisper {args.model}...{C.RESET}", end="", flush=True)
    model = whisper.load_model(args.model)
    print(f"\r  {C.GREEN}✓{C.RESET} Whisper model: {args.model}          ")

    # 4. HuggingFace token
    hf_token = None
    hf_paths = [
        Path.home() / ".huggingface" / "token",
        Path.home() / ".cache" / "huggingface" / "token",
    ]
    for p in hf_paths:
        if p.exists():
            hf_token = p.read_text().strip()
            break
    if hf_token is None:
        # Try environment
        hf_token = os.environ.get("HF_TOKEN") or os.environ.get("HUGGING_FACE_HUB_TOKEN")
    if hf_token is None:
        print_error("HuggingFace token not found"); sys.exit(1)
    print(f"  {C.GREEN}✓{C.RESET} HuggingFace token")

    # 5. Torch device
    if torch.backends.mps.is_available():
        device = "mps"
    elif torch.cuda.is_available():
        device = "cuda"
    else:
        device = "cpu"
    print(f"  {C.GREEN}✓{C.RESET} Device: {device}")

    # Environment info block
    print_environment(args, device)

    # 6. Input path
    n_files = len(ctx["video_files"])
    print(f"  {C.GREEN}✓{C.RESET} Input: {ctx['input_type']} — {n_files} file{'s' if n_files > 1 else ''}")

    # 7. Video files — duration + media metadata
    total_duration = 0.0
    clip_durations = {}
    clip_media_info = {}
    for vf in ctx["video_files"]:
        info = get_media_info(vf)
        dur = info["duration"]
        clip_durations[vf.stem] = dur
        clip_media_info[vf.stem] = info
        total_duration += dur
    ctx["clip_durations"] = clip_durations
    ctx["clip_media_info"] = clip_media_info
    ctx["total_duration"] = total_duration
    # Show resolution/codec from first clip
    first_info = clip_media_info.get(ctx["video_files"][0].stem, {})
    res_str = f"{first_info.get('width', '?')}×{first_info.get('height', '?')}"
    fps_str = f"{first_info.get('fps', '?')}fps"
    codec_str = first_info.get("video_codec", "?").upper()
    print(f"  {C.GREEN}✓{C.RESET} Total duration: {fmt_duration(total_duration)} ({total_duration:.1f}s)")
    print(f"  {C.GREEN}✓{C.RESET} Media: {res_str} {fps_str} {codec_str}")

    # 8. Disk space
    stat = os.statvfs(str(ctx["work_dir"]))
    free_gb = (stat.f_bavail * stat.f_frsize) / (1024**3)
    # Rough estimate: audio ~10MB/min, need 2x
    needed_gb = (total_duration / 60) * 10 / 1024 * 2
    if free_gb < needed_gb:
        print_warn(f"Low disk space: {free_gb:.1f}GB free, ~{needed_gb:.1f}GB needed")
    else:
        print(f"  {C.GREEN}✓{C.RESET} Disk space: {free_gb:.1f}GB free")

    # 9. Previous run
    if ctx["transcription_dir"].exists():
        print_warn(f"Previous run exists: {ctx['transcription_dir'].name}/")
        if not args.y and not args.resume and not args.dry_run:
            resp = input("    Overwrite? [y/N] ")
            if resp.lower() != 'y':
                print("Aborted."); sys.exit(0)
    else:
        print(f"  {C.GREEN}✓{C.RESET} No previous run")

    # 10. Premiere template
    has_template = TEMPLATE_PATH.exists()
    if has_template:
        print(f"  {C.GREEN}✓{C.RESET} Premiere template")
    elif not args.no_prproj:
        print_warn("Premiere template not found, will skip .prproj generation")

    # Per-stage time estimates based on audio duration
    stage_estimates = {
        "1": total_duration * SPEED_COEFFICIENTS["extract_audio"],
        "2": total_duration * SPEED_COEFFICIENTS["diarization"],
        "3": total_duration * SPEED_COEFFICIENTS["whisper"],
        "4": total_duration * SPEED_COEFFICIENTS["speaker_map"],
        "5": total_duration * SPEED_COEFFICIENTS["generate"],
        "5b": total_duration * SPEED_COEFFICIENTS["prproj"],
    }
    ctx["stage_estimates"] = stage_estimates
    est = sum(stage_estimates.values())
    est_min = est / 60
    est_range = f"~{est_min * 0.85:.0f}-{est_min * 1.15:.0f} min"
    print(f"\n{'═' * 60}")
    print(f"  {C.BG_CYAN}{C.WHITE}{C.BOLD} YTAI Transcribe v{VERSION} {C.RESET}")
    print(f"{'─' * 60}")
    print(f"  Project:    {C.BOLD}{ctx['project_name']}{C.RESET}")
    groups = ctx.get("subfolder_groups")
    if groups:
        cam_parts = [f"{name} ({len(grp['clip_ids'])})" for name, grp in groups.items()]
        print(f"  Cameras:    {' + '.join(cam_parts)} = {n_files} clips")
    print(f"  Files:      {n_files} | {fmt_duration(total_duration)} | {args.n or 'auto'} speakers")
    print(f"  Model:      {args.model} | {device.upper()}")
    print(f"  Estimated:  {C.YELLOW}{est_range}{C.RESET}")
    print(f"{'═' * 60}")

    if not args.y and not args.dry_run:
        resp = input("\nProceed? [Y/n] ")
        if resp.lower() == 'n':
            print("Aborted."); sys.exit(0)

    return {
        "whisper_model": model,
        "hf_token": hf_token,
        "device": device,
        "has_template": has_template,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 1: Extract Audio
# ═══════════════════════════════════════════════════════════════════════════════

def stage1_extract_audio(ctx, args):
    """Extract per-clip WAV (16kHz mono) and concatenate to full_audio.wav."""
    import soundfile as sf

    t0 = time.time()
    td = ctx["transcription_dir"]
    per_clip_dir = td / "per_clip"
    clips = ctx["video_files"]

    clip_offsets_data = []
    all_audio = []
    current_offset = 0.0
    total_input_dur = sum(ctx["clip_durations"].get(vf.stem, 0) for vf in clips)
    processed_dur = 0.0

    for i, vf in enumerate(clips):
        clip_id = vf.stem
        clip_dir = get_clip_dir(ctx, clip_id)
        clip_dir.mkdir(parents=True, exist_ok=True)
        wav_path = clip_dir / f"{clip_id}_audio.wav"

        set_terminal_title(f"YTAI > Stage 1/5 > {clip_id} ({i+1}/{len(clips)})")

        # Show progress bar before starting this clip
        if len(clips) > 1:
            eta_str = ""
            if i > 0:
                elapsed_so_far = time.time() - t0
                speed = processed_dur / elapsed_so_far if elapsed_so_far > 0 else 1
                remaining = total_input_dur - processed_dur
                eta_str = fmt_time(remaining / speed) if speed > 0 else ""
            plan_adj = ctx.get("_stage_plan_adj", 0)
            extra = f"plan ~{fmt_time(plan_adj)}" if plan_adj > 0 else ""
            sys.stdout.write(C.CLEAR_LINE + progress_bar(
                processed_dur, total_input_dur, width=30,
                label=f"[{i+1}/{len(clips)}]", eta_str=eta_str, extra=extra))
            sys.stdout.flush()

        # Extract audio
        t1 = time.time()
        cmd = [
            "ffmpeg", "-y", "-i", str(vf),
            "-ar", "16000", "-ac", "1", "-vn",
            "-loglevel", "error",
            str(wav_path)
        ]
        ffmpeg_result = subprocess.run(cmd, capture_output=True, text=True)
        if ffmpeg_result.returncode != 0:
            print_error(f"{clip_id}: ffmpeg failed — {ffmpeg_result.stderr[:200].strip()}")
            sys.exit(1)

        # Read for concatenation
        data, sr = sf.read(str(wav_path))
        duration = len(data) / sr

        clip_offsets_data.append({
            "clip_id": clip_id,
            "file": vf.name,
            "offset": round(current_offset, 3),
            "duration": round(duration, 3),
        })

        all_audio.append(data)
        current_offset = round(current_offset + duration, 3)
        processed_dur += duration

        elapsed = time.time() - t1
        # Clear progress bar, print final result for this clip
        if len(clips) > 1:
            sys.stdout.write(C.CLEAR_LINE)
            sys.stdout.flush()
        print(f"  {clip_id} -> {clip_id}_audio.wav  {fmt_duration(duration):>6s}  "
              f"{C.GREEN}done{C.RESET}  {fmt_time(elapsed):>7s}")

    # Concatenate
    if len(clips) > 1:
        print(f"  Concatenating {len(clips)} clips -> full_audio.wav")
        combined = np.concatenate(all_audio)
    else:
        combined = all_audio[0]

    full_audio_path = td / "full_audio.wav"
    sf.write(str(full_audio_path), combined, 16000)

    total_duration = current_offset

    # Save clip_offsets.json
    clip_offsets_json = {
        "version": VERSION,
        "created_at": now_iso(),
        "clips": clip_offsets_data,
        "total_duration": round(total_duration, 3),
    }
    with open(td / "clip_offsets.json", "w") as f:
        json.dump(clip_offsets_json, f, indent=2, ensure_ascii=False)

    ctx["clip_offsets"] = clip_offsets_data
    ctx["clip_offsets_flat"] = {c["clip_id"]: c["offset"] for c in clip_offsets_data}
    ctx["total_duration"] = total_duration

    elapsed = time.time() - t0
    print_done(f"Stage 1 complete — {len(clips)} clips, {fmt_duration(total_duration)}", elapsed)
    return elapsed


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 2: Global Diarization
# ═══════════════════════════════════════════════════════════════════════════════

def stage2_diarization(ctx, args, resources):
    """Run pyannote speaker diarization on full_audio.wav."""
    from pyannote.audio import Pipeline as PyannotePipeline

    t0 = time.time()
    td = ctx["transcription_dir"]
    full_audio = td / "full_audio.wav"

    set_terminal_title(f"YTAI > Stage 2/5 > Diarization ({fmt_duration(ctx['total_duration'])})")

    pipeline = PyannotePipeline.from_pretrained(
        "pyannote/speaker-diarization-3.1",
        token=resources["hf_token"]
    )

    # Move to device if possible
    import torch
    import soundfile as sf
    if resources["device"] == "mps":
        pipeline = pipeline.to(torch.device("mps"))
    elif resources["device"] == "cuda":
        pipeline.to(torch.device("cuda"))

    # Load audio in-memory to bypass torchcodec (pyannote 4.x + torchcodec
    # fails to find FFmpeg shared libs when given a file path)
    audio_data, sample_rate = sf.read(str(full_audio))
    waveform = torch.tensor(audio_data, dtype=torch.float32).unsqueeze(0)
    if waveform.dim() == 3:
        waveform = waveform.mean(dim=-1)  # mono mixdown if stereo

    diar_kwargs = {}
    if args.n is not None:
        diar_kwargs["num_speakers"] = args.n

    plan_est = ctx.get("_stage_plan_adj", 0)
    spinner = Spinner(f"Diarizing {fmt_duration(ctx['total_duration'])} of audio",
                      plan_est=plan_est)
    spinner.start()
    diarization = pipeline({"waveform": waveform, "sample_rate": sample_rate}, **diar_kwargs)
    spinner.stop(f"Diarization complete ({fmt_duration(ctx['total_duration'])})")

    # Convert to segments list
    # pyannote 4.x returns DiarizeOutput dataclass; extract Annotation
    annotation = getattr(diarization, "speaker_diarization", diarization)
    segments = []
    for turn, _, speaker in annotation.itertracks(yield_label=True):
        segments.append({
            "start": round(turn.start, 3),
            "end": round(turn.end, 3),
            "speaker": speaker,
        })

    # Build speaker map with UUIDs
    unique_speakers = sorted(set(s["speaker"] for s in segments))
    speaker_map = {}
    for i, label in enumerate(unique_speakers):
        speaker_map[label] = {
            "id": str(uuid.uuid4()),
            "name": f"Speaker {i + 1}",
        }

    # Stats
    speaker_stats = {}
    for label in unique_speakers:
        sp_segs = [s for s in segments if s["speaker"] == label]
        sp_dur = sum(s["end"] - s["start"] for s in sp_segs)
        speaker_stats[label] = sp_dur

    total_speech = sum(speaker_stats.values())
    gap_dur = ctx["total_duration"] - total_speech

    print(f"  Found: {len(unique_speakers)} speakers, {len(segments)} segments")
    sp_parts = []
    for label in unique_speakers:
        dur = speaker_stats[label]
        pct = dur / ctx["total_duration"] * 100
        sp_parts.append(f"{speaker_map[label]['name']}: {fmt_duration(dur)} ({pct:.1f}%)")
    gap_pct = gap_dur / ctx["total_duration"] * 100
    print(f"  {'  |  '.join(sp_parts)}  |  Gaps: {fmt_duration(gap_dur)} ({gap_pct:.1f}%)")

    # Save diarization.json
    diar_json = {
        "version": VERSION,
        "created_at": now_iso(),
        "num_speakers": len(unique_speakers),
        "total_duration": round(ctx["total_duration"], 3),
        "segments": segments,
    }
    with open(td / "diarization.json", "w") as f:
        json.dump(diar_json, f, indent=2, ensure_ascii=False)

    # Save speakers.json
    speakers_json = {
        "version": VERSION,
        "created_at": now_iso(),
        "speakers": speaker_map,
    }
    with open(td / "speakers.json", "w") as f:
        json.dump(speakers_json, f, indent=2, ensure_ascii=False)

    ctx["diarization_segments"] = segments
    ctx["speaker_map"] = speaker_map

    elapsed = time.time() - t0
    print_done("Stage 2 complete", elapsed)
    return elapsed


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 3: Per-clip Whisper
# ═══════════════════════════════════════════════════════════════════════════════

def stage3_whisper(ctx, args, resources):
    """Transcribe each clip with Whisper word_timestamps=True."""
    t0 = time.time()
    model = resources["whisper_model"]
    td = ctx["transcription_dir"]
    clips = ctx["video_files"]
    all_clip_results = {}
    total_words = 0
    detected_langs = []
    total_audio_dur = sum(ctx["clip_durations"].get(vf.stem, 0) for vf in clips)
    processed_audio_dur = 0.0

    # Output filter classes (defined once, used per clip)
    class _FilterStdout:
        """Pass through all writes except 'Detected language' messages."""
        def __init__(self, orig): self._orig = orig
        def write(self, text):
            if text and "Detected language" not in text:
                self._orig.write(text)
        def flush(self): self._orig.flush()
        def __getattr__(self, name): return getattr(self._orig, name)

    class _FilterStderr:
        """Suppress tqdm progress bars from Whisper."""
        def __init__(self, orig): self._orig = orig
        def write(self, text):
            if text and "\r" not in text and "%" not in text:
                self._orig.write(text)
        def flush(self): self._orig.flush()
        def __getattr__(self, name): return getattr(self._orig, name)

    failed_clips = []

    for i, vf in enumerate(clips):
        clip_id = vf.stem
        clip_dir = get_clip_dir(ctx, clip_id)
        wav_path = clip_dir / f"{clip_id}_audio.wav"
        checkpoint_path = clip_dir / f"{clip_id}_whisper_raw.json"

        set_terminal_title(f"YTAI > Stage 3/5 > {clip_id} ({i+1}/{len(clips)})")

        dur = ctx["clip_durations"].get(clip_id, 0)

        # Check for existing checkpoint (resume support)
        if checkpoint_path.exists():
            try:
                with open(checkpoint_path) as f:
                    result = json.load(f)
                clip_words = sum(len(s.get("words", [])) for s in result.get("segments", []))
                total_words += clip_words
                detected_langs.append(result.get("language", "unknown"))
                processed_audio_dur += dur
                all_clip_results[clip_id] = result
                if len(clips) > 1 and i > 0:
                    sys.stdout.write(C.CLEAR_LINE)
                    sys.stdout.flush()
                print(f"  [{i+1:2d}/{len(clips)}] {clip_id}  {fmt_duration(dur):>6s}  "
                      f"{clip_words:>5} words  {C.CYAN}cached{C.RESET}")
                continue
            except (json.JSONDecodeError, KeyError):
                pass  # Corrupted checkpoint, re-transcribe

        spinner = Spinner(f"[{i+1:2d}/{len(clips)}] {clip_id}  {fmt_duration(dur):>6s}  transcribing")
        spinner.start()

        t1 = time.time()
        # Suppress Whisper's "Detected language" (stdout) and tqdm (stderr)
        _real_stdout = sys.stdout
        _real_stderr = sys.stderr
        sys.stdout = _FilterStdout(_real_stdout)
        sys.stderr = _FilterStderr(_real_stderr)
        try:
            result = model.transcribe(
                str(wav_path),
                word_timestamps=True,
                language=args.language,
                verbose=False,
            )
        except Exception as e:
            sys.stdout = _real_stdout
            sys.stderr = _real_stderr
            spinner.stop()
            print_error(f"{clip_id}: Whisper failed — {e}")
            result = {"segments": [], "language": "unknown", "text": ""}
            failed_clips.append(clip_id)
        else:
            sys.stdout = _real_stdout
            sys.stderr = _real_stderr

        spinner.stop()

        # Save checkpoint for resume
        try:
            with open(checkpoint_path, "w") as f:
                json.dump(result, f, ensure_ascii=False)
        except Exception:
            pass  # Non-critical

        # Count words
        clip_words = 0
        for seg in result.get("segments", []):
            clip_words += len(seg.get("words", []))
        total_words += clip_words

        lang = result.get("language", "unknown")
        detected_langs.append(lang)

        elapsed_clip = time.time() - t1
        speed = dur / elapsed_clip if elapsed_clip > 0 else 0
        processed_audio_dur += dur

        # Clear previous progress bar if present
        if len(clips) > 1 and i > 0:
            sys.stdout.write(C.CLEAR_LINE)
            sys.stdout.flush()

        status = f"{C.GREEN}done{C.RESET}" if clip_id not in failed_clips else f"{C.RED}FAILED{C.RESET}"
        print(f"  [{i+1:2d}/{len(clips)}] {clip_id}  {fmt_duration(dur):>6s}  "
              f"{clip_words:>5} words  {status}  {fmt_time(elapsed_clip):>7s}  "
              f"{C.DIM}({speed:>5.2f}x){C.RESET}")

        all_clip_results[clip_id] = result

        # Continuous progress bar with ETA after every clip
        if len(clips) > 1:
            elapsed_total = time.time() - t0
            avg_speed = processed_audio_dur / elapsed_total if elapsed_total > 0 else 1
            remaining_dur = total_audio_dur - processed_audio_dur
            eta_sec = remaining_dur / avg_speed if avg_speed > 0 else 0
            pct = processed_audio_dur / total_audio_dur * 100 if total_audio_dur > 0 else 100

            plan_adj = ctx.get("_stage_plan_adj", 0)
            plan_part = f"plan ~{fmt_time(plan_adj)}" if plan_adj > 0 else ""
            extra_parts = [f"{avg_speed:.2f}x"]
            if plan_part:
                extra_parts.append(plan_part)
            bar_line = progress_bar(
                processed_audio_dur, total_audio_dur, width=30,
                label="Stage 3",
                eta_str=fmt_time(eta_sec) if remaining_dur > 0 else "",
                extra=" · ".join(extra_parts)
            )
            sys.stdout.write(C.CLEAR_LINE + bar_line)
            sys.stdout.flush()

            set_terminal_title(f"YTAI > Stage 3/5 > {pct:.0f}% > ~{fmt_time(eta_sec)} left")

    # Clear the last progress bar line
    if len(clips) > 1:
        sys.stdout.write(C.CLEAR_LINE)
        sys.stdout.flush()

    # Determine main language
    from collections import Counter
    lang_counts = Counter(detected_langs)
    main_lang = lang_counts.most_common(1)[0][0] if lang_counts else "unknown"
    lang_pct = lang_counts[main_lang] / len(detected_langs) * 100 if detected_langs else 0

    print(f"  Total words: {total_words:,} | Language: {main_lang} ({lang_pct:.1f}%)")

    ctx["whisper_results"] = all_clip_results
    ctx["main_language"] = main_lang
    ctx["total_words"] = total_words

    elapsed = time.time() - t0
    print_done("Stage 3 complete", elapsed)
    return elapsed


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 4: Speaker Mapping
# ═══════════════════════════════════════════════════════════════════════════════

def find_nearest_speaker(global_time, diar_segments):
    """Find nearest diarization segment for a time that falls in a gap."""
    best = None
    best_dist = float('inf')
    for seg in diar_segments:
        mid = (seg["start"] + seg["end"]) / 2
        dist = abs(global_time - mid)
        if dist < best_dist:
            best_dist = dist
            best = seg["speaker"]
    return best


def assign_speaker(word_start, word_end, clip_id, clip_offsets_flat, diar_segments):
    """Assign speaker to a word based on global timecode."""
    global_mid = clip_offsets_flat[clip_id] + (word_start + word_end) / 2
    for seg in diar_segments:
        if seg["start"] <= global_mid <= seg["end"]:
            return seg["speaker"]
    return find_nearest_speaker(global_mid, diar_segments)


def group_words_into_segments(words_with_speakers):
    """Group consecutive words by speaker, splitting on speaker change or long pause."""
    if not words_with_speakers:
        return []

    segments = []
    current = {
        "speaker": words_with_speakers[0]["speaker"],
        "words": [words_with_speakers[0]],
        "start": words_with_speakers[0]["start"],
        "end": words_with_speakers[0]["end"],
    }

    for w in words_with_speakers[1:]:
        pause = w["start"] - current["end"]
        same_speaker = w["speaker"] == current["speaker"]

        if not same_speaker or pause > PAUSE_SPLIT_THRESHOLD:
            # Close current segment
            segments.append(current)
            current = {
                "speaker": w["speaker"],
                "words": [w],
                "start": w["start"],
                "end": w["end"],
            }
        else:
            current["words"].append(w)
            current["end"] = w["end"]

    segments.append(current)

    # Split long segments (>30s) at pauses
    final_segments = []
    for seg in segments:
        if seg["end"] - seg["start"] > 30.0 and len(seg["words"]) > 1:
            final_segments.extend(split_long_segment(seg))
        else:
            final_segments.append(seg)

    return final_segments


def split_long_segment(seg):
    """Split a >30s segment at pauses > LONG_SEGMENT_PAUSE."""
    result = []
    current_words = [seg["words"][0]]

    for w in seg["words"][1:]:
        pause = w["start"] - current_words[-1]["end"]
        if pause > LONG_SEGMENT_PAUSE:
            result.append({
                "speaker": seg["speaker"],
                "words": current_words,
                "start": current_words[0]["start"],
                "end": current_words[-1]["end"],
            })
            current_words = [w]
        else:
            current_words.append(w)

    if current_words:
        result.append({
            "speaker": seg["speaker"],
            "words": current_words,
            "start": current_words[0]["start"],
            "end": current_words[-1]["end"],
        })

    # If no splits happened, return original
    return result if len(result) > 1 else [seg]


def stage4_speaker_mapping(ctx, args):
    """Assign speakers to Whisper words via diarization, group into segments."""
    t0 = time.time()
    clips = ctx["video_files"]
    diar_segments = ctx["diarization_segments"]
    offsets = ctx["clip_offsets_flat"]
    whisper_results = ctx["whisper_results"]

    set_terminal_title("YTAI > Stage 4/5 > Mapping speakers")

    plan_est = ctx.get("_stage_plan_adj", 0)
    spinner = Spinner(f"Mapping speakers for {len(clips)} clip{'s' if len(clips) > 1 else ''}",
                      plan_est=plan_est)
    spinner.start()

    all_clip_transcripts = {}
    total_segments = 0
    gap_words = 0

    for vf in clips:
        clip_id = vf.stem
        result = whisper_results[clip_id]

        # Flatten all words with timing
        words_with_speakers = []
        for seg in result.get("segments", []):
            for w in seg.get("words", []):
                if "start" not in w or "end" not in w:
                    continue
                speaker = assign_speaker(
                    w["start"], w["end"], clip_id, offsets, diar_segments
                )
                if speaker is None:
                    gap_words += 1
                    speaker = find_nearest_speaker(
                        offsets[clip_id] + (w["start"] + w["end"]) / 2,
                        diar_segments
                    )
                words_with_speakers.append({
                    "word": w.get("word", "").strip(),
                    "start": round(w["start"], 3),
                    "end": round(w["end"], 3),
                    "confidence": round(w.get("probability", 0.0), 4),
                    "speaker": speaker,
                })

        # Group into segments
        grouped = group_words_into_segments(words_with_speakers)

        # Build transcript segments
        transcript_segments = []
        for idx, seg in enumerate(grouped):
            text = " ".join(w["word"] for w in seg["words"])
            confidences = [w["confidence"] for w in seg["words"]]
            avg_conf = sum(confidences) / len(confidences) if confidences else 0.0
            transcript_segments.append({
                "id": idx,
                "start": seg["start"],
                "end": seg["end"],
                "speaker": seg["speaker"],
                "text": text,
                "words": [{
                    "word": w["word"],
                    "start": w["start"],
                    "end": w["end"],
                    "confidence": w["confidence"],
                } for w in seg["words"]],
                "avg_confidence": round(avg_conf, 4),
                "low_confidence": bool(avg_conf < LOW_CONFIDENCE_THRESHOLD),
            })

        total_segments += len(transcript_segments)

        # Per-clip stats
        clip_duration = ctx["clip_durations"].get(clip_id, 0)
        clip_offset = offsets.get(clip_id, 0.0)
        all_words = [w for s in transcript_segments for w in s["words"]]

        speaker_stats = {}
        for s in transcript_segments:
            sp = s["speaker"]
            if sp not in speaker_stats:
                speaker_stats[sp] = {"segments": 0, "words": 0, "duration": 0.0}
            speaker_stats[sp]["segments"] += 1
            speaker_stats[sp]["words"] += len(s["words"])
            speaker_stats[sp]["duration"] += s["end"] - s["start"]

        # Round stats
        for sp in speaker_stats:
            speaker_stats[sp]["duration"] = round(speaker_stats[sp]["duration"], 3)

        clip_transcript = {
            "version": VERSION,
            "created_at": now_iso(),
            "whisper_model": args.model,
            "clip_id": clip_id,
            "clip_file": vf.name,
            "duration": round(clip_duration, 3),
            "clip_offset_in_full": round(clip_offset, 3),
            "language": whisper_results[clip_id].get("language", "unknown"),
            "segments": transcript_segments,
            "stats": {
                "total_words": len(all_words),
                "total_segments": len(transcript_segments),
                "speakers": speaker_stats,
                "avg_confidence": round(
                    sum(w["confidence"] for w in all_words) / len(all_words) if all_words else 0, 4
                ),
                "low_confidence_segments": sum(1 for s in transcript_segments if s["low_confidence"]),
            },
        }

        all_clip_transcripts[clip_id] = clip_transcript

    ctx["clip_transcripts"] = all_clip_transcripts

    total_words = sum(t["stats"]["total_words"] for t in all_clip_transcripts.values())
    spinner.stop(f"{total_words:,} words | {total_segments} segments | {gap_words} gap words resolved")

    elapsed = time.time() - t0
    print_done("Stage 4 complete", elapsed)
    return elapsed


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 5: Generate Outputs
# ═══════════════════════════════════════════════════════════════════════════════

def split_word_punctuation(whisper_word):
    text = whisper_word["word"].strip()
    s_ms = int(whisper_word["start"] * 1000)
    e_ms = int(whisper_word["end"] * 1000)
    dur_ms = e_ms - s_ms
    conf = whisper_word.get("confidence", whisper_word.get("probability", 0.0))
    match = PUNCT_PATTERN.match(text)
    if match and match.group(1):
        cw, punct = match.group(1), match.group(2)
        return [
            {"text": cw, "start": s_ms, "duration": dur_ms, "confidence": conf,
             "eos": any(c in EOS_CHARS for c in punct), "tags": [], "type": "word"},
            {"text": punct, "start": e_ms, "duration": 0,
             "confidence": 1.0, "eos": False, "tags": [], "type": "punctuation"},
        ]
    elif match and not match.group(1):
        return [{"text": match.group(2), "start": s_ms, "duration": dur_ms,
                 "confidence": conf, "eos": False, "tags": [], "type": "punctuation"}]
    else:
        return [{"text": text, "start": s_ms, "duration": dur_ms,
                 "confidence": conf, "eos": False, "tags": [], "type": "word"}]


def generate_premiere_json(clip_transcript, speaker_map, language):
    prem_lang = premiere_lang(language)
    prem_segments = []
    for seg in clip_transcript["segments"]:
        sid = speaker_map.get(seg["speaker"], {}).get("id", str(uuid.uuid4()))
        prem_words = []
        for w in seg["words"]:
            prem_words.extend(split_word_punctuation(w))
        prem_segments.append({"language": prem_lang, "speakerId": sid, "words": prem_words})
    speakers_list = [{"id": v["id"], "name": v["name"]} for v in speaker_map.values()]
    return {"language": prem_lang, "segments": prem_segments, "speakers": speakers_list}


def generate_srt(clip_transcript, speaker_map):
    lines = []
    for i, seg in enumerate(clip_transcript["segments"], 1):
        name = speaker_map.get(seg["speaker"], {}).get("name", seg["speaker"])
        lines += [str(i), f"{fmt_srt_time(seg['start'])} --> {fmt_srt_time(seg['end'])}",
                  f"[{name}] {seg['text']}", ""]
    return "\n".join(lines)


def generate_combined_srt(ctx, clip_filter=None, output_path=None):
    """Generate combined SRT for the whole project (or filtered clips)."""
    smap = ctx["speaker_map"]
    offsets = ctx["clip_offsets_flat"]
    video_files = ctx["video_files"]
    if clip_filter:
        video_files = [vf for vf in video_files if vf.stem in clip_filter]
    lines = []; idx = 1
    for vf in video_files:
        cid = vf.stem
        tr = ctx["clip_transcripts"].get(cid)
        if not tr:
            continue
        co = offsets.get(cid, 0.0)
        for seg in tr["segments"]:
            name = smap.get(seg["speaker"], {}).get("name", seg["speaker"])
            gs = co + seg["start"]
            ge = co + seg["end"]
            lines += [str(idx), f"{fmt_srt_time(gs)} --> {fmt_srt_time(ge)}",
                      f"[{name}] {seg['text']}", ""]
            idx += 1
    srt_path = output_path or (ctx["work_dir"] / f"{ctx['project_name']}_transcript.srt")
    with open(srt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return srt_path


def generate_txt(clip_transcript, speaker_map):
    lines = []
    for seg in clip_transcript["segments"]:
        name = speaker_map.get(seg["speaker"], {}).get("name", seg["speaker"])
        lines += [f"[{fmt_timecode(seg['start'])}] {name}:", f"  {seg['text']}", ""]
    return "\n".join(lines)



def generate_xlsx(ctx, args, clip_filter=None, output_path=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    video_files = ctx["video_files"]
    if clip_filter:
        video_files = [vf for vf in video_files if vf.stem in clip_filter]
    wb = openpyxl.Workbook()

    # ── Sheet 1: Full Transcript (grouped by clip) ──
    ws = wb.active
    ws.title = "Transcript"
    headers = ["Timecode", "#", "Clip", "Start", "End", "Duration",
               "Words", "Speaker", "Text", "Confidence"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    yf = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    clip_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    clip_font = Font(bold=True, size=11, color="1F3864")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2; seg_n = 0
    offsets = ctx["clip_offsets_flat"]; smap = ctx["speaker_map"]

    for clip_idx, vf in enumerate(video_files):
        cid = vf.stem
        tr = ctx["clip_transcripts"].get(cid)
        if not tr:
            continue
        co = offsets.get(cid, 0.0)
        clip_dur = tr.get("duration", 0)
        clip_words = tr.get("stats", {}).get("total_words", 0)
        clip_segs = tr.get("stats", {}).get("total_segments", 0)

        # Clip separator row
        sep_text = (f"{cid}  |  {fmt_duration(clip_dur)}  |  "
                    f"{clip_words} words  |  {clip_segs} segments")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = clip_fill
        ws.cell(row=row, column=1, value=sep_text).font = clip_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1

        if not tr["segments"]:
            seg_n += 1
            ws.cell(row=row, column=1, value=fmt_timecode(co))
            ws.cell(row=row, column=2, value=seg_n)
            ws.cell(row=row, column=3, value=cid)
            ws.cell(row=row, column=9, value="[no speech]")
            row += 1
            continue

        for seg in tr["segments"]:
            seg_n += 1; gs = co + seg["start"]
            sn = smap.get(seg["speaker"], {}).get("name", seg["speaker"])
            ws.cell(row=row, column=1, value=fmt_timecode(gs))
            ws.cell(row=row, column=2, value=seg_n)
            ws.cell(row=row, column=3, value=cid)
            ws.cell(row=row, column=4, value=round(seg["start"], 2))
            ws.cell(row=row, column=5, value=round(seg["end"], 2))
            ws.cell(row=row, column=6, value=round(seg["end"] - seg["start"], 2))
            ws.cell(row=row, column=7, value=len(seg["text"].split()))
            ws.cell(row=row, column=8, value=sn)
            tc = ws.cell(row=row, column=9, value=seg["text"])
            tc.alignment = wrap_align
            ws.cell(row=row, column=10, value=round(seg["avg_confidence"], 3))
            if seg["low_confidence"]:
                tc.fill = yf
            # Thin bottom border on each row
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

    ws.freeze_panes = "A2"
    # Column widths: Timecode, #, Clip, Start, End, Duration, Words, Speaker, Text, Confidence
    col_widths = {"A": 12, "B": 5, "C": 10, "D": 8, "E": 8, "F": 10,
                  "G": 7, "H": 12, "I": 80, "J": 12}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    # ── Sheet 2: Summary (per-clip stats + timing) ──
    ws2 = wb.create_sheet("Summary")
    s_hf = Font(bold=True, color="FFFFFF", size=11)
    s_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    # Project info
    ws2.cell(row=1, column=1, value="Project").font = Font(bold=True)
    ws2.cell(row=1, column=2, value=ctx["project_name"])
    ws2.cell(row=2, column=1, value="Version").font = Font(bold=True)
    ws2.cell(row=2, column=2, value=VERSION)
    ws2.cell(row=3, column=1, value="Date").font = Font(bold=True)
    ws2.cell(row=3, column=2, value=ctx.get("meta_created_at", now_iso()))
    ws2.cell(row=4, column=1, value="Language").font = Font(bold=True)
    ws2.cell(row=4, column=2, value=ctx.get("main_language", "?"))
    ws2.cell(row=5, column=1, value="Speakers").font = Font(bold=True)
    ws2.cell(row=5, column=2, value=len(ctx.get("speaker_map", {})))

    # Per-clip table
    r = 7
    clip_headers = ["Clip", "File", "Duration", "Words", "Segments",
                     "Speakers", "Avg Conf", "Low Conf Segs"]
    for col, h in enumerate(clip_headers, 1):
        cell = ws2.cell(row=r, column=col, value=h)
        cell.font = s_hf
        cell.fill = s_fill
    r += 1

    total_words_all = 0
    total_segs_all = 0
    for vf in video_files:
        cid = vf.stem
        tr = ctx["clip_transcripts"].get(cid)
        if not tr:
            continue
        stats = tr.get("stats", {})
        ws2.cell(row=r, column=1, value=cid)
        ws2.cell(row=r, column=2, value=tr.get("clip_file", vf.name))
        ws2.cell(row=r, column=3, value=fmt_duration(tr.get("duration", 0)))
        ws2.cell(row=r, column=4, value=stats.get("total_words", 0))
        ws2.cell(row=r, column=5, value=stats.get("total_segments", 0))
        ws2.cell(row=r, column=6, value=len(stats.get("speakers", {})))
        ws2.cell(row=r, column=7, value=round(stats.get("avg_confidence", 0), 3))
        ws2.cell(row=r, column=8, value=stats.get("low_confidence_segments", 0))
        total_words_all += stats.get("total_words", 0)
        total_segs_all += stats.get("total_segments", 0)
        r += 1

    # Totals row
    tot_font = Font(bold=True)
    ws2.cell(row=r, column=1, value="TOTAL").font = tot_font
    ws2.cell(row=r, column=3, value=fmt_duration(ctx.get("total_duration", 0))).font = tot_font
    ws2.cell(row=r, column=4, value=total_words_all).font = tot_font
    ws2.cell(row=r, column=5, value=total_segs_all).font = tot_font

    # Plan-fact timing table (if available)
    st = ctx.get("stages_timing", {})
    est_times = ctx.get("stage_estimates", {})
    if st:
        r += 2
        timing_headers = ["Stage", "Estimated", "Actual", "Delta"]
        for col, h in enumerate(timing_headers, 1):
            cell = ws2.cell(row=r, column=col, value=h)
            cell.font = s_hf
            cell.fill = s_fill
        r += 1
        stage_labels = [("1", "Extract audio"), ("2", "Diarization"),
                        ("3", "Whisper"), ("4", "Speaker map"),
                        ("5", "Generate"), ("5b", "Premiere proj")]
        total_est = 0; total_act = 0
        for key, label in stage_labels:
            act = st.get(key, 0)
            est = est_times.get(key, 0)
            if act == 0 and key not in st:
                continue
            delta = act - est
            total_est += est; total_act += act
            ws2.cell(row=r, column=1, value=label)
            ws2.cell(row=r, column=2, value=fmt_time(est))
            ws2.cell(row=r, column=3, value=fmt_time(act))
            delta_str = f"+{fmt_time(delta)}" if delta > 0 else f"-{fmt_time(abs(delta))}"
            delta_cell = ws2.cell(row=r, column=4, value=delta_str)
            if delta > est * 0.3:
                delta_cell.font = Font(color="FF0000")
            elif delta < -est * 0.3:
                delta_cell.font = Font(color="008000")
            r += 1
        # Total
        ws2.cell(row=r, column=1, value="TOTAL").font = tot_font
        ws2.cell(row=r, column=2, value=fmt_time(total_est)).font = tot_font
        ws2.cell(row=r, column=3, value=fmt_time(total_act)).font = tot_font
        total_delta = total_act - total_est
        d_str = f"+{fmt_time(total_delta)}" if total_delta > 0 else f"-{fmt_time(abs(total_delta))}"
        ws2.cell(row=r, column=4, value=d_str).font = tot_font

    # Auto-width for summary sheet
    for col_cells in ws2.columns:
        cl = col_cells[0].column_letter
        ml = max((len(str(c.value)) for c in col_cells if c.value), default=8)
        ws2.column_dimensions[cl].width = min(ml + 3, 30)

    # ── Sheet 3: Media ──
    ws3 = wb.create_sheet("Media")
    media_headers = ["Clip", "File", "Duration", "Resolution", "FPS", "Frames",
                     "Video Codec", "Bitrate (Mbps)", "Pixel Format",
                     "Audio", "Size (GB)", "Created", "Timecode"]
    for ci, h in enumerate(media_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = hf
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    clip_media = ctx.get("clip_media_info", {})
    mr = 2
    total_frames = 0
    total_size = 0
    for vf in video_files:
        info = clip_media.get(vf.stem, {})
        w, h = info.get("width", 0), info.get("height", 0)
        fps = info.get("fps", 0)
        frames = info.get("nb_frames", 0)
        vc = info.get("video_codec", "").upper()
        vp = info.get("video_profile", "")
        vbr = info.get("video_bitrate", 0) / 1_000_000
        pf = info.get("pix_fmt", "")
        bits = info.get("bits_per_raw_sample", "")
        ac = info.get("audio_codec", "")
        ar = info.get("audio_sample_rate", 0)
        ach = info.get("audio_channels", 0)
        fs = info.get("file_size", 0)
        ct_raw = info.get("creation_time", "")
        tc = info.get("timecode", "")
        dur = info.get("duration", 0)

        total_frames += frames
        total_size += fs

        res_str = f"{w}×{h}" if w else ""
        codec_str = f"{vc} {vp}" if vp else vc
        pix_str = f"{pf} ({bits}-bit)" if bits else pf
        ch_label = "mono" if ach == 1 else ("stereo" if ach == 2 else f"{ach}ch")
        audio_str = f"{ac.upper()} {ar // 1000}kHz {ch_label}" if ac else ""
        size_gb = round(fs / (1024 ** 3), 2) if fs else 0
        created_str = ct_raw[:16].replace("T", " ") if ct_raw else ""

        ws3.cell(row=mr, column=1, value=vf.stem)
        ws3.cell(row=mr, column=2, value=vf.name)
        ws3.cell(row=mr, column=3, value=fmt_duration(dur))
        ws3.cell(row=mr, column=4, value=res_str)
        ws3.cell(row=mr, column=5, value=fps)
        ws3.cell(row=mr, column=6, value=frames)
        ws3.cell(row=mr, column=7, value=codec_str)
        ws3.cell(row=mr, column=8, value=round(vbr, 1))
        ws3.cell(row=mr, column=9, value=pix_str)
        ws3.cell(row=mr, column=10, value=audio_str)
        ws3.cell(row=mr, column=11, value=size_gb)
        ws3.cell(row=mr, column=12, value=created_str)
        ws3.cell(row=mr, column=13, value=tc)
        for c in range(1, len(media_headers) + 1):
            ws3.cell(row=mr, column=c).border = thin_border
        mr += 1

    # Total row
    bf = Font(bold=True)
    total_dur = sum(clip_media.get(vf.stem, {}).get("duration", 0) for vf in video_files)
    ws3.cell(row=mr, column=1, value="TOTAL").font = bf
    ws3.cell(row=mr, column=3, value=fmt_duration(total_dur)).font = bf
    ws3.cell(row=mr, column=6, value=total_frames).font = bf
    ws3.cell(row=mr, column=11, value=round(total_size / (1024 ** 3), 2)).font = bf
    for c in range(1, len(media_headers) + 1):
        ws3.cell(row=mr, column=c).border = thin_border

    ws3.freeze_panes = "A2"
    media_widths = {"A": 10, "B": 14, "C": 8, "D": 14, "E": 6, "F": 8,
                    "G": 22, "H": 14, "I": 22, "J": 22, "K": 10, "L": 18, "M": 14}
    for letter, mw in media_widths.items():
        ws3.column_dimensions[letter].width = mw

    xlsx_path = output_path or (ctx["work_dir"] / f"{ctx['project_name']}_transcript.xlsx")
    wb.save(str(xlsx_path))
    return xlsx_path


def generate_combined_transcript(ctx):
    offsets = ctx["clip_offsets_flat"]
    all_segs = []; gid = 0; tw = ts = 0; spa = {}
    for vf in ctx["video_files"]:
        cid = vf.stem; tr = ctx["clip_transcripts"].get(cid)
        if not tr: continue
        co = offsets.get(cid, 0.0)
        for seg in tr["segments"]:
            all_segs.append({"id": gid, "clip_id": cid,
                "local_start": seg["start"], "local_end": seg["end"],
                "global_start": round(co+seg["start"],3), "global_end": round(co+seg["end"],3),
                "speaker": seg["speaker"], "text": seg["text"], "avg_confidence": seg["avg_confidence"]})
            gid += 1; sp = seg["speaker"]
            if sp not in spa: spa[sp] = {"segments":0,"words":0,"duration":0.0,"_confs":[]}
            spa[sp]["segments"] += 1; spa[sp]["words"] += len(seg["words"])
            spa[sp]["duration"] += seg["end"]-seg["start"]
            spa[sp]["_confs"].append(seg["avg_confidence"])
        tw += tr["stats"]["total_words"]; ts += tr["stats"]["total_segments"]
    for sp in spa:
        confs = spa[sp].pop("_confs")
        spa[sp]["duration"] = round(spa[sp]["duration"], 3)
        spa[sp]["avg_confidence"] = round(sum(confs)/len(confs), 4) if confs else 0.0
    ac = [s["avg_confidence"] for s in all_segs]
    avg_c = sum(ac)/len(ac) if ac else 0.0
    return {"version": VERSION, "created_at": now_iso(), "project": ctx["project_name"],
        "total_duration": round(ctx["total_duration"],3), "num_speakers": len(ctx["speaker_map"]),
        "language": ctx.get("main_language","unknown"), "clips": [vf.stem for vf in ctx["video_files"]],
        "segments": all_segs, "stats": {"total_words": tw, "total_segments": ts, "speakers": spa,
            "avg_confidence": round(avg_c,4),
            "low_confidence_segments": sum(1 for c in ac if c < LOW_CONFIDENCE_THRESHOLD)}}


def generate_project_json(ctx, args, clip_filter=None, output_path=None):
    """Generate {project}_transcript.json — central project manifest with all data."""
    video_files = ctx["video_files"]
    if clip_filter:
        video_files = [vf for vf in video_files if vf.stem in clip_filter]
    smap = ctx["speaker_map"]
    offsets = ctx["clip_offsets_flat"]
    media_info = ctx.get("clip_media_info", {})
    combined_stats = ctx.get("combined_stats", {})
    # Speakers
    speakers_out = {}
    for label, info in smap.items():
        speakers_out[label] = {"id": info["id"], "name": info["name"]}
    # Params
    params = {
        "whisper_model": args.model,
        "language": args.language,
        "num_speakers": args.n,
    }
    # Stats
    stats = {
        "total_duration": round(ctx["total_duration"], 3),
        "total_words": combined_stats.get("total_words", 0),
        "total_segments": combined_stats.get("total_segments", 0),
        "num_speakers": len(smap),
        "avg_confidence": combined_stats.get("avg_confidence", 0),
        "low_confidence_segments": combined_stats.get("low_confidence_segments", 0),
    }
    # Determine output path early so we can compute relative file paths
    project_path = output_path or (ctx["work_dir"] / f"{ctx['project_name']}_transcript.json")
    project_json_dir = project_path.parent
    # Clips
    clips_out = []
    for vf in video_files:
        cid = vf.stem
        tr = ctx["clip_transcripts"].get(cid)
        if not tr:
            continue
        co = offsets.get(cid, 0.0)
        # Media metadata
        mi = media_info.get(cid, {})
        media_block = {
            "width": mi.get("width", 0),
            "height": mi.get("height", 0),
            "fps": mi.get("fps", 0),
            "nb_frames": mi.get("nb_frames", 0),
            "video_codec": mi.get("video_codec", ""),
            "video_profile": mi.get("video_profile", ""),
            "video_bitrate_mbps": round(int(mi.get("video_bitrate", 0)) / 1_000_000, 1) if mi.get("video_bitrate") else 0,
            "pix_fmt": mi.get("pix_fmt", ""),
            "bit_depth": int(mi.get("bits_per_raw_sample", 0) or 0),
            "audio_codec": mi.get("audio_codec", ""),
            "audio_sample_rate": mi.get("audio_sample_rate", 0),
            "audio_channels": mi.get("audio_channels", 0),
            "file_size_bytes": mi.get("file_size", 0),
            "creation_time": mi.get("creation_time", ""),
            "timecode": mi.get("timecode", ""),
        }
        # File paths (relative to project.json location)
        clip_dir = get_clip_dir(ctx, cid)
        try:
            rel_prefix = clip_dir.relative_to(project_json_dir)
        except ValueError:
            rel_prefix = clip_dir  # fallback to absolute
        files_block = {
            "premiere_transcript": str(rel_prefix / f"{cid}_premiere_transcript.json"),
            "transcript": str(rel_prefix / f"{cid}_transcript.json"),
            "srt": str(rel_prefix / f"{cid}_transcript.srt"),
            "txt": str(rel_prefix / f"{cid}_transcript.txt"),
        }
        # Segments
        segs_out = []
        for seg in tr["segments"]:
            sp_label = seg["speaker"]
            sp_name = smap.get(sp_label, {}).get("name", sp_label)
            dur = round(seg["end"] - seg["start"], 2)
            segs_out.append({
                "id": seg["id"],
                "start": seg["start"],
                "end": seg["end"],
                "duration": dur,
                "timecode": fmt_timecode(co + seg["start"]),
                "words": len(seg["words"]),
                "speaker": sp_name,
                "speaker_id": sp_label,
                "text": seg["text"],
                "confidence": round(seg["avg_confidence"], 3),
                "low_confidence": seg.get("low_confidence", False),
                "use": False,
                "notes": "",
            })
        clips_out.append({
            "clip_id": cid,
            "filename": vf.name,
            "duration": round(tr.get("duration", 0), 2),
            "offset": round(co, 3),
            "media": media_block,
            "files": files_block,
            "segments": segs_out,
        })
    # Structure: project folder layout
    groups = ctx.get("subfolder_groups")
    if groups and not clip_filter:
        # Combined project.json — full multi-camera structure
        cameras = {}
        for grp_name, grp in groups.items():
            cameras[grp_name] = {
                "path": grp_name,
                "clips": len(grp["clip_ids"]),
                "clip_ids": grp["clip_ids"],
                "transcription_dir": f"{grp_name}/{grp_name}_transcription",
                "xlsx": f"{grp_name}/{grp_name}_transcript.xlsx",
                "transcript_json": f"{grp_name}/{grp_name}_transcript.json",
                "srt": f"{grp_name}/{grp_name}_transcript.srt",
            }
        structure = {
            "type": "multi-camera",
            "input_path": str(ctx.get("input_path", ctx["work_dir"])),
            "cameras": cameras,
            "combined": {
                "transcription_dir": f"{ctx['project_name']}_transcription",
                "xlsx": f"{ctx['project_name']}_transcript.xlsx",
                "transcript_json": f"{ctx['project_name']}_transcript.json",
                "srt": f"{ctx['project_name']}_transcript.srt",
                "prproj": f"{ctx['project_name']}.prproj",
            },
        }
    elif groups and clip_filter:
        # Per-camera project.json — show which camera this is
        cam_name = None
        for grp_name, grp in groups.items():
            if clip_filter and set(clip_filter) & set(grp["clip_ids"]):
                cam_name = grp_name
                break
        structure = {
            "type": "camera",
            "camera": cam_name or "unknown",
            "input_path": str(ctx.get("input_path", ctx["work_dir"])),
            "transcription_dir": f"{cam_name}_transcription" if cam_name else "",
        }
    else:
        # Flat / single file
        structure = {
            "type": "folder" if ctx.get("input_type") == "folder" else "single_file",
            "input_path": str(ctx.get("input_path", ctx["work_dir"])),
            "transcription_dir": f"{ctx['project_name']}_transcription",
        }

    # Add full paths to all structure types
    work_dir = ctx["work_dir"]
    structure["work_dir"] = str(work_dir)
    structure["transcription_dir_full"] = str(ctx["transcription_dir"])
    structure["transcript_json"] = str(project_path)
    structure["transcript_xlsx"] = str(work_dir / f"{ctx['project_name']}_transcript.xlsx")
    structure["transcript_srt"] = str(work_dir / f"{ctx['project_name']}_transcript.srt")
    structure["video_files"] = [str(vf) for vf in video_files]

    project = {
        "version": VERSION,
        "project": ctx["project_name"],
        "created_at": ctx.get("meta_created_at", now_iso()),
        "language": ctx.get("main_language", "unknown"),
        "params": params,
        "speakers": speakers_out,
        "stats": stats,
        "structure": structure,
        "clips": clips_out,
    }
    with open(project_path, "w") as f:
        json.dump(project, f, indent=2, ensure_ascii=False)
    return project_path


def stage5_generate_outputs(ctx, args):
    t0 = time.time(); td = ctx["transcription_dir"]; smap = ctx["speaker_map"]
    set_terminal_title("YTAI > Stage 5/5 > Generating outputs"); fc = 0
    clips = ctx["video_files"]
    for idx, vf in enumerate(clips):
        cid = vf.stem; tr = ctx["clip_transcripts"].get(cid)
        if not tr: continue
        cd = get_clip_dir(ctx, cid); cd.mkdir(parents=True, exist_ok=True)
        lang = tr.get("language", ctx.get("main_language","unknown"))
        with open(cd/f"{cid}_transcript.json","w") as f: json.dump(tr,f,indent=2,ensure_ascii=False)
        with open(cd/f"{cid}_transcript.srt","w") as f: f.write(generate_srt(tr,smap))
        with open(cd/f"{cid}_transcript.txt","w") as f: f.write(generate_txt(tr,smap))
        pj = generate_premiere_json(tr, smap, lang)
        with open(cd/f"{cid}_premiere_transcript.json","w") as f: json.dump(pj,f,indent=2,ensure_ascii=False)
        fc += 4
        if len(clips) > 1:
            sys.stdout.write(C.CLEAR_LINE + progress_bar(
                idx + 1, len(clips), width=20, label="Files"))
            sys.stdout.flush()
    if len(clips) > 1:
        sys.stdout.write(C.CLEAR_LINE)
        sys.stdout.flush()
    print(f"  per_clip files: {fc}")
    xlsx_path = generate_xlsx(ctx, args); print(f"  {xlsx_path.name}")
    combined = generate_combined_transcript(ctx)
    with open(td/"combined_transcript.json","w") as f: json.dump(combined,f,indent=2,ensure_ascii=False)
    print(f"  combined_transcript.json"); ctx["combined_stats"] = combined["stats"]
    project_path = generate_project_json(ctx, args); print(f"  {project_path.name}")
    srt_path = generate_combined_srt(ctx); print(f"  {srt_path.name}")
    # Per-camera outputs (multi-camera mode)
    groups = ctx.get("subfolder_groups")
    if groups:
        for grp_name, grp in groups.items():
            grp_xlsx = generate_xlsx(ctx, args, clip_filter=grp["clip_ids"],
                output_path=grp["source_dir"] / f"{grp_name}_transcript.xlsx")
            print(f"  {grp_name}/{grp_xlsx.name}")
            grp_proj = generate_project_json(ctx, args, clip_filter=grp["clip_ids"],
                output_path=grp["source_dir"] / f"{grp_name}_transcript.json")
            print(f"  {grp_name}/{grp_proj.name}")
            grp_srt = generate_combined_srt(ctx, clip_filter=grp["clip_ids"],
                output_path=grp["source_dir"] / f"{grp_name}_transcript.srt")
            print(f"  {grp_name}/{grp_srt.name}")
    elapsed = time.time() - t0; print_done("Stage 5 complete", elapsed); return elapsed



def stage5b_generate_prproj(ctx, args, resources):
    """Stage 5b: Generate .prproj from template.
    TODO: Full clip injection (Media, MasterClips, ClipTrackItems on V1/A1).
    Currently: renames sequence only. User imports clips manually in Premiere."""
    if args.no_prproj or not resources.get("has_template"):
        if not args.no_prproj:
            print_warn("Premiere template not found, skipping .prproj generation")
        return 0.0
    t0 = time.time()
    set_terminal_title("YTAI > [5b] > Generating .prproj")
    with open(TEMPLATE_PATH, "rb") as f:
        xml_bytes = gzip.decompress(f.read())
    xml_str = xml_bytes.decode("utf-8")
    # Update sequence name in XML
    xml_str = re.sub(
        r"(<Sequence[^>]*>.*?<n>)(.*?)(</n>)",
        lambda m: m.group(1) + ctx["project_name"] + m.group(3),
        xml_str, count=1, flags=re.DOTALL)
    prproj_path = ctx["work_dir"] / f"{ctx['project_name']}.prproj"
    with open(prproj_path, "wb") as f:
        f.write(gzip.compress(xml_str.encode("utf-8")))
    elapsed = time.time() - t0
    print(f"  {C.CYAN}[5b]{C.RESET} Generating {ctx['project_name']}.prproj "
          f"(template + sequence named)  {C.GREEN}done{C.RESET}  {fmt_time(elapsed)}")
    return elapsed




def save_meta(ctx, args, resources, st):
    td = ctx["transcription_dir"]
    comp = [k for k in ["1","2","3","4","5","5b"] if k in st]
    status = ctx.get("pipeline_status", "in_progress")
    meta = {"version": VERSION, "status": status,
        "created_at": ctx.get("meta_created_at", now_iso()),
        "updated_at": now_iso(), "project_name": ctx["project_name"],
        "input_type": ctx["input_type"], "input_path": str(ctx["input_path"]),
        "params": {"num_speakers": args.n, "whisper_model": args.model,
            "language": args.language, "device": resources["device"]},
        "clips": [vf.stem for vf in ctx["video_files"]],
        "total_duration_sec": round(ctx["total_duration"], 3), "stages_completed": comp,
        "stages_timing": {"1_extract_audio": round(st.get("1",0),1),
            "2_diarization": round(st.get("2",0),1), "3_whisper": round(st.get("3",0),1),
            "4_speaker_mapping": round(st.get("4",0),1), "5_generate_outputs": round(st.get("5",0),1),
            "5b_premiere_project": round(st.get("5b",0),1)},
        "total_time": round(sum(st.values()),1)}
    with open(td/"meta.json","w") as f:
        json.dump(meta,f,indent=2,ensure_ascii=False)



def write_log(ctx, args, resources, st):
    td = ctx["transcription_dir"]
    lp = td / f"{ctx['project_name']}_transcribe_{now_stamp()}.log"
    cs = ctx.get("combined_stats", {})
    total = sum(st.values())

    lines = [
        f"{'=' * 60}",
        f"YTAI Transcribe v{VERSION}",
        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"{'=' * 60}",
        "",
        "=== Configuration ===",
        f"Project:    {ctx['project_name']}",
        f"Input:      {ctx.get('input_type', '?')} — {args.project}",
        f"Clips:      {len(ctx['video_files'])}",
        f"Duration:   {fmt_duration(ctx['total_duration'])} ({ctx['total_duration']:.1f}s)",
        f"Speakers:   {args.n or 'auto'} (detected: {len(ctx.get('speaker_map', {}))})",
        f"Model:      {args.model}",
        f"Language:   {args.language or 'auto'} (detected: {ctx.get('main_language', '?')})",
        f"Device:     {resources['device']}",
        "",
        "=== Clips ===",
    ]
    clip_media = ctx.get("clip_media_info", {})
    for vf in ctx["video_files"]:
        cid = vf.stem
        dur = ctx.get("clip_durations", {}).get(cid, 0)
        offset = ctx.get("clip_offsets_flat", {}).get(cid, 0)
        info = clip_media.get(cid, {})
        w, h = info.get("width", 0), info.get("height", 0)
        res = f"{w}×{h}" if w else ""
        fps = info.get("fps", "")
        vc = info.get("video_codec", "").upper()
        media_str = f"  {res} {fps}fps {vc}" if w else ""
        lines.append(f"  {cid:<20s}  {fmt_duration(dur):>6s}{media_str}  offset {offset:.3f}s  {vf.name}")

    lines += [
        "",
        "=== Diarization ===",
        f"Speakers: {len(ctx.get('speaker_map', {}))}",
    ]
    smap = ctx.get("speaker_map", {})
    diar_segs = ctx.get("diarization_segments", [])
    for label, info in smap.items():
        sp_segs = [s for s in diar_segs if s["speaker"] == label]
        sp_dur = sum(s["end"] - s["start"] for s in sp_segs)
        pct = sp_dur / ctx["total_duration"] * 100 if ctx["total_duration"] > 0 else 0
        lines.append(f"  {info['name']:<12s} ({label})  {fmt_duration(sp_dur):>6s}  {pct:.1f}%  {len(sp_segs)} segments")
    gap_dur = ctx["total_duration"] - sum(s["end"] - s["start"] for s in diar_segs)
    if ctx["total_duration"] > 0:
        lines.append(f"  {'Gaps':<12s}             {fmt_duration(gap_dur):>6s}  {gap_dur / ctx['total_duration'] * 100:.1f}%")

    lines += [
        "",
        "=== Transcription ===",
        f"Total words:     {cs.get('total_words', 0):,}",
        f"Total segments:  {cs.get('total_segments', 0)}",
        f"Avg confidence:  {cs.get('avg_confidence', 0):.4f}",
        f"Low confidence:  {cs.get('low_confidence_segments', 0)} segments",
    ]
    # List low-confidence segments by clip
    if cs.get("low_confidence_segments", 0) > 0:
        ct = ctx.get("clip_transcripts", {})
        lines.append("  Low confidence details:")
        for vf in ctx["video_files"]:
            tr = ct.get(vf.stem)
            if not tr:
                continue
            for seg in tr.get("segments", []):
                if seg.get("low_confidence"):
                    preview = seg["text"][:60] + "..." if len(seg["text"]) > 60 else seg["text"]
                    conf = seg.get("avg_confidence", 0)
                    lines.append(f"    [{vf.stem}] seg {seg['id']}: \"{preview}\" (conf={conf:.2f})")
    lines.append("")

    # Per-clip details
    ct = ctx.get("clip_transcripts", {})
    if ct:
        lines.append("=== Per-Clip Details ===")
        for vf in ctx["video_files"]:
            cid = vf.stem
            tr = ct.get(cid)
            if not tr:
                continue
            trs = tr.get("stats", {})
            lines.append(f"  {cid}:")
            lines.append(f"    Words: {trs.get('total_words', 0)}  "
                         f"Segments: {trs.get('total_segments', 0)}  "
                         f"Confidence: {trs.get('avg_confidence', 0):.4f}  "
                         f"Low: {trs.get('low_confidence_segments', 0)}")
            for sp_label, sp_stats in trs.get("speakers", {}).items():
                sp_name = smap.get(sp_label, {}).get("name", sp_label)
                lines.append(f"    {sp_name}: {sp_stats.get('words', 0)} words, "
                             f"{sp_stats.get('segments', 0)} segs, "
                             f"{fmt_duration(sp_stats.get('duration', 0))}")
        lines.append("")

    est_times = ctx.get("stage_estimates", {})
    if est_times:
        lines += ["=== Timings (Plan vs Fact) ==="]
        lines.append(f"  {'Stage':<16s}  {'Est':>6s}  {'Actual':>8s}  {'Delta':>8s}  {'%':>5s}")
        lines.append(f"  {'─' * 48}")
    else:
        lines += ["=== Timings ==="]
    for k, label in [("1", "Extract audio"), ("2", "Diarization"), ("3", "Whisper"),
                      ("4", "Speaker map"), ("5", "Generate"),
                      ("5b", "Premiere proj")]:
        t = st.get(k, 0)
        pct = t / total * 100 if total > 0 else 0
        if k in st:
            if est_times:
                e = est_times.get(k, 0)
                delta = t - e
                d_sign = "+" if delta >= 0 else "-"
                lines.append(f"  {label:<16s}  {fmt_time(e):>6s}  {fmt_time(t):>8s}  "
                             f"{d_sign}{fmt_time(abs(delta)):>6s}  {pct:5.1f}%")
            else:
                lines.append(f"  {label:<16s}  {fmt_time(t):>8s}  {pct:5.1f}%")
    total_est = sum(est_times.values()) if est_times else 0
    lines += [
        f"  {'─' * 48}" if est_times else f"  {'─' * 36}",
    ]
    if est_times:
        d = total - total_est
        d_sign = "+" if d >= 0 else "-"
        lines.append(f"  {'Total':<16s}  {fmt_time(total_est):>6s}  {fmt_time(total):>8s}  "
                     f"{d_sign}{fmt_time(abs(d)):>6s}  100.0%")
    else:
        lines.append(f"  {'Total':<16s}  {fmt_time(total):>8s}  100.0%")
    lines += [
        "",
        "=== Output Files ===",
        f"  {ctx['transcription_dir'].name}/",
    ]
    # List output files
    for f in sorted(td.iterdir()):
        if f.is_file() and not f.name.startswith('.'):
            size_kb = f.stat().st_size / 1024
            lines.append(f"    {f.name:<45s}  {size_kb:>8.1f} KB")
    for vf in ctx["video_files"]:
        clip_dir = get_clip_dir(ctx, vf.stem)
        if clip_dir.exists():
            for f in sorted(clip_dir.iterdir()):
                if f.is_file() and not f.name.startswith('.'):
                    size_kb = f.stat().st_size / 1024
                    lines.append(f"    per_clip/{vf.stem}/{f.name:<30s}  {size_kb:>8.1f} KB")

    # Project-level outputs
    xlsx_path = ctx["work_dir"] / f"{ctx['project_name']}_transcript.xlsx"
    if xlsx_path.exists():
        lines.append(f"  {xlsx_path.name}  ({xlsx_path.stat().st_size / 1024:.1f} KB)")

    lines += ["", f"{'=' * 60}"]

    with open(lp, "w") as f:
        f.write("\n".join(lines) + "\n")


def print_summary(ctx, st):
    total = sum(st.values())
    max_stage_time = max(st.values()) if st else 1
    cs = ctx.get("combined_stats", {})
    est_times = ctx.get("stage_estimates", {})
    total_est = sum(est_times.values()) if est_times else 0
    avg_conf = cs.get("avg_confidence", 0)
    low_conf_n = cs.get("low_confidence_segments", 0)
    speaker_stats = cs.get("speakers", {})
    smap = ctx.get("speaker_map", {})
    total_dur = ctx.get("total_duration", 0)

    W = 62
    B = "║"

    print(f"\n  {C.BG_GREEN}{C.WHITE}{C.BOLD} Summary {C.RESET}")
    print(f"  {C.DIM}╔{'═' * W}╗{C.RESET}")

    # ── Key stats ──
    time_line = f"  Total time:   {C.BOLD}{fmt_time(total)}{C.RESET}"
    if total_est > 0:
        delta = total - total_est
        d_color = C.RED if delta > 0 else C.GREEN
        d_sign = "+" if delta > 0 else "−"
        time_line += (f"  {C.DIM}(est: {fmt_time(total_est)}, "
                      f"{d_color}{d_sign}{fmt_time(abs(delta))}{C.RESET}{C.DIM}){C.RESET}")
    print(_trow(time_line, W, B))

    # Processing speed
    if total > 0 and total_dur > 0:
        speed = total_dur / total
        speed_line = (f"  Speed:        {C.BOLD}{speed:.1f}x{C.RESET} realtime"
                      f"  {C.DIM}({fmt_duration(total_dur)} video → {fmt_time(total)}){C.RESET}")
        print(_trow(speed_line, W, B))

    print(_trow(f"  Words:        {cs.get('total_words', 0):,}", W, B))
    print(_trow(f"  Segments:     {cs.get('total_segments', 0)}", W, B))
    print(_trow(f"  Speakers:     {len(smap)}", W, B))
    print(_trow(f"  Language:     {ctx.get('main_language', '?')}", W, B))
    conf_color = C.GREEN if avg_conf >= 0.85 else (C.YELLOW if avg_conf >= 0.75 else C.RED)
    conf_line = (f"  Confidence:   {conf_color}{avg_conf:.2f}{C.RESET}"
                 f"  {C.DIM}({low_conf_n} low){C.RESET}")
    print(_trow(conf_line, W, B))
    print(_trow(f"  Output:       {ctx['transcription_dir'].name}/", W, B))
    print(_trow(f"                {ctx['project_name']}_transcript.xlsx", W, B))

    # ── Per-speaker breakdown ──
    if speaker_stats:
        print(f"  {C.DIM}╠{'═' * W}╣{C.RESET}")
        print(_trow(f"  {C.BOLD}Speakers{C.RESET}", W, B))
        print(_trow(f"  {C.DIM}{'─' * 58}{C.RESET}", W, B))
        # Calculate total speaker duration for gap calculation
        total_spk_dur = sum(sp.get("duration", 0) for sp in speaker_stats.values())
        gap_dur = max(total_dur - total_spk_dur, 0)
        for label, sp_info in sorted(speaker_stats.items(), key=lambda x: -x[1].get("duration", 0)):
            name = smap.get(label, {}).get("name", label)
            dur = sp_info.get("duration", 0)
            words = sp_info.get("words", 0)
            sp_conf = sp_info.get("avg_confidence", 0)
            pct = dur / total_dur * 100 if total_dur > 0 else 0
            conf_c = C.GREEN if sp_conf >= 0.85 else (C.YELLOW if sp_conf >= 0.75 else C.RED)
            row = (f"  {name:<14s}"
                   f"{fmt_duration(dur):>6s} ({pct:4.1f}%)"
                   f"  {words:>5,} words"
                   f"  {conf_c}{sp_conf:.2f}{C.RESET} conf")
            print(_trow(row, W, B))
        # Gaps
        if gap_dur > 0 and total_dur > 0:
            gap_pct = gap_dur / total_dur * 100
            print(_trow(f"  {C.DIM}{'Gaps':<14s}{fmt_duration(gap_dur):>6s} ({gap_pct:4.1f}%){C.RESET}", W, B))

    # ── Stage breakdown ──
    print(f"  {C.DIM}╠{'═' * W}╣{C.RESET}")
    if est_times:
        hdr = f"  {C.DIM}{'Stage':<16s}{'Est':>8s}{'Actual':>9s}{'Δ':>9s} {'%':>4s} Bar{C.RESET}"
        print(_trow(hdr, W, B))
        print(_trow(f"  {C.DIM}{'─' * 58}{C.RESET}", W, B))
    else:
        hdr = f"  {C.DIM}{'Stage':<16s}{'Time':>9s} {'%':>4s} Bar{C.RESET}"
        print(_trow(hdr, W, B))
        print(_trow(f"  {C.DIM}{'─' * 58}{C.RESET}", W, B))

    for k, label in STAGE_LABELS:
        t = st.get(k, 0)
        if t == 0 and k not in st:
            continue
        pct = t / total * 100 if total > 0 else 0
        bar = summary_bar(t, max_stage_time, width=10)
        color = C.GREEN if pct < 5 else (C.YELLOW if pct < 40 else C.RED)
        if est_times:
            e = est_times.get(k, 0)
            delta = t - e
            d_color = C.RED if delta > e * 0.3 else (C.GREEN if delta < -e * 0.3 else C.DIM)
            d_sign = "+" if delta >= 0 else "−"
            delta_txt = f"{d_sign}{fmt_time(abs(delta))}"
            row = (f"  {label:<16s}"
                   f"{fmt_time(e):>8s}"
                   f"{fmt_time(t):>9s} "
                   f"{d_color}{delta_txt:>8s}{C.RESET}"
                   f" {pct:3.0f}% {color}{bar}{C.RESET}")
        else:
            row = f"  {label:<16s}{fmt_time(t):>9s} {pct:3.0f}% {color}{bar}{C.RESET}"
        print(_trow(row, W, B))

    # ── Quality analysis ──
    print(f"  {C.DIM}╠{'═' * W}╣{C.RESET}")
    print(_trow(f"  {C.BOLD}Quality{C.RESET}", W, B))
    print(_trow(f"  {C.DIM}{'─' * 56}{C.RESET}", W, B))

    # Overall confidence
    if avg_conf >= 0.85:
        print(_trow(f"  {C.GREEN}✓{C.RESET}  High confidence ({avg_conf:.2f} avg)", W, B))
    elif avg_conf >= 0.75:
        print(_trow(f"  {C.YELLOW}⚠{C.RESET}  Medium confidence ({avg_conf:.2f} avg) — review flagged segments", W, B))
    else:
        print(_trow(f"  {C.RED}⚠{C.RESET}  Low confidence ({avg_conf:.2f} avg) — check audio quality", W, B))

    # Low confidence segments
    if low_conf_n > 0:
        print(_trow(f"  {C.YELLOW}⚠{C.RESET}  {low_conf_n} low-confidence segments — check manually", W, B))

    # Per-speaker confidence warnings
    for label, sp_info in speaker_stats.items():
        sp_conf = sp_info.get("avg_confidence", 1.0)
        name = smap.get(label, {}).get("name", label)
        if sp_conf < 0.75:
            print(_trow(f"  {C.YELLOW}⚠{C.RESET}  {name}: low confidence ({sp_conf:.2f}) — may be background", W, B))

    # Speaker balance
    if speaker_stats and total_dur > 0:
        durs = sorted([(sp.get("duration", 0), label) for label, sp in speaker_stats.items()], reverse=True)
        max_pct = durs[0][0] / total_dur * 100
        if len(durs) == 1:
            print(_trow(f"  {C.GREEN}✓{C.RESET}  Single speaker detected", W, B))
        elif max_pct > 90:
            print(_trow(f"  {C.GREEN}✓{C.RESET}  Clear main speaker ({max_pct:.0f}%)", W, B))
        elif max_pct > 65:
            parts = "/".join(f"{d/total_dur*100:.0f}%" for d, _ in durs)
            print(_trow(f"  {C.GREEN}✓{C.RESET}  Speaker balance: {parts}", W, B))
        else:
            parts = "/".join(f"{d/total_dur*100:.0f}%" for d, _ in durs)
            print(_trow(f"  {C.GREEN}✓{C.RESET}  Balanced dialogue: {parts}", W, B))

    # Gap analysis
    if speaker_stats and total_dur > 0:
        total_spk_dur = sum(sp.get("duration", 0) for sp in speaker_stats.values())
        gap_pct = (total_dur - total_spk_dur) / total_dur * 100
        if gap_pct > 15:
            print(_trow(f"  {C.YELLOW}⚠{C.RESET}  Large gaps ({gap_pct:.0f}%) — check diarization or audio", W, B))

    print(f"  {C.DIM}╚{'═' * W}╝{C.RESET}")


def try_resume(ctx, args):
    td = ctx["transcription_dir"]; mp = td/"meta.json"
    if not mp.exists(): print_warn("Cannot resume: no meta.json"); sys.exit(1)
    try:
        with open(mp) as f: meta = json.load(f)
    except Exception: print_warn("Cannot resume: corrupted meta.json"); sys.exit(1)
    # Warn about incomplete previous run
    if meta.get("status") == "in_progress":
        print_warn("Previous run did not complete (status: in_progress)")
    # Validate parameter mismatch
    saved = meta.get("params", {})
    mismatches = []
    if args.n is not None and saved.get("num_speakers") != args.n:
        mismatches.append(f"speakers: saved={saved.get('num_speakers')}, current={args.n}")
    if args.model != saved.get("whisper_model", args.model):
        mismatches.append(f"model: saved={saved.get('whisper_model')}, current={args.model}")
    if args.language and args.language != saved.get("language"):
        mismatches.append(f"language: saved={saved.get('language')}, current={args.language}")
    if mismatches:
        print_warn(f"Parameter mismatch: {'; '.join(mismatches)}")
        print_warn("Run without --resume to start fresh, or continue at your own risk")
    if set(meta.get("clips",[])) != set(vf.stem for vf in ctx["video_files"]):
        print_warn("Clips changed. Run without --resume."); sys.exit(1)
    completed = set(meta.get("stages_completed",[]))
    for s in ["1","2","3","4","5","5b"]:
        if s not in completed:
            print(f"  Resuming from stage {s}")
            if "1" in completed:
                with open(td/"clip_offsets.json") as f: co = json.load(f)
                ctx["clip_offsets"] = co["clips"]
                ctx["clip_offsets_flat"] = {c["clip_id"]: c["offset"] for c in co["clips"]}
                ctx["total_duration"] = co["total_duration"]
                ctx["clip_durations"] = {c["clip_id"]: c["duration"] for c in co["clips"]}
            if "2" in completed:
                with open(td/"diarization.json") as f: ctx["diarization_segments"] = json.load(f)["segments"]
                with open(td/"speakers.json") as f: ctx["speaker_map"] = json.load(f)["speakers"]
            if "4" in completed and s in ("5", "5b"):
                ct = {}
                for vf in ctx["video_files"]:
                    jp = get_clip_dir(ctx, vf.stem)/f"{vf.stem}_transcript.json"
                    if jp.exists():
                        try:
                            with open(jp) as f: ct[vf.stem] = json.load(f)
                        except json.JSONDecodeError:
                            print_warn(f"Corrupted {jp.name}, re-running from stage 3")
                            return "3"
                ctx["clip_transcripts"] = ct
                # Restore main_language from per-clip data
                langs = [t.get("language","unknown") for t in ct.values()]
                if langs:
                    from collections import Counter
                    ctx["main_language"] = Counter(langs).most_common(1)[0][0]
            ctx["meta_created_at"] = meta.get("created_at", now_iso())
            # Restore timing data from previous stages so XLSX plan-fact works
            prev_timing = meta.get("stages_timing", {})
            stage_key_map = {"1_extract_audio": "1", "2_diarization": "2",
                             "3_whisper": "3", "4_speaker_mapping": "4",
                             "5_generate_outputs": "5", "5b_premiere_project": "5b"}
            for tkey, skey in stage_key_map.items():
                if skey in completed and tkey in prev_timing:
                    ctx["stages_timing"][skey] = prev_timing[tkey]
            return s
    print("  All stages done."); sys.exit(0)


def main():
    args = parse_args()
    ctx = detect_input(args)
    ctx["meta_created_at"] = now_iso()
    resources = preflight(ctx, args)

    # --dry-run: print plan and exit
    if args.dry_run:
        str_set = set(args.stages.split(",")) if args.stages else {"1","2","3","4","5","5b"}
        est = ctx.get("stage_estimates", {})
        total_est = sum(est.get(k, 0) for k in str_set)
        est_range = f"~{total_est / 60 * 0.85:.0f}-{total_est / 60 * 1.15:.0f} min"
        print(f"\n  {C.BG_CYAN}{C.WHITE}{C.BOLD} Dry run {C.RESET}  — would process:\n")
        clip_media = ctx.get("clip_media_info", {})
        groups = ctx.get("subfolder_groups")
        if groups:
            for grp_name, grp in groups.items():
                print(f"    {C.BOLD}[{grp_name}]{C.RESET}  ({len(grp['clip_ids'])} clips)")
                for vf in grp["clips"]:
                    dur = ctx["clip_durations"].get(vf.stem, 0)
                    info = clip_media.get(vf.stem, {})
                    w, h = info.get("width", 0), info.get("height", 0)
                    res = f"{w}×{h}" if w else ""
                    fps = f"{info.get('fps', '')}fps" if info.get("fps") else ""
                    vc = info.get("video_codec", "").upper()
                    print(f"      {vf.name:<25s}  {fmt_duration(dur):>6s}  {res} {fps} {vc}")
                print()
        else:
            for vf in ctx["video_files"]:
                dur = ctx["clip_durations"].get(vf.stem, 0)
                info = clip_media.get(vf.stem, {})
                w, h = info.get("width", 0), info.get("height", 0)
                res = f"{w}×{h}" if w else ""
                fps = f"{info.get('fps', '')}fps" if info.get("fps") else ""
                vc = info.get("video_codec", "").upper()
                print(f"    {vf.name:<20s}  {fmt_duration(dur):>6s}  {res} {fps} {vc}")
        print(f"\n    Stages:    {','.join(sorted(str_set))}")
        print(f"    Estimated: {est_range}")
        print()
        sys.exit(0)

    str_set = set(args.stages.split(",")) if args.stages else {"1","2","3","4","5","5b"}
    start_s = "1"
    st = {}
    ctx["stages_timing"] = st  # Updated incrementally so Stage 5 XLSX can access it
    if args.resume: start_s = try_resume(ctx, args)
    ctx["pipeline_status"] = "in_progress"
    pipeline_prog = PipelineProgress()
    ctx["transcription_dir"].mkdir(parents=True, exist_ok=True)
    started = False
    est_times = ctx.get("stage_estimates", {})
    for s in ["1","2","3","4","5","5b"]:
        if not started:
            if s == start_s: started = True
            else: continue
        if s not in str_set: continue
        # Dashboard before stage: show current stage as ▶
        print_status_table(st, est_times, str_set, current_stage=s, ctx=ctx)
        # Compute drift-adjusted plan for this stage and store in ctx
        drift = compute_drift(st, est_times)
        ctx["_current_drift"] = drift
        ctx["_stage_plan_adj"] = est_times.get(s, 0) * drift
        if s == "1":
            print_stage_header("1", "5", "Extracting audio",
                               s, est_times, st, str_set)
            st["1"] = stage1_extract_audio(ctx, args)
        elif s == "2":
            print_stage_header("2", "5",
                               f"Speaker diarization ({fmt_duration(ctx['total_duration'])})",
                               s, est_times, st, str_set)
            st["2"] = stage2_diarization(ctx, args, resources)
        elif s == "3":
            print_stage_header("3", "5", "Whisper transcription (word_timestamps)",
                               s, est_times, st, str_set)
            st["3"] = stage3_whisper(ctx, args, resources)
        elif s == "4":
            print_stage_header("4", "5", "Mapping speakers",
                               s, est_times, st, str_set)
            st["4"] = stage4_speaker_mapping(ctx, args)
        elif s == "5":
            print_stage_header("5", "5", "Generating outputs",
                               s, est_times, st, str_set)
            st["5"] = stage5_generate_outputs(ctx, args)
        elif s == "5b":
            st["5b"] = stage5b_generate_prproj(ctx, args, resources)
        pipeline_prog.stage_completed(s)
        save_meta(ctx, args, resources, st)
    # Final dashboard: all stages done
    print_status_table(st, est_times, str_set, current_stage=None, ctx=ctx)
    ctx["pipeline_status"] = "completed"
    save_meta(ctx, args, resources, st)
    ctx["stages_timing"] = st
    write_log(ctx, args, resources, st)
    print_summary(ctx, st)
    set_terminal_title(f"YTAI > Done > {ctx['project_name']}")
    print("\a")  # Terminal bell


if __name__ == "__main__":
    main()
