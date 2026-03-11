#!/usr/bin/env python3
"""
whisper_batch.py — Video transcription with speaker diarization

Whisper (transcription) + pyannote (diarization) = text with speaker labels

Features:
  - Automatic preflight checks before processing
  - Speaker diarization with pyannote 3.x
  - Excel output with speaker columns
  - Support for single files and folders
  - Folder mode: organized output in {folder}_transcription/ subdirectory
  - Smart project naming (auto-detects from parent folder or --name)

Modes:
  1. Preflight only: python whisper_batch.py --preflight
  2. Single file:    python whisper_batch.py video.mp4
  3. Folder:         python whisper_batch.py /path/to/folder --folder

Output structure (folder mode):
  folder/
  └── folder_transcription/
      ├── folder_transcription.xlsx   (combined table)
      ├── video1/
      │   ├── video1.wav
      │   ├── video1.json
      │   ├── video1.srt
      │   ├── video1.txt
      │   └── video1.xlsx
      └── video2/
          └── ...

Requirements:
    pip install openai-whisper openpyxl torch pyannote.audio soundfile
    brew install ffmpeg
    
    # HuggingFace token for pyannote:
    # 1. https://huggingface.co/settings/tokens → create token
    # 2. https://huggingface.co/pyannote/speaker-diarization-3.1 → Accept
    # 3. https://huggingface.co/pyannote/segmentation-3.0 → Accept
    # 4. export HF_TOKEN="hf_xxx" or huggingface-cli login
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple

# ============================================================================
# Configuration
# ============================================================================

VERSION = "2.3.0"
VIDEO_EXTS = {".mp4", ".mov", ".m4v", ".mts", ".avi", ".mkv", ".webm"}
AUDIO_EXTS = {".m4a", ".mp3", ".wav", ".ogg", ".flac", ".aac", ".opus"}
MEDIA_EXTS = VIDEO_EXTS | AUDIO_EXTS
DEFAULT_WHISPER_MODEL = "large-v3"
AUDIO_SAMPLE_RATE = 48000  # 48kHz for Premiere Pro compatibility
MIN_DISK_SPACE_GB = 1.0
HF_CONFIG_PATH = Path.home() / "YTAI" / "config" / "HuggingFace-yt-prod.conf"

# Model cache
_whisper_model = None
_whisper_model_name = None
_pyannote_pipeline = None


# ============================================================================
# Preflight Checks
# ============================================================================

class PreflightResult:
    """Container for preflight check results."""
    def __init__(self):
        self.errors: List[Tuple[str, str]] = []      # (check_name, message)
        self.warnings: List[Tuple[str, str]] = []    # (check_name, message)
        self.passed: List[Tuple[str, str]] = []      # (check_name, details)
    
    def add_error(self, check: str, message: str):
        self.errors.append((check, message))
    
    def add_warning(self, check: str, message: str):
        self.warnings.append((check, message))
    
    def add_passed(self, check: str, details: str = ""):
        self.passed.append((check, details))
    
    @property
    def ok(self) -> bool:
        return len(self.errors) == 0


def get_ffmpeg_version() -> Optional[str]:
    """Get ffmpeg version string."""
    try:
        result = subprocess.run(
            ["ffmpeg", "-version"],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            # Parse first line: "ffmpeg version 7.1 ..."
            first_line = result.stdout.split('\n')[0]
            parts = first_line.split()
            if len(parts) >= 3:
                return parts[2]
        return "installed"
    except Exception:
        return None


def get_package_version(package_name: str) -> Optional[str]:
    """Get installed package version."""
    try:
        if package_name == "openai-whisper":
            import whisper
            return getattr(whisper, '__version__', 'installed')
        elif package_name == "torch":
            import torch
            return torch.__version__
        elif package_name == "pyannote.audio":
            import pyannote.audio
            return getattr(pyannote.audio, '__version__', 'installed')
        elif package_name == "soundfile":
            import soundfile
            return getattr(soundfile, '__version__', 'installed')
        elif package_name == "openpyxl":
            import openpyxl
            return openpyxl.__version__
        return None
    except ImportError:
        return None


def get_hf_token() -> Optional[str]:
    """Get HuggingFace token from multiple sources."""
    # 1. Environment variable
    token = os.environ.get("HF_TOKEN") or os.environ.get("HUGGINGFACE_TOKEN")
    if token:
        return token
    
    # 2. YTAI config file
    if HF_CONFIG_PATH.exists():
        try:
            with open(HF_CONFIG_PATH) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and 'HF_TOKEN' in line and '=' in line:
                        return line.split('=', 1)[1].strip().strip('"\'')
        except Exception:
            pass
    
    # 3. HuggingFace cache
    token_path = Path.home() / ".cache" / "huggingface" / "token"
    if token_path.exists():
        try:
            return token_path.read_text().strip()
        except Exception:
            pass
    
    return None


def get_disk_space_gb(path: Path) -> float:
    """Get free disk space in GB for the given path."""
    try:
        stat = shutil.disk_usage(path)
        return stat.free / (1024 ** 3)
    except Exception:
        return 0.0


def get_device_info() -> Tuple[str, bool]:
    """Get available compute device and whether it's GPU."""
    try:
        import torch
        if torch.backends.mps.is_available():
            return "Apple Silicon (MPS)", True
        elif torch.cuda.is_available():
            return f"CUDA ({torch.cuda.get_device_name(0)})", True
        else:
            return "CPU", False
    except Exception:
        return "Unknown", False


def run_preflight_checks(
    input_path: Optional[Path] = None,
    is_folder: bool = False,
) -> PreflightResult:
    """Run all preflight checks and return results."""
    result = PreflightResult()
    
    print("")
    print("╔" + "═" * 68 + "╗")
    print("║" + "  PREFLIGHT CHECKS".center(68) + "║")
    print("╚" + "═" * 68 + "╝")
    print("")
    
    # ========================================
    # 1. System Dependencies
    # ========================================
    print("[1/6] System dependencies")
    
    ffmpeg_version = get_ffmpeg_version()
    if ffmpeg_version:
        result.add_passed("ffmpeg", ffmpeg_version)
        print(f"  ✅ ffmpeg: {ffmpeg_version}")
    else:
        result.add_error("ffmpeg", "Not installed")
        print(f"  ❌ ffmpeg: NOT INSTALLED")
    
    # Check ffprobe (comes with ffmpeg)
    try:
        subprocess.run(["ffprobe", "-version"], capture_output=True, timeout=10)
        result.add_passed("ffprobe", "available")
        print(f"  ✅ ffprobe: available")
    except Exception:
        result.add_error("ffprobe", "Not installed")
        print(f"  ❌ ffprobe: NOT INSTALLED")
    
    print("")
    
    # ========================================
    # 2. Python Packages
    # ========================================
    print("[2/6] Python packages")
    
    required_packages = [
        ("openai-whisper", "Transcription"),
        ("torch", "ML framework"),
        ("pyannote.audio", "Speaker diarization"),
        ("soundfile", "Audio loading"),
        ("openpyxl", "Excel output"),
    ]
    
    missing_packages = []
    for pkg_name, purpose in required_packages:
        version = get_package_version(pkg_name)
        if version:
            result.add_passed(pkg_name, version)
            print(f"  ✅ {pkg_name}: {version}")
        else:
            result.add_error(pkg_name, f"Not installed ({purpose})")
            missing_packages.append(pkg_name)
            print(f"  ❌ {pkg_name}: NOT INSTALLED ({purpose})")
    
    print("")
    
    # ========================================
    # 3. Authentication
    # ========================================
    print("[3/6] Authentication")
    
    hf_token = get_hf_token()
    if hf_token:
        # Mask token for display
        masked = hf_token[:8] + "..." + hf_token[-4:] if len(hf_token) > 12 else "***"
        result.add_passed("HF_TOKEN", f"found ({masked})")
        print(f"  ✅ HF_TOKEN: found ({masked})")
        
        # Check token source
        if os.environ.get("HF_TOKEN") or os.environ.get("HUGGINGFACE_TOKEN"):
            print(f"     Source: environment variable")
        elif HF_CONFIG_PATH.exists():
            print(f"     Source: {HF_CONFIG_PATH}")
        else:
            print(f"     Source: ~/.cache/huggingface/token")
    else:
        result.add_error("HF_TOKEN", "Not found (required for speaker diarization)")
        print(f"  ❌ HF_TOKEN: NOT FOUND")
        print(f"     Required for speaker diarization")
    
    print("")
    
    # ========================================
    # 4. Models
    # ========================================
    print("[4/6] Models")
    
    # Check Whisper model cache
    whisper_cache = Path.home() / ".cache" / "whisper" / f"{DEFAULT_WHISPER_MODEL}.pt"
    if whisper_cache.exists():
        size_gb = whisper_cache.stat().st_size / (1024 ** 3)
        result.add_passed(f"Whisper {DEFAULT_WHISPER_MODEL}", f"cached ({size_gb:.1f} GB)")
        print(f"  ✅ Whisper {DEFAULT_WHISPER_MODEL}: cached ({size_gb:.1f} GB)")
    else:
        result.add_warning(f"Whisper {DEFAULT_WHISPER_MODEL}", "Will download on first run (~2.9 GB)")
        print(f"  ⚠️  Whisper {DEFAULT_WHISPER_MODEL}: will download (~2.9 GB)")
    
    # Check pyannote model (harder to check, just note it)
    print(f"  ℹ️  pyannote/speaker-diarization-3.1: checked at runtime")
    
    print("")
    
    # ========================================
    # 5. Input Validation
    # ========================================
    print("[5/6] Input validation")
    
    if input_path is None:
        print(f"  ℹ️  No input specified (preflight only mode)")
    elif not input_path.exists():
        result.add_error("input_path", f"Path not found: {input_path}")
        print(f"  ❌ Path: NOT FOUND")
        print(f"     {input_path}")
    else:
        result.add_passed("input_path", str(input_path))
        print(f"  ✅ Path exists: {input_path.name}")
        
        # Count videos
        if is_folder:
            videos = [f for f in input_path.iterdir() 
                     if f.is_file() and f.suffix.lower() in MEDIA_EXTS and not f.name.startswith(".")]
            if videos:
                total_size = sum(v.stat().st_size for v in videos)
                size_str = f"{total_size / (1024**3):.2f} GB" if total_size > 1024**3 else f"{total_size / (1024**2):.1f} MB"
                result.add_passed("videos_found", f"{len(videos)} files ({size_str})")
                print(f"  ✅ Videos found: {len(videos)} files ({size_str})")
            else:
                result.add_error("videos_found", "No video files in folder")
                print(f"  ❌ Videos: NONE FOUND in folder")
        else:
            if input_path.suffix.lower() in MEDIA_EXTS:
                size = input_path.stat().st_size
                size_str = f"{size / (1024**3):.2f} GB" if size > 1024**3 else f"{size / (1024**2):.1f} MB"
                result.add_passed("video_file", f"{input_path.name} ({size_str})")
                print(f"  ✅ Video file: {size_str}")
            else:
                result.add_error("video_file", f"Not a supported media file: {input_path.suffix}")
                print(f"  ❌ Not a supported media file: {input_path.suffix}")
                print(f"     Supported: {', '.join(sorted(MEDIA_EXTS))}")
        
        # Check disk space
        free_space = get_disk_space_gb(input_path.parent if input_path.is_file() else input_path)
        if free_space >= MIN_DISK_SPACE_GB:
            result.add_passed("disk_space", f"{free_space:.1f} GB free")
            print(f"  ✅ Disk space: {free_space:.1f} GB free")
        else:
            result.add_error("disk_space", f"Only {free_space:.1f} GB free (need {MIN_DISK_SPACE_GB} GB)")
            print(f"  ❌ Disk space: {free_space:.1f} GB free (need {MIN_DISK_SPACE_GB} GB)")
    
    print("")
    
    # ========================================
    # 6. GPU/Device
    # ========================================
    print("[6/6] Compute device")
    
    device_name, is_gpu = get_device_info()
    if is_gpu:
        result.add_passed("device", device_name)
        print(f"  ✅ {device_name}")
    else:
        result.add_warning("device", "CPU only (will be slow)")
        print(f"  ⚠️  {device_name} (processing will be slow)")
    
    print("")
    
    # ========================================
    # Summary
    # ========================================
    print("─" * 70)
    
    if result.ok:
        if result.warnings:
            print(f"✅ All critical checks passed ({len(result.warnings)} warnings)")
        else:
            print(f"✅ All checks passed! Ready to process.")
    else:
        print(f"❌ {len(result.errors)} error(s) found. Cannot proceed.")
        print("")
        print("Fix with:")
        print("")
        
        # Generate fix commands
        if missing_packages:
            pkg_list = " ".join(missing_packages)
            print(f"  pip install {pkg_list} --break-system-packages")
        
        if not hf_token:
            print("")
            print("  # Set HuggingFace token:")
            print("  export HF_TOKEN='hf_your_token_here'")
            print("")
            print("  # Or login with CLI:")
            print("  huggingface-cli login")
        
        if not ffmpeg_version:
            print("")
            print("  # Install ffmpeg:")
            print("  brew install ffmpeg")
    
    print("─" * 70)
    print("")
    
    return result


# ============================================================================
# Utilities
# ============================================================================

def natural_sort_key(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]


def format_timestamp(seconds: float) -> str:
    """Format as MM:SS or HH:MM:SS."""
    total_secs = int(round(seconds))
    h = total_secs // 3600
    m = (total_secs % 3600) // 60
    s = total_secs % 60
    
    if h > 0:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def format_srt_timestamp(seconds: float) -> str:
    """Format as HH:MM:SS,mmm for SRT."""
    total_ms = int(round(seconds * 1000))
    ms = total_ms % 1000
    total_secs = total_ms // 1000
    h = total_secs // 3600
    m = (total_secs % 3600) // 60
    s = total_secs % 60
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


def format_duration(seconds: float) -> str:
    """Human-readable duration."""
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    
    if h > 0:
        return f"{h}h {m}m {s}s"
    elif m > 0:
        return f"{m}m {s}s"
    return f"{s}s"


def format_size(size_bytes: int) -> str:
    """Human-readable file size."""
    if size_bytes >= 1024 * 1024 * 1024:
        return f"{size_bytes / (1024**3):.2f} GB"
    elif size_bytes >= 1024 * 1024:
        return f"{size_bytes / (1024**2):.1f} MB"
    elif size_bytes >= 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes} bytes"


def log(msg: str, indent: int = 0):
    """Print log message with timestamp."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    prefix = "  " * indent
    print(f"[{timestamp}] {prefix}{msg}")


# ============================================================================
# Project naming
# ============================================================================

# Folder names that are "technical" (not good project names)
TECHNICAL_FOLDER_NAMES = {
    "01_01_video", "01_02_audio", "01_raw", "01_source",
    "video", "videos", "audio", "raw", "source", "media", "clips",
    "footage", "input", "inputs",
}


def resolve_project_name(folder: Path, user_name: Optional[str] = None) -> str:
    """
    Determine project name for output folder and files.
    
    Priority:
      1. --name parameter (user override)
      2. Folder name if it's meaningful (not technical)
      3. Parent folder name if current folder is technical (e.g. 01_01_Video)
      4. Fallback to folder name anyway
    """
    if user_name:
        return user_name
    
    folder_name = folder.name
    
    # Check if folder name is "technical" → use parent
    if folder_name.lower().replace(" ", "_") in TECHNICAL_FOLDER_NAMES:
        parent_name = folder.parent.name
        # Don't go up if parent is also technical or is root
        if parent_name and parent_name.lower().replace(" ", "_") not in TECHNICAL_FOLDER_NAMES:
            return parent_name
    
    return folder_name


# ============================================================================
# Video discovery
# ============================================================================

def find_media_in_folder(folder: Path) -> List[Path]:
    """Find all media files (video + audio) in folder (not in subdirectories)."""
    media = []
    for f in folder.iterdir():
        if f.is_file() and f.suffix.lower() in MEDIA_EXTS and not f.name.startswith("."):
            media.append(f)
    media.sort(key=lambda p: natural_sort_key(p.name))
    return media


def is_media_file(path: Path) -> bool:
    """Check if path is a supported media file (video or audio)."""
    return path.is_file() and path.suffix.lower() in MEDIA_EXTS


# ============================================================================
# Audio extraction
# ============================================================================

def get_media_duration(path: Path) -> float:
    try:
        result = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(path)],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0 and result.stdout.strip():
            return float(result.stdout.strip())
    except Exception:
        pass
    return 0.0


def get_media_info(path: Path) -> dict:
    """Get detailed media information."""
    info = {"duration": 0, "size": 0, "video_codec": "", "audio_codec": ""}
    
    try:
        info["size"] = path.stat().st_size
        
        result = subprocess.run(
            ["ffprobe", "-v", "error", "-select_streams", "v:0",
             "-show_entries", "stream=codec_name,width,height,r_frame_rate",
             "-of", "json", str(path)],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if data.get("streams"):
                s = data["streams"][0]
                info["video_codec"] = s.get("codec_name", "")
                info["width"] = s.get("width", 0)
                info["height"] = s.get("height", 0)
        
        info["duration"] = get_media_duration(path)
    except Exception:
        pass
    
    return info


def extract_audio(video_path: Path, output_path: Path) -> bool:
    """Extract audio from video to WAV."""
    log("Extracting audio track...", 1)
    log(f"Input: {video_path.name}", 2)
    log(f"Output: {output_path.name}", 2)
    log(f"Format: WAV {AUDIO_SAMPLE_RATE}Hz mono 16-bit PCM", 2)
    
    start_time = datetime.now()
    
    cmd = [
        "ffmpeg", "-hide_banner", "-loglevel", "error", "-y",
        "-i", str(video_path),
        "-vn", "-sn", "-dn",
        "-ar", str(AUDIO_SAMPLE_RATE),
        "-ac", "1",
        "-c:a", "pcm_s16le",
        str(output_path),
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    elapsed = (datetime.now() - start_time).total_seconds()
    
    if result.returncode != 0:
        log(f"ERROR: ffmpeg failed: {result.stderr.strip()}", 2)
        return False
    
    if not output_path.exists() or output_path.stat().st_size < 1000:
        log("ERROR: Audio file not created or empty", 2)
        return False
    
    size = output_path.stat().st_size
    duration = get_media_duration(output_path)
    
    log(f"DONE: {format_size(size)}, {format_duration(duration)} ({elapsed:.1f}s)", 2)
    
    return True


# ============================================================================
# Whisper transcription
# ============================================================================

def get_whisper_model(model_size: str):
    global _whisper_model, _whisper_model_name
    
    if _whisper_model is not None and _whisper_model_name == model_size:
        log(f"Using cached Whisper model: {model_size}", 2)
        return _whisper_model
    
    import whisper
    
    cache_dir = Path.home() / ".cache" / "whisper"
    model_file = cache_dir / f"{model_size}.pt"
    
    if model_file.exists():
        log(f"Loading Whisper model: {model_size}", 2)
        log(f"Cache: {model_file}", 3)
    else:
        log(f"Downloading Whisper model: {model_size}", 2)
        log(f"This may take a few minutes...", 3)
    
    start_time = datetime.now()
    
    _whisper_model = whisper.load_model(model_size)
    _whisper_model_name = model_size
    
    elapsed = (datetime.now() - start_time).total_seconds()
    log(f"Model loaded in {elapsed:.1f}s", 2)
    
    # Log device info
    import torch
    if torch.cuda.is_available():
        log(f"Device: CUDA ({torch.cuda.get_device_name(0)})", 2)
    elif torch.backends.mps.is_available():
        log("Device: Apple Silicon (MPS)", 2)
    else:
        log("Device: CPU", 2)
    
    return _whisper_model


def transcribe_audio(
    audio_path: Path,
    model_size: str,
    language: Optional[str],
    word_timestamps: bool = True,
    beam_size: int = 5,
    initial_prompt: Optional[str] = None,
) -> dict:
    """Transcribe audio using Whisper."""
    model = get_whisper_model(model_size)
    
    duration = get_media_duration(audio_path)
    log(f"Starting transcription...", 2)
    log(f"Audio duration: {format_duration(duration)}", 2)
    log(f"Beam size: {beam_size}", 2)
    log(f"Word timestamps: {word_timestamps}", 2)
    
    opts = {
        "word_timestamps": word_timestamps,
        "verbose": False,
        "beam_size": beam_size,
    }
    
    if language and language.lower() not in ("auto", ""):
        opts["language"] = language
        log(f"Language: {language} (specified)", 2)
    else:
        log("Language: auto-detect", 2)
    
    if initial_prompt:
        opts["initial_prompt"] = initial_prompt
        log(f"Prompt: {initial_prompt[:50]}{'...' if len(initial_prompt) > 50 else ''}", 2)
    
    start_time = datetime.now()
    result = model.transcribe(str(audio_path), **opts)
    elapsed = (datetime.now() - start_time).total_seconds()
    
    detected_lang = result.get("language", "unknown")
    num_segments = len(result.get("segments", []))
    total_words = sum(len(s.get("text", "").split()) for s in result.get("segments", []))
    
    log(f"DONE: {num_segments} segments, {total_words} words", 2)
    log(f"Detected language: {detected_lang}", 2)
    log(f"Processing time: {elapsed:.1f}s ({duration/elapsed:.1f}x realtime)" if elapsed > 0 else "Processing time: instant", 2)
    
    return result


# ============================================================================
# pyannote diarization
# ============================================================================

def get_pyannote_pipeline(hf_token: str):
    """Get pyannote pipeline with caching."""
    global _pyannote_pipeline
    
    if _pyannote_pipeline is not None:
        log("Using cached pyannote pipeline", 2)
        return _pyannote_pipeline
    
    from pyannote.audio import Pipeline
    
    log("Loading pyannote speaker-diarization-3.1...", 2)
    start_time = datetime.now()
    
    _pyannote_pipeline = Pipeline.from_pretrained(
        "pyannote/speaker-diarization-3.1",
        token=hf_token
    )
    
    # Use GPU if available
    import torch
    if torch.backends.mps.is_available():
        _pyannote_pipeline = _pyannote_pipeline.to(torch.device("mps"))
        log("Device: Apple Silicon (MPS)", 2)
    elif torch.cuda.is_available():
        _pyannote_pipeline = _pyannote_pipeline.to(torch.device("cuda"))
        log(f"Device: CUDA ({torch.cuda.get_device_name(0)})", 2)
    else:
        log("Device: CPU (this will be slow)", 2)
    
    elapsed = (datetime.now() - start_time).total_seconds()
    log(f"Pipeline loaded in {elapsed:.1f}s", 2)
    
    return _pyannote_pipeline


def run_diarization(
    audio_path: Path,
    hf_token: str,
    num_speakers: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """Run speaker diarization on audio using soundfile for loading."""
    import soundfile as sf
    import torch
    
    pipeline = get_pyannote_pipeline(hf_token)
    
    duration = get_media_duration(audio_path)
    log("Starting speaker diarization...", 2)
    log(f"Audio duration: {format_duration(duration)}", 2)
    
    if num_speakers:
        log(f"Expected speakers: {num_speakers}", 2)
    else:
        log("Speakers: auto-detect", 2)
    
    start_time = datetime.now()
    
    # Load audio with soundfile (bypasses torchcodec issues)
    waveform, sample_rate = sf.read(str(audio_path), dtype='float32')
    waveform = torch.from_numpy(waveform)
    if waveform.ndim == 1:
        waveform = waveform.unsqueeze(0)
    else:
        waveform = waveform.T
    
    audio_input = {"waveform": waveform, "sample_rate": sample_rate}
    
    # Run diarization
    if num_speakers:
        result = pipeline(audio_input, num_speakers=num_speakers)
    else:
        result = pipeline(audio_input)
    
    elapsed = (datetime.now() - start_time).total_seconds()
    
    # Convert to list - handle both old and new pyannote API
    segments = []
    
    # New API (pyannote 3.x): result.speaker_diarization
    if hasattr(result, 'speaker_diarization'):
        diarization = result.speaker_diarization
    else:
        diarization = result
    
    for turn, _, speaker in diarization.itertracks(yield_label=True):
        segments.append({
            "start": turn.start,
            "end": turn.end,
            "speaker": speaker
        })
    
    # Statistics
    speakers = sorted(set(s["speaker"] for s in segments))
    speaker_times = {}
    for s in segments:
        spk = s["speaker"]
        speaker_times[spk] = speaker_times.get(spk, 0) + (s["end"] - s["start"])
    
    log(f"DONE: {len(speakers)} speakers detected", 2)
    for spk in speakers:
        time_pct = speaker_times[spk] / duration * 100 if duration > 0 else 0
        log(f"  {spk}: {format_duration(speaker_times[spk])} ({time_pct:.1f}%)", 2)
    log(f"Processing time: {elapsed:.1f}s ({duration/elapsed:.1f}x realtime)" if elapsed > 0 else "Processing time: instant", 2)
    
    return segments


def assign_speakers_to_segments(
    whisper_segments: List[dict],
    diarization_segments: List[dict],
) -> List[dict]:
    """Match Whisper transcription with pyannote speaker labels."""
    log("Merging transcription with speaker labels...", 2)
    
    result = []
    
    for seg in whisper_segments:
        text = seg.get("text", "").strip()
        if not text:
            continue
        
        seg_start = seg["start"]
        seg_end = seg["end"]
        
        # Calculate overlap time for each speaker
        speaker_times: Dict[str, float] = {}
        
        for diar in diarization_segments:
            overlap_start = max(seg_start, diar["start"])
            overlap_end = min(seg_end, diar["end"])
            
            if overlap_start < overlap_end:
                overlap_duration = overlap_end - overlap_start
                speaker = diar["speaker"]
                speaker_times[speaker] = speaker_times.get(speaker, 0) + overlap_duration
        
        # Select speaker with maximum overlap
        if speaker_times:
            speaker = max(speaker_times, key=speaker_times.get)
        else:
            speaker = "UNKNOWN"
        
        result.append({
            "start": seg_start,
            "end": seg_end,
            "speaker": speaker,
            "text": text,
        })
    
    log(f"DONE: {len(result)} segments with speaker labels", 2)
    
    return result


# ============================================================================
# Save results
# ============================================================================

def save_json(segments: List[dict], output_path: Path, metadata: dict) -> Path:
    """Save full data as JSON."""
    data = {
        "metadata": metadata,
        "segments": segments
    }
    
    json_path = output_path.with_suffix(".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    log(f"Saved: {json_path.name} ({format_size(json_path.stat().st_size)})", 2)
    return json_path


def save_srt(segments: List[dict], output_path: Path, include_speaker: bool = True) -> Path:
    """Save SRT subtitles."""
    lines = []
    
    for idx, seg in enumerate(segments, 1):
        start = seg["start"]
        end = seg["end"]
        text = seg["text"]
        speaker = seg.get("speaker", "")
        
        lines.append(str(idx))
        lines.append(f"{format_srt_timestamp(start)} --> {format_srt_timestamp(end)}")
        
        if include_speaker and speaker and speaker != "UNKNOWN":
            lines.append(f"[{speaker}] {text}")
        else:
            lines.append(text)
        
        lines.append("")
    
    srt_path = output_path.with_suffix(".srt")
    with open(srt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    
    log(f"Saved: {srt_path.name} ({len(segments)} subtitles)", 2)
    return srt_path


def save_txt(segments: List[dict], output_path: Path, include_speaker: bool = True) -> Path:
    """Save plain text."""
    lines = []
    
    for seg in segments:
        text = seg["text"]
        speaker = seg.get("speaker", "")
        
        if include_speaker and speaker and speaker != "UNKNOWN":
            lines.append(f"[{speaker}] {text}")
        else:
            lines.append(text)
    
    txt_path = output_path.with_suffix(".txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    
    total_words = sum(len(seg["text"].split()) for seg in segments)
    log(f"Saved: {txt_path.name} ({total_words} words)", 2)
    return txt_path


def sanitize_sheet_name(name: str) -> str:
    """Sanitize string for use as Excel sheet name (max 31 chars, no special chars)."""
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name[:31]


def save_xlsx_single(
    segments: List[dict],
    output_path: Path,
    metadata: dict,
    include_speaker_col: bool = True,
) -> Path:
    """Save Excel file for single video."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    
    # Use video name as sheet name
    video_name = metadata.get("source_file", "Transcription")
    if video_name:
        video_name = Path(video_name).stem
    ws.title = sanitize_sheet_name(video_name or "Transcription")
    
    # Headers - Timecode first
    if include_speaker_col:
        headers = ["Timecode", "Start", "End", "Duration", "Speaker", "Text"]
    else:
        headers = ["Timecode", "Start", "End", "Duration", "Text"]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    row_num = 2
    for seg in segments:
        start = seg["start"]
        end = seg["end"]
        duration = end - start
        
        col = 1
        # Timecode column first (same as Start but positioned first for easy navigation)
        ws.cell(row=row_num, column=col, value=format_timestamp(start)); col += 1
        ws.cell(row=row_num, column=col, value=format_timestamp(start)); col += 1
        ws.cell(row=row_num, column=col, value=format_timestamp(end)); col += 1
        ws.cell(row=row_num, column=col, value=f"{duration:.1f}s"); col += 1
        
        if include_speaker_col:
            ws.cell(row=row_num, column=col, value=seg.get("speaker", "")); col += 1
        
        ws.cell(row=row_num, column=col, value=seg["text"])
        row_num += 1
    
    # Column widths
    ws.column_dimensions["A"].width = 10  # Timecode
    ws.column_dimensions["B"].width = 10  # Start
    ws.column_dimensions["C"].width = 10  # End
    ws.column_dimensions["D"].width = 10  # Duration
    if include_speaker_col:
        ws.column_dimensions["E"].width = 15  # Speaker
        ws.column_dimensions["F"].width = 80  # Text
    else:
        ws.column_dimensions["E"].width = 80  # Text
    
    ws.freeze_panes = "A2"
    
    # Statistics sheet
    ws_stats = wb.create_sheet("Statistics")
    
    total_duration = max((seg["end"] for seg in segments), default=0)
    total_words = sum(len(seg["text"].split()) for seg in segments)
    speakers = sorted(set(seg.get("speaker", "") for seg in segments if seg.get("speaker")))
    wpm = (total_words / total_duration * 60) if total_duration > 0 else 0
    
    stats = [
        ("Source File", metadata.get("source_file", "")),
        ("Processed", metadata.get("processed_at", "")),
        ("Whisper Model", metadata.get("model", "")),
        ("Language", metadata.get("detected_language", metadata.get("language", ""))),
        ("", ""),
        ("Segments", len(segments)),
        ("Speakers", len(speakers)),
        ("Speaker List", ", ".join(speakers) if speakers else "—"),
        ("Duration", format_duration(total_duration)),
        ("Words", total_words),
        ("Words/min", f"{wpm:.1f}"),
    ]
    
    for row, (label, value) in enumerate(stats, 1):
        ws_stats.cell(row=row, column=1, value=label)
        ws_stats.cell(row=row, column=2, value=value)
    
    ws_stats.column_dimensions["A"].width = 20
    ws_stats.column_dimensions["B"].width = 50
    
    xlsx_path = output_path.with_suffix(".xlsx")
    wb.save(xlsx_path)
    
    log(f"Saved: {xlsx_path.name} ({len(segments)} rows)", 2)
    return xlsx_path


def save_xlsx_combined(
    all_results: List[Dict[str, Any]],
    output_path: Path,
    metadata: dict,
    include_video_col: bool = True,
    include_speaker_col: bool = True,
    include_index_col: bool = True,
    folder_name: Optional[str] = None,
) -> Path:
    """Save combined Excel file for multiple videos."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    
    # Use folder name as sheet name
    sheet_name = folder_name or "Transcription"
    ws.title = sanitize_sheet_name(sheet_name)
    
    # Headers - Timecode first
    headers = ["Timecode"]  # Always first
    if include_index_col:
        headers.append("#")
    if include_video_col:
        headers.append("Video")
    headers.extend(["Start", "End", "Duration"])
    if include_speaker_col:
        headers.append("Speaker")
    headers.append("Text")
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    video_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    video_font = Font(bold=True)
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    row_num = 2
    
    for video_idx, result in enumerate(all_results, 1):
        video_name = result["video_name"]
        segments = result["segments"]
        
        if not segments:
            continue
        
        first_row = True
        video_has_content = False
        
        for seg in segments:
            col = 1
            
            # Timecode always first
            ws.cell(row=row_num, column=col, value=format_timestamp(seg["start"])); col += 1
            
            if include_index_col:
                idx_cell = ws.cell(row=row_num, column=col)
                if first_row:
                    idx_cell.value = video_idx
                    idx_cell.fill = video_fill
                    idx_cell.font = video_font
                    idx_cell.alignment = Alignment(horizontal="center")
                col += 1
            
            if include_video_col:
                video_cell = ws.cell(row=row_num, column=col)
                if first_row:
                    video_cell.value = video_name
                    video_cell.fill = video_fill
                    video_cell.font = video_font
                    first_row = False
                col += 1
            
            ws.cell(row=row_num, column=col, value=format_timestamp(seg["start"])); col += 1
            ws.cell(row=row_num, column=col, value=format_timestamp(seg["end"])); col += 1
            ws.cell(row=row_num, column=col, value=f"{seg['end'] - seg['start']:.1f}s"); col += 1
            
            if include_speaker_col:
                ws.cell(row=row_num, column=col, value=seg.get("speaker", "")); col += 1
            
            ws.cell(row=row_num, column=col, value=seg["text"])
            
            row_num += 1
            video_has_content = True
        
        if video_has_content:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num - 1, column=col).border = thin_border
    
    # Column widths - adjust for new Timecode column
    col_idx = 1
    ws.column_dimensions["A"].width = 10  # Timecode
    col_idx += 1
    
    if include_index_col:
        ws.column_dimensions[chr(64 + col_idx)].width = 5  # #
        col_idx += 1
    if include_video_col:
        ws.column_dimensions[chr(64 + col_idx)].width = 25  # Video
        col_idx += 1
    
    for _ in range(3):  # Start, End, Duration
        ws.column_dimensions[chr(64 + col_idx)].width = 10
        col_idx += 1
    
    if include_speaker_col:
        ws.column_dimensions[chr(64 + col_idx)].width = 15  # Speaker
        col_idx += 1
    
    ws.column_dimensions[chr(64 + col_idx)].width = 80  # Text
    
    ws.freeze_panes = "A2"
    
    # Statistics sheet
    ws_stats = wb.create_sheet("Statistics")
    
    total_segments = sum(len(r["segments"]) for r in all_results)
    total_words = sum(sum(len(s["text"].split()) for s in r["segments"]) for r in all_results)
    all_speakers = set()
    for r in all_results:
        for s in r["segments"]:
            if s.get("speaker"):
                all_speakers.add(s["speaker"])
    
    stats_rows = [
        ("SUMMARY", ""),
        ("Project", folder_name or "—"),
        ("Processed", metadata.get("processed_at", "")),
        ("Whisper Model", metadata.get("model", "")),
        ("Videos", len(all_results)),
        ("Total Segments", total_segments),
        ("Total Words", total_words),
        ("Total Speakers", len(all_speakers)),
        ("", ""),
        ("PER VIDEO", ""),
    ]
    
    for idx, result in enumerate(all_results, 1):
        segments = result["segments"]
        if segments:
            dur = max(s["end"] for s in segments)
            words = sum(len(s["text"].split()) for s in segments)
            spk = sorted(set(s.get("speaker", "") for s in segments if s.get("speaker")))
            stats_rows.append((
                f"#{idx} {result['video_name']}",
                f"{format_duration(dur)}, {len(segments)} seg, {words} words, speakers: {', '.join(spk) or '—'}"
            ))
    
    for row, (label, value) in enumerate(stats_rows, 1):
        cell_a = ws_stats.cell(row=row, column=1, value=label)
        ws_stats.cell(row=row, column=2, value=value)
        if label in ("SUMMARY", "PER VIDEO"):
            cell_a.font = Font(bold=True)
    
    ws_stats.column_dimensions["A"].width = 35
    ws_stats.column_dimensions["B"].width = 70
    
    xlsx_path = output_path.with_suffix(".xlsx")
    wb.save(xlsx_path)
    
    log(f"Saved: {xlsx_path.name} ({total_segments} rows)", 2)
    return xlsx_path


# ============================================================================
# Video processing
# ============================================================================

def process_video(
    video_path: Path,
    model_size: str,
    language: Optional[str],
    num_speakers: Optional[int],
    hf_token: str,
    keep_audio: bool,
    word_timestamps: bool,
    beam_size: int,
    initial_prompt: Optional[str],
    output_dir: Optional[Path] = None,
) -> Optional[Dict[str, Any]]:
    """
    Full video processing pipeline:
    1. Extract audio
    2. Transcribe (Whisper)
    3. Diarize (pyannote)
    4. Merge and save
    
    Args:
        output_dir: If specified, results go into output_dir/video_name/
                    If None, results go into video_path.parent/video_name/ (legacy)
    """
    video_path = video_path.resolve()
    video_name = video_path.stem
    
    if output_dir:
        result_dir = output_dir / video_name
    else:
        result_dir = video_path.parent / video_name
    
    print("")
    print("=" * 70)
    log(f"VIDEO: {video_path.name}")
    print("=" * 70)
    
    # Video info
    info = get_media_info(video_path)
    log(f"Path: {video_path.parent}")
    log(f"Size: {format_size(info['size'])}")
    log(f"Duration: {format_duration(info['duration'])}")
    if info.get('video_codec'):
        log(f"Video: {info['video_codec']} {info.get('width', '?')}x{info.get('height', '?')}")
    log(f"Output: {result_dir}/")
    print("")
    
    result_dir.mkdir(parents=True, exist_ok=True)
    
    audio_path = result_dir / f"{video_name}.wav"
    output_base = result_dir / video_name
    
    metadata = {
        "source_file": video_path.name,
        "source_path": str(video_path),
        "processed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "model": model_size,
        "language": language or "auto",
        "sample_rate": AUDIO_SAMPLE_RATE,
        "num_speakers": num_speakers,
    }
    
    pipeline_start = datetime.now()
    
    # === PHASE 1: Audio extraction ===
    print("-" * 70)
    log("PHASE 1: Audio Extraction")
    print("-" * 70)
    
    if audio_path.exists() and audio_path.stat().st_size > 1000:
        log(f"Audio already exists: {audio_path.name}", 1)
        log(f"Size: {format_size(audio_path.stat().st_size)}", 2)
        log(f"Duration: {format_duration(get_media_duration(audio_path))}", 2)
    else:
        if not extract_audio(video_path, audio_path):
            return None
    print("")
    
    # === PHASE 2: Transcription ===
    print("-" * 70)
    log("PHASE 2: Transcription (Whisper)")
    print("-" * 70)
    
    whisper_result = transcribe_audio(
        audio_path, model_size, language,
        word_timestamps=word_timestamps,
        beam_size=beam_size,
        initial_prompt=initial_prompt,
    )
    
    detected_language = whisper_result.get("language", "unknown")
    metadata["detected_language"] = detected_language
    print("")
    
    # === PHASE 3: Diarization ===
    print("-" * 70)
    log("PHASE 3: Speaker Diarization (pyannote)")
    print("-" * 70)
    
    diarization_segments = run_diarization(audio_path, hf_token, num_speakers)
    print("")
    
    # === PHASE 4: Merge and save ===
    print("-" * 70)
    log("PHASE 4: Merge & Save")
    print("-" * 70)
    
    segments = assign_speakers_to_segments(
        whisper_result.get("segments", []),
        diarization_segments
    )
    
    # Count speakers
    speakers = sorted(set(s.get("speaker", "") for s in segments if s.get("speaker")))
    metadata["speakers"] = speakers
    
    # Always include speaker column (we require diarization)
    include_speaker = True
    
    log("Saving files:", 1)
    save_json(segments, output_base, metadata)
    save_srt(segments, output_base, include_speaker=include_speaker)
    save_txt(segments, output_base, include_speaker=include_speaker)
    save_xlsx_single(segments, output_base, metadata, include_speaker_col=include_speaker)
    
    if not keep_audio and audio_path.exists():
        audio_path.unlink()
        log(f"Deleted: {audio_path.name}", 2)
    else:
        log(f"Kept: {audio_path.name} ({format_size(audio_path.stat().st_size)})", 2)
    
    print("")
    
    # Summary
    pipeline_elapsed = (datetime.now() - pipeline_start).total_seconds()
    duration = max((s["end"] for s in segments), default=0)
    words = sum(len(s["text"].split()) for s in segments)
    
    print("=" * 70)
    log("COMPLETED")
    print("=" * 70)
    log(f"Segments: {len(segments)}")
    log(f"Words: {words}")
    log(f"Duration: {format_duration(duration)}")
    log(f"Speakers: {', '.join(speakers) if speakers else 'none detected'}")
    log(f"Total time: {format_duration(pipeline_elapsed)} ({duration/pipeline_elapsed:.1f}x realtime)" if pipeline_elapsed > 0 else "Total time: instant")
    log(f"Output: {result_dir}/")
    
    return {
        "video_name": video_name,
        "segments": segments,
        "speakers": speakers,
        "metadata": metadata,
    }


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Video transcription with speaker diarization (Whisper + pyannote)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Version: {VERSION}

Examples:
  Preflight check only:
    %(prog)s --preflight

  Single video:
    %(prog)s video.mp4
    %(prog)s video.mp4 --language en --num-speakers 2

  Folder of videos:
    %(prog)s /path/to/folder --folder
    %(prog)s /path/to/folder --folder --name MyProject

  Skip preflight (not recommended):
    %(prog)s video.mp4 --skip-preflight

For speaker diarization:
  1. export HF_TOKEN="hf_xxx"  (or: huggingface-cli login)
  2. Accept licenses:
     - https://huggingface.co/pyannote/speaker-diarization-3.1
     - https://huggingface.co/pyannote/segmentation-3.0
        """
    )
    
    parser.add_argument("input", nargs="*", help="Video file(s) or folder (with --folder)")
    parser.add_argument("-f", "--folder", action="store_true", help="Folder mode → combined table")
    parser.add_argument("--name", type=str, default=None,
                       help="Project name for output folder (default: auto-detect from folder name)")
    
    parser.add_argument("-m", "--model", default=DEFAULT_WHISPER_MODEL,
                       choices=["tiny", "base", "small", "medium", "large", "large-v2", "large-v3"],
                       help=f"Whisper model (default: {DEFAULT_WHISPER_MODEL})")
    parser.add_argument("-l", "--language", default=None, help="Language code (ru, en, ar) — speeds up")
    
    parser.add_argument("-n", "--num-speakers", type=int, default=None,
                       help="Number of speakers (helps pyannote)")
    
    parser.add_argument("--beam-size", type=int, default=5, help="Beam size (1-10, lower=faster)")
    parser.add_argument("--no-word-timestamps", action="store_true", help="Disable word-level timestamps")
    parser.add_argument("--prompt", type=str, default=None, help="Initial prompt for Whisper")
    
    parser.add_argument("--no-keep-audio", action="store_true", help="Delete WAV after processing")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be processed")
    
    parser.add_argument("--preflight", action="store_true", help="Run preflight checks only")
    parser.add_argument("--skip-preflight", action="store_true", help="Skip preflight checks (not recommended)")
    
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    
    args = parser.parse_args()
    
    # Determine input path for preflight
    input_path: Optional[Path] = None
    is_folder = args.folder
    
    if args.input:
        input_path = Path(args.input[0]).resolve()
    
    # Preflight only mode
    if args.preflight:
        result = run_preflight_checks(input_path, is_folder)
        sys.exit(0 if result.ok else 1)
    
    # No input provided
    if not args.input:
        parser.print_help()
        print("\nError: No input specified. Use --preflight to check setup.")
        sys.exit(1)
    
    # Run preflight unless skipped
    if not args.skip_preflight:
        result = run_preflight_checks(input_path, is_folder)
        if not result.ok:
            sys.exit(1)
        print("")  # Extra newline before processing
    
    # Get HF token (already validated in preflight)
    hf_token = get_hf_token()
    if not hf_token:
        print("ERROR: HF_TOKEN not found. Run with --preflight to see details.")
        sys.exit(1)
    
    # Collect videos
    videos: List[Path] = []
    folder: Optional[Path] = None
    output_dir: Optional[Path] = None
    project_name: Optional[str] = None
    
    if args.folder:
        if len(args.input) != 1:
            print("ERROR: In --folder mode, specify exactly one folder")
            sys.exit(1)
        
        folder = Path(args.input[0]).resolve()
        
        if not folder.exists() or not folder.is_dir():
            print(f"ERROR: Folder not found: {folder}")
            sys.exit(1)
        
        videos = find_media_in_folder(folder)
        
        if not videos:
            print(f"ERROR: No videos found in: {folder}")
            sys.exit(1)
        
        # Determine project name and output directory
        project_name = resolve_project_name(folder, args.name)
        output_dir = folder / f"{project_name}_transcription"
        output_dir.mkdir(parents=True, exist_ok=True)
    else:
        for p in args.input:
            path = Path(p).resolve()
            
            if path.is_dir():
                print(f"ERROR: '{path.name}' is a folder.")
                print(f"")
                print(f"To process all videos in a folder, use --folder (or -f):")
                print(f"  python {sys.argv[0]} '{path}' --folder")
                sys.exit(1)
            
            if not path.exists():
                print(f"ERROR: File not found: {path}")
                sys.exit(1)
            
            if not is_media_file(path):
                print(f"ERROR: Not a supported media file: {path.name}")
                print(f"       Supported: {', '.join(sorted(MEDIA_EXTS))}")
                sys.exit(1)
            
            videos.append(path)
    
    # Header
    print("")
    print("╔" + "═" * 68 + "╗")
    print("║" + f"  WHISPER BATCH v{VERSION} + SPEAKER DIARIZATION".center(68) + "║")
    print("╚" + "═" * 68 + "╝")
    print("")
    
    log("Configuration:")
    if args.folder:
        log(f"Mode: FOLDER", 1)
        log(f"Folder: {folder}", 1)
        log(f"Project name: {project_name}", 1)
        log(f"Output: {output_dir}/", 1)
    else:
        log(f"Mode: FILES", 1)
    
    log(f"Videos: {len(videos)}", 1)
    log(f"Whisper model: {args.model}", 1)
    log(f"Language: {args.language or 'auto-detect'}", 1)
    log(f"Beam size: {args.beam_size}", 1)
    log(f"Word timestamps: {not args.no_word_timestamps}", 1)
    
    if args.num_speakers:
        log(f"Expected speakers: {args.num_speakers}", 1)
    
    log(f"Diarization: ENABLED (pyannote)", 1)
    
    if args.prompt:
        log(f"Prompt: {args.prompt[:40]}{'...' if len(args.prompt) > 40 else ''}", 1)
    
    # Dry run
    if args.dry_run:
        print("")
        log("Videos to process:")
        total_duration = 0
        for i, v in enumerate(videos, 1):
            dur = get_media_duration(v)
            total_duration += dur
            log(f"{i}. {v.name} ({format_duration(dur)})", 1)
        print("")
        log(f"Total duration: {format_duration(total_duration)}")
        log("(--dry-run: no processing performed)")
        sys.exit(0)
    
    # Process videos
    start_time = datetime.now()
    all_results: List[Dict[str, Any]] = []
    success_count = 0
    fail_count = 0
    
    for i, video_path in enumerate(videos, 1):
        if len(videos) > 1:
            print("")
            print("╔" + "═" * 68 + "╗")
            print(f"║  [{i}/{len(videos)}] Processing video...".ljust(69) + "║")
            print("╚" + "═" * 68 + "╝")
        
        try:
            result = process_video(
                video_path,
                args.model,
                args.language,
                args.num_speakers,
                hf_token,
                keep_audio=not args.no_keep_audio,
                word_timestamps=not args.no_word_timestamps,
                beam_size=args.beam_size,
                initial_prompt=args.prompt,
                output_dir=output_dir,
            )
            
            if result:
                all_results.append(result)
                success_count += 1
            else:
                fail_count += 1
                
        except KeyboardInterrupt:
            print("\n")
            log("INTERRUPTED by user (Ctrl+C)")
            sys.exit(1)
        except Exception as e:
            log(f"ERROR: {e}")
            fail_count += 1
    
    # Combined table for folder mode
    if args.folder and all_results:
        print("")
        print("=" * 70)
        log("Creating combined table...")
        print("=" * 70)
        
        # Always include columns (we require speakers)
        include_video_col = len(all_results) > 1
        include_speaker_col = True
        include_index_col = len(all_results) > 1
        
        log(f"Include # column: {include_index_col}", 1)
        log(f"Include Video column: {include_video_col}", 1)
        log(f"Include Speaker column: {include_speaker_col}", 1)
        
        xlsx_path = output_dir / f"{project_name}_transcription.xlsx"
        meta = {
            "processed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "model": args.model,
        }
        
        log(f"Project name: {project_name}", 1)
        
        save_xlsx_combined(
            all_results, xlsx_path, meta,
            include_video_col=include_video_col,
            include_speaker_col=include_speaker_col,
            include_index_col=include_index_col,
            folder_name=project_name,
        )
    
    # Final summary
    elapsed = (datetime.now() - start_time).total_seconds()
    
    print("")
    print("╔" + "═" * 68 + "╗")
    print("║" + "  BATCH COMPLETED".center(68) + "║")
    print("╚" + "═" * 68 + "╝")
    print("")
    log(f"Successful: {success_count}")
    if fail_count > 0:
        log(f"Failed: {fail_count}")
    log(f"Total time: {format_duration(elapsed)}")
    
    if all_results:
        total_segments = sum(len(r["segments"]) for r in all_results)
        total_words = sum(sum(len(s["text"].split()) for s in r["segments"]) for r in all_results)
        total_duration = sum(max((s["end"] for s in r["segments"]), default=0) for r in all_results)
        all_speakers = set()
        for r in all_results:
            all_speakers.update(r.get("speakers", []))
        
        log(f"Total segments: {total_segments}")
        log(f"Total words: {total_words}")
        log(f"Total audio: {format_duration(total_duration)}")
        log(f"Speakers found: {', '.join(sorted(all_speakers)) if all_speakers else 'none'}")
        
        if output_dir:
            log(f"Output: {output_dir}/")
    
    sys.exit(0 if fail_count == 0 else 1)


if __name__ == "__main__":
    main()
