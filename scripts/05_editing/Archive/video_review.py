#!/usr/bin/env python3
"""
video_review.py — Video review pipeline for finished YouTube videos

Analyze a final edited video before publishing:
  - Extract frames (scene detection + regular sampling)
  - OCR all on-screen text (titles, lower thirds, graphics)
  - Vision LLM analysis of key scenes (via Ollama)
  - Speech transcription with speaker diarization (Whisper + pyannote)
  - Audio levels check (LUFS, clipping, balance)
  - Cross-check with source transcript (optional)
  - Generate comprehensive review report (Excel + Markdown + JSON)

Features:
  - Automatic preflight checks before processing
  - Work estimate upfront (frames, time, disk space)
  - Scene change detection + regular frame sampling
  - OCR text grouping (merges duplicate detections into appearances)
  - Jump cut / black frame / flash frame detection
  - Tempo analysis (shot duration, speaker balance, content mix)
  - Structure check (intro, outro, end screen)
  - Detailed progress with ETA in terminal
  - File logging to logs/ directory
  - Resume support (skip completed phases)
  - Flexible modes: --quick, --full, --skip-speech, --skip-vision

Modes:
  1. Preflight only: python video_review.py --preflight
  2. Quick review:   python video_review.py video.mp4 --quick
  3. Full review:    python video_review.py video.mp4
  4. With transcript: python video_review.py video.mp4 --transcript path/to/apply_names.txt

Output structure:
  video_review/
  ├── frames/
  │   ├── frame_0001_000005_00m05s.jpg
  │   └── ...
  ├── _intermediate/
  │   ├── frames.json
  │   ├── ocr.json
  │   └── speech.json
  ├── video_review.xlsx            (main report)
  ├── video_review.md              (Markdown report)
  ├── video_review.json            (machine-readable)
  └── logs/
      └── review_20260217_143000.log

Requirements:
    pip install easyocr Pillow openpyxl soundfile numpy
    pip install openai-whisper torch pyannote.audio   # for speech
    brew install ffmpeg

    # Ollama + vision model (for --full mode):
    ollama pull minicpm-v

    # HuggingFace token (for speaker diarization):
    export HF_TOKEN="hf_xxx"
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import shutil
import subprocess
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ============================================================================
# Configuration
# ============================================================================

VERSION = "1.3.0"
VIDEO_EXTS = {".mp4", ".mov", ".m4v", ".mts", ".avi", ".mkv", ".webm"}
DEFAULT_WHISPER_MODEL = "large-v3"
DEFAULT_VISION_MODEL = "minicpm-v"
OLLAMA_URL = "http://localhost:11434"
AUDIO_SAMPLE_RATE = 16000  # 16kHz for Whisper
MIN_DISK_SPACE_GB = 2.0
HF_CONFIG_PATH = Path.home() / "YTAI" / "config" / "HuggingFace-yt-prod.conf"

# Frame extraction settings
REGULAR_INTERVAL_SEC = 3       # Extract a frame every N seconds
SCENE_THRESHOLD = 0.3          # ffmpeg scene detection threshold (0-1)
FRAME_JPEG_QUALITY = 85        # JPEG quality for extracted frames
MIN_FRAME_DIFF_SEC = 1.0       # Minimum gap between frames after dedup

# OCR settings
OCR_LANGUAGES = ["en", "ru"]   # EasyOCR language list
OCR_CONFIDENCE_THRESHOLD = 0.3 # Minimum OCR confidence to include
TEXT_GROUP_SIMILARITY = 0.85   # Threshold for grouping similar OCR text
TEXT_GROUP_MAX_GAP_SEC = 5.0   # Max gap between frames to group same text

# Vision LLM settings
VISION_TIMEOUT_SEC = 120       # Timeout per frame for vision LLM

# Audio analysis settings
SILENCE_THRESHOLD_DB = -40     # Below this = silence
LONG_SILENCE_SEC = 5.0         # Flag silences longer than this

# Jump cut detection
JUMP_CUT_MIN_DIFF = 0.15      # Minimum scene diff to consider a cut
JUMP_CUT_MAX_RETURN_SEC = 2.0 # If similar frame returns within this = jump cut
BLACK_FRAME_THRESHOLD = 10     # Average pixel value below this = black frame

# Model cache
_whisper_model = None
_whisper_model_name = None
_pyannote_pipeline = None
_easyocr_reader = None

# Global log file handle
_log_file = None


# ============================================================================
# Utilities
# ============================================================================

def format_timestamp(seconds: float) -> str:
    """Format as HH:MM:SS or MM:SS."""
    total_secs = int(round(seconds))
    h = total_secs // 3600
    m = (total_secs % 3600) // 60
    s = total_secs % 60
    if h > 0:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def format_timestamp_filename(seconds: float) -> str:
    """Format as 00m05s for filenames."""
    total_secs = int(round(seconds))
    m = total_secs // 60
    s = total_secs % 60
    return f"{m:02d}m{s:02d}s"


def format_duration(seconds: float) -> str:
    """Human-readable duration."""
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    if h > 0:
        return f"{h}h {m}m {s}s"
    elif m > 0:
        return f"{m}m {s}s"
    return f"{s}s"


def format_size(size_bytes: int) -> str:
    """Human-readable file size."""
    if size_bytes >= 1024 ** 3:
        return f"{size_bytes / (1024**3):.2f} GB"
    elif size_bytes >= 1024 ** 2:
        return f"{size_bytes / (1024**2):.1f} MB"
    elif size_bytes >= 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes} bytes"


def log(msg: str, indent: int = 0):
    """Print log message with timestamp. Also writes to log file if open."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    prefix = "  " * indent
    line = f"[{timestamp}] {prefix}{msg}"
    print(line)
    _write_to_log(line)


def out(msg: str):
    """Print message without timestamp. Also writes to log file."""
    print(msg)
    _write_to_log(msg)


def _write_to_log(line: str):
    """Write a line to the log file if open."""
    if _log_file:
        try:
            _log_file.write(line + "\n")
            _log_file.flush()
        except Exception:
            pass


def init_log_file(logs_dir: Path) -> None:
    """Open log file for writing."""
    global _log_file
    logs_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = logs_dir / f"review_{ts}.log"
    _log_file = open(log_path, "w", encoding="utf-8")
    log(f"Log file: {log_path}")


def close_log_file() -> None:
    """Close log file."""
    global _log_file
    if _log_file:
        _log_file.close()
        _log_file = None


def progress_bar(current: int, total: int, width: int = 30) -> str:
    """Generate a text progress bar."""
    if total == 0:
        return "[" + "=" * width + "]"
    filled = int(width * current / total)
    bar = "█" * filled + "░" * (width - filled)
    pct = current / total * 100
    return f"[{bar}] {current}/{total}  {pct:.0f}%"


class ETATracker:
    """Track processing speed and estimate remaining time."""

    def __init__(self):
        self.start_time = time.time()
        self.items_done = 0

    def update(self, items_done: int):
        self.items_done = items_done

    def get_speed(self) -> float:
        elapsed = time.time() - self.start_time
        if elapsed == 0 or self.items_done == 0:
            return 0.0
        return self.items_done / elapsed

    def get_eta(self, total: int) -> str:
        speed = self.get_speed()
        if speed == 0:
            return "calculating..."
        remaining = (total - self.items_done) / speed
        return format_duration(remaining)


# ============================================================================
# Intermediate state (resume support)
# ============================================================================

def save_intermediate(data: Any, output_dir: Path, name: str) -> Path:
    """Save intermediate results as JSON for resume."""
    inter_dir = output_dir / "_intermediate"
    inter_dir.mkdir(parents=True, exist_ok=True)
    path = inter_dir / f"{name}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, default=str)
    return path


def load_intermediate(output_dir: Path, name: str) -> Optional[Any]:
    """Load intermediate results if available."""
    path = output_dir / "_intermediate" / f"{name}.json"
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return None


# ============================================================================
# Preflight Checks
# ============================================================================

class PreflightResult:
    """Container for preflight check results."""

    def __init__(self):
        self.errors: List[Tuple[str, str]] = []
        self.warnings: List[Tuple[str, str]] = []
        self.passed: List[Tuple[str, str]] = []

    def add_error(self, check: str, message: str):
        self.errors.append((check, message))

    def add_warning(self, check: str, message: str):
        self.warnings.append((check, message))

    def add_passed(self, check: str, details: str = ""):
        self.passed.append((check, details))

    @property
    def ok(self) -> bool:
        return len(self.errors) == 0


def get_ffmpeg_version() -> Optional[str]:
    try:
        result = subprocess.run(
            ["ffmpeg", "-version"],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            first_line = result.stdout.split('\n')[0]
            parts = first_line.split()
            if len(parts) >= 3:
                return parts[2]
        return "installed"
    except Exception:
        return None


def get_package_version(package_name: str) -> Optional[str]:
    try:
        if package_name == "easyocr":
            import easyocr
            return getattr(easyocr, '__version__', 'installed')
        elif package_name == "Pillow":
            from PIL import Image
            import PIL
            return getattr(PIL, '__version__', 'installed')
        elif package_name == "openpyxl":
            import openpyxl
            return openpyxl.__version__
        elif package_name == "numpy":
            import numpy
            return numpy.__version__
        elif package_name == "openai-whisper":
            import whisper
            return getattr(whisper, '__version__', 'installed')
        elif package_name == "torch":
            import torch
            return torch.__version__
        elif package_name == "pyannote.audio":
            import pyannote.audio
            return getattr(pyannote.audio, '__version__', 'installed')
        elif package_name == "soundfile":
            import soundfile
            return getattr(soundfile, '__version__', 'installed')
        return None
    except ImportError:
        return None


def check_ollama_running() -> Optional[str]:
    """Check if Ollama is running and return version."""
    try:
        import urllib.request
        req = urllib.request.Request(f"{OLLAMA_URL}/api/version")
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
            return data.get("version", "running")
    except Exception:
        return None


def check_ollama_model(model_name: str) -> bool:
    """Check if a specific model is available in Ollama."""
    try:
        import urllib.request
        req = urllib.request.Request(f"{OLLAMA_URL}/api/tags")
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
            models = [m.get("name", "") for m in data.get("models", [])]
            for m in models:
                if m == model_name or m.startswith(model_name + ":"):
                    return True
            return False
    except Exception:
        return False


def get_hf_token() -> Optional[str]:
    """Get HuggingFace token from multiple sources."""
    token = os.environ.get("HF_TOKEN") or os.environ.get("HUGGINGFACE_TOKEN")
    if token:
        return token

    if HF_CONFIG_PATH.exists():
        try:
            with open(HF_CONFIG_PATH) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and 'HF_TOKEN' in line and '=' in line:
                        return line.split('=', 1)[1].strip().strip('"\'')
        except Exception:
            pass

    token_path = Path.home() / ".cache" / "huggingface" / "token"
    if token_path.exists():
        try:
            return token_path.read_text().strip()
        except Exception:
            pass

    return None


def get_disk_space_gb(path: Path) -> float:
    try:
        stat = shutil.disk_usage(path)
        return stat.free / (1024 ** 3)
    except Exception:
        return 0.0


def get_device_info() -> Tuple[str, bool]:
    try:
        import torch
        if torch.backends.mps.is_available():
            return "Apple Silicon (MPS)", True
        elif torch.cuda.is_available():
            return f"CUDA ({torch.cuda.get_device_name(0)})", True
        else:
            return "CPU", False
    except Exception:
        return "Unknown", False


def get_video_info(video_path: Path) -> Dict[str, Any]:
    """Get video metadata via ffprobe."""
    info = {
        "duration": 0.0,
        "width": 0,
        "height": 0,
        "fps": 0.0,
        "codec": "",
        "size_bytes": video_path.stat().st_size if video_path.exists() else 0,
        "audio_codec": "",
        "audio_channels": 0,
        "audio_sample_rate": 0,
        "has_audio": False,
        "is_valid": False,
    }

    try:
        cmd = [
            "ffprobe", "-v", "error",
            "-print_format", "json",
            "-show_format", "-show_streams",
            str(video_path)
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode != 0:
            info["probe_error"] = result.stderr.strip()
            return info

        data = json.loads(result.stdout)
        info["is_valid"] = True

        # Duration
        fmt = data.get("format", {})
        info["duration"] = float(fmt.get("duration", 0))

        # Video stream
        for stream in data.get("streams", []):
            if stream.get("codec_type") == "video":
                info["width"] = int(stream.get("width", 0))
                info["height"] = int(stream.get("height", 0))
                info["codec"] = stream.get("codec_name", "")
                fps_str = stream.get("r_frame_rate", "0/1")
                if "/" in fps_str:
                    num, den = fps_str.split("/")
                    info["fps"] = float(num) / float(den) if float(den) > 0 else 0
                break

        # Audio stream
        for stream in data.get("streams", []):
            if stream.get("codec_type") == "audio":
                info["audio_codec"] = stream.get("codec_name", "")
                info["audio_channels"] = int(stream.get("channels", 0))
                info["audio_sample_rate"] = int(stream.get("sample_rate", 0))
                info["has_audio"] = True
                break

    except Exception as e:
        info["probe_error"] = str(e)

    return info


def estimate_work(
    video_info: Dict[str, Any],
    mode: str,
    do_speech: bool,
    do_vision: bool,
    interval_sec: float,
) -> Dict[str, Any]:
    """Estimate work to be done: frame count, disk, time."""
    duration = video_info["duration"]
    has_audio = video_info.get("has_audio", True)

    est_regular_frames = int(duration / interval_sec) if interval_sec > 0 else 0
    est_scene_frames = int(duration / 10)  # rough: ~1 scene per 10 sec
    est_total_frames = est_regular_frames + est_scene_frames
    # After dedup, ~70% remain
    est_unique_frames = int(est_total_frames * 0.7)
    # Vision LLM frames
    est_vision_frames = int(est_unique_frames * 0.2) if do_vision else 0

    # Time estimates (seconds)
    t_frames = est_unique_frames * 0.15  # ~0.15s per frame extraction
    t_ocr = est_unique_frames * 0.5      # ~0.5s per frame OCR
    t_vision = est_vision_frames * 15    # ~15s per frame vision LLM
    t_speech = (duration * 0.8) if (do_speech and has_audio) else 0
    t_audio = 10 if has_audio else 0
    t_total = t_frames + t_ocr + t_vision + t_speech + t_audio

    # Disk estimate: ~50KB per JPEG frame + audio WAV (only if speech enabled)
    est_disk_mb = est_unique_frames * 50 / 1024
    if do_speech and has_audio:
        est_disk_mb += duration * 16000 * 2 / 1024 / 1024

    return {
        "est_regular_frames": est_regular_frames,
        "est_scene_frames": est_scene_frames,
        "est_unique_frames": est_unique_frames,
        "est_vision_frames": est_vision_frames,
        "est_time_sec": t_total,
        "est_disk_mb": est_disk_mb,
        "breakdown": {
            "frames": t_frames,
            "ocr": t_ocr,
            "vision": t_vision,
            "speech": t_speech,
            "audio": t_audio,
        },
    }


def run_preflight_checks(
    input_path: Optional[Path] = None,
    do_vision: bool = True,
    do_speech: bool = True,
) -> PreflightResult:
    """Run all preflight checks and return results."""
    result = PreflightResult()

    print("")
    print("╔" + "═" * 68 + "╗")
    print("║" + "  VIDEO REVIEW — PREFLIGHT CHECKS".center(68) + "║")
    print("╚" + "═" * 68 + "╝")
    print("")

    # ========================================
    # 1. System Dependencies
    # ========================================
    print("[1/7] System dependencies")

    ffmpeg_version = get_ffmpeg_version()
    if ffmpeg_version:
        result.add_passed("ffmpeg", ffmpeg_version)
        print(f"  ✅ ffmpeg: {ffmpeg_version}")
    else:
        result.add_error("ffmpeg", "Not installed")
        print(f"  ❌ ffmpeg: NOT INSTALLED")

    try:
        subprocess.run(["ffprobe", "-version"], capture_output=True, timeout=10)
        result.add_passed("ffprobe", "available")
        print(f"  ✅ ffprobe: available")
    except Exception:
        result.add_error("ffprobe", "Not installed")
        print(f"  ❌ ffprobe: NOT INSTALLED")

    print("")

    # ========================================
    # 2. Core Python Packages
    # ========================================
    print("[2/7] Core Python packages")

    core_packages = [
        ("easyocr", "OCR text extraction"),
        ("Pillow", "Image processing"),
        ("openpyxl", "Excel output"),
        ("numpy", "Array operations"),
    ]

    missing_packages = []
    for pkg_name, purpose in core_packages:
        version = get_package_version(pkg_name)
        if version:
            result.add_passed(pkg_name, version)
            print(f"  ✅ {pkg_name}: {version}")
        else:
            result.add_error(pkg_name, f"Not installed ({purpose})")
            missing_packages.append(pkg_name)
            print(f"  ❌ {pkg_name}: NOT INSTALLED ({purpose})")

    print("")

    # ========================================
    # 3. Speech Packages
    # ========================================
    print("[3/7] Speech packages")

    if not do_speech:
        print(f"  ℹ️  Skipped (--skip-speech)")
    else:
        speech_packages = [
            ("openai-whisper", "Transcription"),
            ("torch", "ML framework"),
            ("pyannote.audio", "Speaker diarization"),
            ("soundfile", "Audio loading"),
        ]
        for pkg_name, purpose in speech_packages:
            version = get_package_version(pkg_name)
            if version:
                result.add_passed(pkg_name, version)
                print(f"  ✅ {pkg_name}: {version}")
            else:
                result.add_error(pkg_name, f"Not installed ({purpose})")
                missing_packages.append(pkg_name)
                print(f"  ❌ {pkg_name}: NOT INSTALLED ({purpose})")

    print("")

    # ========================================
    # 4. Ollama + Vision Model
    # ========================================
    print("[4/7] Vision LLM (Ollama)")

    if not do_vision:
        print(f"  ℹ️  Skipped (--skip-vision / --quick)")
    else:
        ollama_version = check_ollama_running()
        if ollama_version:
            result.add_passed("ollama", ollama_version)
            print(f"  ✅ Ollama: running (v{ollama_version})")

            if check_ollama_model(DEFAULT_VISION_MODEL):
                result.add_passed(DEFAULT_VISION_MODEL, "available")
                print(f"  ✅ {DEFAULT_VISION_MODEL}: available")
            else:
                result.add_error(DEFAULT_VISION_MODEL, "Model not pulled")
                print(f"  ❌ {DEFAULT_VISION_MODEL}: NOT FOUND")
                print(f"     Run: ollama pull {DEFAULT_VISION_MODEL}")
        else:
            result.add_error("ollama", "Not running")
            print(f"  ❌ Ollama: NOT RUNNING")
            print(f"     Run: OLLAMA_MAX_VRAM=20g ollama serve")

    print("")

    # ========================================
    # 5. Authentication
    # ========================================
    print("[5/7] Authentication")

    hf_token = get_hf_token()
    if do_speech:
        if hf_token:
            masked = hf_token[:8] + "..." + hf_token[-4:] if len(hf_token) > 12 else "***"
            result.add_passed("HF_TOKEN", f"found ({masked})")
            print(f"  ✅ HF_TOKEN: found ({masked})")
        else:
            result.add_error("HF_TOKEN", "Not found (required for speaker diarization)")
            print(f"  ❌ HF_TOKEN: NOT FOUND")
    else:
        print(f"  ℹ️  Skipped (--skip-speech)")

    print("")

    # ========================================
    # 6. Input Validation
    # ========================================
    print("[6/7] Input validation")

    video_info = None
    if input_path is None:
        print(f"  ℹ️  No input specified (preflight only mode)")
    elif not input_path.exists():
        result.add_error("input_path", f"Path not found: {input_path}")
        print(f"  ❌ Path: NOT FOUND")
        print(f"     {input_path}")
    else:
        if input_path.suffix.lower() in VIDEO_EXTS:
            size = input_path.stat().st_size
            result.add_passed("video_file", f"{input_path.name} ({format_size(size)})")
            print(f"  ✅ Video file: {input_path.name} ({format_size(size)})")

            # Validate video is readable
            video_info = get_video_info(input_path)
            if video_info["is_valid"]:
                print(f"     Duration: {format_duration(video_info['duration'])}")
                print(f"     Resolution: {video_info['width']}×{video_info['height']}")
                print(f"     Codec: {video_info['codec']} @ {video_info['fps']:.1f} fps")
                if video_info["audio_codec"]:
                    print(f"     Audio: {video_info['audio_codec']}, "
                          f"{video_info['audio_channels']}ch, "
                          f"{video_info['audio_sample_rate']}Hz")
            else:
                err = video_info.get("probe_error", "Unknown error")
                result.add_error("video_probe", f"Cannot read video: {err}")
                print(f"  ❌ Video UNREADABLE: {err}")
        else:
            result.add_error("video_file", f"Not a video file: {input_path.suffix}")
            print(f"  ❌ Not a video file: {input_path.suffix}")
            print(f"     Supported: {', '.join(sorted(VIDEO_EXTS))}")

        free_space = get_disk_space_gb(input_path.parent)
        if free_space >= MIN_DISK_SPACE_GB:
            result.add_passed("disk_space", f"{free_space:.1f} GB free")
            print(f"  ✅ Disk space: {free_space:.1f} GB free")
        else:
            result.add_error("disk_space", f"Only {free_space:.1f} GB free")
            print(f"  ❌ Disk space: {free_space:.1f} GB free (need {MIN_DISK_SPACE_GB} GB)")

    print("")

    # ========================================
    # 7. GPU/Device
    # ========================================
    print("[7/7] Compute device")

    device_name, is_gpu = get_device_info()
    if is_gpu:
        result.add_passed("device", device_name)
        print(f"  ✅ {device_name}")
    else:
        result.add_warning("device", "CPU only (will be slow)")
        print(f"  ⚠️  {device_name} (processing will be slow)")

    print("")

    # ========================================
    # Summary
    # ========================================
    print("─" * 70)

    if result.ok:
        if result.warnings:
            print(f"✅ All critical checks passed ({len(result.warnings)} warnings)")
        else:
            print(f"✅ All checks passed! Ready to process.")
    else:
        print(f"❌ {len(result.errors)} error(s) found. Cannot proceed.")
        print("")
        print("Fix with:")
        print("")

        if missing_packages:
            pkg_list = " ".join(missing_packages)
            print(f"  pip install {pkg_list} --break-system-packages")

        if do_speech and not hf_token:
            print("")
            print("  # Set HuggingFace token:")
            print("  export HF_TOKEN='hf_your_token_here'")

        if do_vision and not check_ollama_running():
            print("")
            print("  # Start Ollama:")
            print("  OLLAMA_MAX_VRAM=20g ollama serve")
            print(f"  ollama pull {DEFAULT_VISION_MODEL}")

        if not ffmpeg_version:
            print("")
            print("  brew install ffmpeg")

    print("─" * 70)
    print("")

    return result


# ============================================================================
# Phase 1: Frame Extraction
# ============================================================================

def extract_frames_scene_detect(
    video_path: Path,
    threshold: float = SCENE_THRESHOLD,
) -> List[float]:
    """Extract scene change timestamps using ffmpeg scene detection.
    Returns list of timestamps (float seconds)."""
    log("Detecting scene changes...", 1)

    cmd = [
        "ffmpeg", "-hide_banner", "-i", str(video_path),
        "-vf", f"select='gt(scene,{threshold})',showinfo",
        "-vsync", "vfr",
        "-f", "null", "-"
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)

    scene_times = []
    for line in result.stderr.split('\n'):
        if 'pts_time:' in line:
            match = re.search(r'pts_time:([\d.]+)', line)
            if match:
                t = float(match.group(1))
                scene_times.append(t)

    log(f"Found {len(scene_times)} scene changes", 2)
    return scene_times


def make_regular_timestamps(
    duration: float,
    interval_sec: float = REGULAR_INTERVAL_SEC,
) -> List[float]:
    """Generate regular timestamps across video duration."""
    timestamps = []
    t = 0.0
    while t < duration:
        timestamps.append(t)
        t += interval_sec
    return timestamps


def merge_and_deduplicate(
    scene_times: List[float],
    regular_times: List[float],
    min_gap: float = MIN_FRAME_DIFF_SEC,
) -> List[Tuple[float, str]]:
    """
    Merge scene and regular timestamps, deduplicate, and preserve origin.
    Returns list of (timestamp, origin) tuples where origin is 'scene' or 'regular'.
    Scene timestamps get priority when close to regular timestamps.
    """
    # Create tagged list
    tagged = [(t, "scene") for t in scene_times] + [(t, "regular") for t in regular_times]
    tagged.sort(key=lambda x: x[0])

    if not tagged:
        return []

    result = [tagged[0]]
    for t, origin in tagged[1:]:
        prev_t, prev_origin = result[-1]
        gap = t - prev_t
        if gap < min_gap:
            # Keep scene over regular when too close
            if origin == "scene" and prev_origin == "regular":
                result[-1] = (t, origin)
            # Otherwise keep what we already have
        else:
            result.append((t, origin))

    return result


def extract_frames_batch(
    video_path: Path,
    frames_dir: Path,
    frame_list: List[Tuple[float, str]],
) -> List[Dict[str, Any]]:
    """
    Extract frames at specified timestamps.
    Uses individual ffmpeg calls with -ss before -i (fast keyframe seek).
    Returns list of frame dicts: {timestamp, path, source, index}
    """
    frames_dir.mkdir(parents=True, exist_ok=True)

    if not frame_list:
        return []

    # Build output frame info
    frames = []
    eta = ETATracker()
    total = len(frame_list)

    # Individual extraction (more reliable, -ss before -i is fast with keyframe seek)
    for i, (t, origin) in enumerate(frame_list):
        time_str = format_timestamp_filename(t)
        # Use sequential index to avoid filename collisions
        frame_path = frames_dir / f"frame_{i+1:04d}_{int(t):06d}_{time_str}.jpg"

        cmd = [
            "ffmpeg", "-hide_banner", "-loglevel", "error", "-y",
            "-ss", str(t),
            "-i", str(video_path),
            "-frames:v", "1",
            "-q:v", str(max(1, min(31, int((100 - FRAME_JPEG_QUALITY) * 31 / 100)))),
            str(frame_path),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True)

        if result.returncode == 0 and frame_path.exists():
            frames.append({
                "index": i + 1,
                "timestamp": t,
                "path": frame_path,
                "source": origin,
            })

        eta.update(i + 1)
        if (i + 1) % 20 == 0 or i + 1 == total:
            line = f"  {progress_bar(i + 1, total)}  ETA: {eta.get_eta(total)}"
            print(f"\r{line:<75}", end="", flush=True)

    print()
    return frames


def extract_all_frames(
    video_path: Path,
    frames_dir: Path,
    video_info: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Phase 1: Extract frames using scene detection + regular sampling.
    Returns list of frame dicts with origin preserved.
    """
    duration = video_info["duration"]

    # Scene detection
    scene_times = extract_frames_scene_detect(video_path)

    # Regular sampling
    regular_times = make_regular_timestamps(duration, REGULAR_INTERVAL_SEC)
    log(f"Regular sampling: {len(regular_times)} frames @ every {REGULAR_INTERVAL_SEC}s", 2)

    # Merge with origin tracking
    frame_list = merge_and_deduplicate(scene_times, regular_times)
    log(f"After merge+dedup: {len(frame_list)} unique frames", 2)

    # Extract frames
    frames = extract_frames_batch(video_path, frames_dir, frame_list)

    return frames


# ============================================================================
# Phase 1b: Black Frame / Jump Cut Detection
# ============================================================================

def detect_black_frames(frames: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Detect black/near-black frames (editing artifacts)."""
    import numpy as np
    from PIL import Image

    black_frames = []
    for frame in frames:
        try:
            img = Image.open(frame["path"]).convert("L")
            avg = np.mean(np.array(img))
            if avg < BLACK_FRAME_THRESHOLD:
                black_frames.append({
                    "timestamp": frame["timestamp"],
                    "type": "black_frame",
                    "detail": f"Average brightness: {avg:.1f}/255",
                })
        except Exception:
            pass

    return black_frames


def detect_jump_cuts(
    video_path: Path,
    duration: float,
) -> List[Dict[str, Any]]:
    """
    Detect potential jump cuts: scene changes that revert quickly.
    """
    cmd = [
        "ffmpeg", "-hide_banner", "-i", str(video_path),
        "-vf", f"select='gt(scene,{JUMP_CUT_MIN_DIFF})',metadata=print",
        "-f", "null", "-"
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)

    cuts = []
    prev_time = -999
    for line in result.stderr.split('\n'):
        if 'pts_time:' in line:
            match = re.search(r'pts_time:([\d.]+)', line)
            if match:
                t = float(match.group(1))
                gap = t - prev_time
                if 0.3 < gap < JUMP_CUT_MAX_RETURN_SEC:
                    cuts.append({
                        "timestamp": prev_time,
                        "type": "jump_cut",
                        "detail": f"Quick cut at {format_timestamp(prev_time)} "
                                  f"→ {format_timestamp(t)} ({gap:.1f}s gap)",
                    })
                prev_time = t

    return cuts


# ============================================================================
# Phase 2: OCR
# ============================================================================

def get_easyocr_reader():
    """Get or create EasyOCR reader with GPU fallback."""
    global _easyocr_reader
    if _easyocr_reader is None:
        import easyocr

        # Try GPU first, fall back to CPU
        try:
            log(f"Loading EasyOCR ({', '.join(OCR_LANGUAGES)}) with GPU...", 2)
            _easyocr_reader = easyocr.Reader(OCR_LANGUAGES, gpu=True)
            log("EasyOCR loaded (GPU)", 2)
        except Exception:
            log("GPU not available for EasyOCR, falling back to CPU...", 2)
            _easyocr_reader = easyocr.Reader(OCR_LANGUAGES, gpu=False)
            log("EasyOCR loaded (CPU)", 2)
    return _easyocr_reader


def run_ocr_on_frame(frame_path: Path) -> List[Dict[str, Any]]:
    """Run OCR on a single frame. Returns list of text detections."""
    reader = get_easyocr_reader()
    try:
        results = reader.readtext(str(frame_path))
        detections = []
        for bbox, text, confidence in results:
            if confidence >= OCR_CONFIDENCE_THRESHOLD and text.strip():
                detections.append({
                    "text": text.strip(),
                    "confidence": round(confidence, 3),
                    "bbox": [[int(round(c)) for c in point] for point in bbox],
                })
        return detections
    except Exception as e:
        log(f"OCR error on {frame_path.name}: {e}", 2)
        return []


def run_ocr_all_frames(
    frames: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Phase 2: Run OCR on all frames.
    Enriches frames with 'ocr_detections', 'ocr_text', 'has_text' fields.
    """
    eta = ETATracker()
    text_count = 0

    for i, frame in enumerate(frames):
        detections = run_ocr_on_frame(frame["path"])
        frame["ocr_detections"] = detections
        frame["ocr_text"] = " | ".join(d["text"] for d in detections) if detections else ""
        frame["has_text"] = len(detections) > 0

        if frame["has_text"]:
            text_count += 1

        eta.update(i + 1)
        if (i + 1) % 10 == 0 or i + 1 == len(frames):
            speed = eta.get_speed()
            speed_str = f"{speed:.1f} f/s" if speed > 0 else "..."
            line = (f"  {progress_bar(i + 1, len(frames))}  "
                    f"{speed_str}  Text: {text_count}  "
                    f"ETA: {eta.get_eta(len(frames))}")
            print(f"\r{line:<75}", end="", flush=True)

    print()
    return frames


def group_text_appearances(
    frames: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Group similar OCR text across consecutive frames into 'appearances'.
    Only groups when text persists continuously (does not disappear).
    """

    def text_similarity(a: str, b: str) -> float:
        if not a or not b:
            return 0.0
        a_low, b_low = a.lower(), b.lower()
        if a_low == b_low:
            return 1.0
        def trigrams(s):
            return set(s[i:i+3] for i in range(len(s) - 2))
        ta, tb = trigrams(a_low), trigrams(b_low)
        if not ta or not tb:
            return 1.0 if a_low == b_low else 0.0
        intersection = len(ta & tb)
        union = len(ta | tb)
        return intersection / union if union > 0 else 0.0

    appearances = []
    active: Dict[str, Dict] = {}

    all_sorted = sorted(frames, key=lambda f: f["timestamp"])

    for frame in all_sorted:
        current_texts = set()
        if frame.get("has_text"):
            for det in frame.get("ocr_detections", []):
                current_texts.add(det["text"])

        # Check which active appearances are still present
        finished_keys = []
        matched_texts = set()

        for key, app in active.items():
            matched = False
            for ct in current_texts:
                if ct not in matched_texts and text_similarity(key, ct) >= TEXT_GROUP_SIMILARITY:
                    matched = True
                    app["end_time"] = frame["timestamp"]
                    app["frame_count"] += 1
                    matched_texts.add(ct)
                    break

            if not matched:
                gap = frame["timestamp"] - app["end_time"]
                if gap > TEXT_GROUP_MAX_GAP_SEC:
                    finished_keys.append(key)

        for key in finished_keys:
            appearances.append(active.pop(key))

        # Start new appearances for unmatched text
        unmatched = current_texts - matched_texts
        for ct in unmatched:
            already_tracked = False
            for key in active:
                if text_similarity(key, ct) >= TEXT_GROUP_SIMILARITY:
                    already_tracked = True
                    break
            if not already_tracked:
                active[ct] = {
                    "text": ct,
                    "start_time": frame["timestamp"],
                    "end_time": frame["timestamp"],
                    "frame_count": 1,
                }

    # Flush remaining
    for app in active.values():
        appearances.append(app)

    appearances.sort(key=lambda a: a["start_time"])

    for app in appearances:
        app["duration"] = round(app["end_time"] - app["start_time"], 1)

    return appearances


# ============================================================================
# Phase 3: Vision LLM
# ============================================================================

def analyze_frame_with_vision(
    frame_path: Path,
    timestamp: float,
    model: str = DEFAULT_VISION_MODEL,
    prev_description: str = "",
    vision_context: str = "",
) -> Dict[str, Any]:
    """Send a frame to Ollama vision model for analysis.
    Includes previous frame description for continuity awareness."""
    import urllib.request

    with open(frame_path, "rb") as f:
        img_b64 = base64.b64encode(f.read()).decode("utf-8")

    context_parts = []
    if prev_description:
        context_parts.append(
            f"\nPrevious frame description: {prev_description[:200]}\n"
            "Note any significant changes from the previous frame."
        )

    channel_context = ""
    if vision_context:
        channel_context = f" for a YouTube channel about {vision_context}"

    prompt = (
        f"You are a video reviewer{channel_context}. "
        "Analyze this frame from a finished video. Be concise.\n\n"
        "Report:\n"
        "1. SCENE: What type of shot? (talking head, B-roll, graphic, intro, outro, lower third, etc.)\n"
        "2. TEXT: List ALL visible text on screen (titles, names, numbers, watermarks, logos).\n"
        "3. PEOPLE: How many people visible? What are they doing?\n"
        "4. ISSUES: Any visual problems? (bad framing, blurry, wrong text, color issues, artifacts)\n"
        "5. BRANDING: Is there a logo or watermark? Channel branding visible?\n"
        f"{''.join(context_parts)}\n"
        "\nBe factual and brief. If nothing notable, say 'Standard shot, no issues.'"
    )

    payload = {
        "model": model,
        "messages": [{
            "role": "user",
            "content": prompt,
            "images": [img_b64],
        }],
        "stream": False,
    }

    try:
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            f"{OLLAMA_URL}/api/chat",
            data=data,
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=VISION_TIMEOUT_SEC) as resp:
            response = json.loads(resp.read())
            content = response.get("message", {}).get("content", "")
            return {
                "timestamp": timestamp,
                "description": content,
                "model": model,
            }
    except Exception as e:
        return {
            "timestamp": timestamp,
            "description": f"ERROR: {e}",
            "model": model,
        }


def run_vision_analysis(
    frames: List[Dict[str, Any]],
    model: str = DEFAULT_VISION_MODEL,
    vision_context: str = "",
) -> List[Dict[str, Any]]:
    """
    Phase 3: Run vision LLM on priority frames.
    Priority: frames with OCR text → scene changes → sample of regular.
    Passes previous description for continuity awareness.
    """
    priority_frames = []
    seen_timestamps = set()

    # 1. Frames with text (highest priority — most likely to have errors)
    for frame in frames:
        if frame.get("has_text") and frame["timestamp"] not in seen_timestamps:
            priority_frames.append(frame)
            seen_timestamps.add(frame["timestamp"])

    # 2. Scene change frames
    for frame in frames:
        if frame.get("source") == "scene" and frame["timestamp"] not in seen_timestamps:
            priority_frames.append(frame)
            seen_timestamps.add(frame["timestamp"])

    # 3. Every 5th regular frame for coverage
    regular_only = [f for f in frames if f["timestamp"] not in seen_timestamps]
    for i, frame in enumerate(regular_only):
        if i % 5 == 0:
            priority_frames.append(frame)
            seen_timestamps.add(frame["timestamp"])

    priority_frames.sort(key=lambda f: f["timestamp"])
    log(f"Vision analysis on {len(priority_frames)} priority frames", 2)
    log(f"  With text: {sum(1 for f in priority_frames if f.get('has_text'))}", 2)
    log(f"  Scene changes: {sum(1 for f in priority_frames if f.get('source') == 'scene')}", 2)

    eta = ETATracker()
    results = []
    prev_desc = ""

    for i, frame in enumerate(priority_frames):
        result = analyze_frame_with_vision(
            frame["path"], frame["timestamp"], model,
            prev_description=prev_desc,
            vision_context=vision_context,
        )
        results.append(result)
        frame["vision_description"] = result["description"]
        prev_desc = result["description"]

        eta.update(i + 1)
        speed = eta.get_speed()
        speed_str = f"{1/speed:.1f}s/f" if speed > 0 else "..."
        line = (f"  {progress_bar(i + 1, len(priority_frames))}  "
                f"{speed_str}  ETA: {eta.get_eta(len(priority_frames))}")
        print(f"\r{line:<75}", end="", flush=True)

    print()
    return results


# ============================================================================
# Phase 4: Speech Transcription
# ============================================================================

def extract_audio(video_path: Path, output_path: Path) -> bool:
    """Extract audio from video to WAV (16kHz mono for Whisper)."""
    cmd = [
        "ffmpeg", "-hide_banner", "-loglevel", "error", "-y",
        "-i", str(video_path),
        "-vn", "-sn", "-dn",
        "-ar", str(AUDIO_SAMPLE_RATE),
        "-ac", "1",
        "-c:a", "pcm_s16le",
        str(output_path),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    return result.returncode == 0 and output_path.exists()


def get_whisper_model(model_size: str):
    global _whisper_model, _whisper_model_name

    if _whisper_model is not None and _whisper_model_name == model_size:
        return _whisper_model

    import whisper

    log(f"Loading Whisper model: {model_size}", 2)
    start_time = time.time()
    _whisper_model = whisper.load_model(model_size)
    _whisper_model_name = model_size
    elapsed = time.time() - start_time
    log(f"Model loaded in {elapsed:.1f}s", 2)

    return _whisper_model


def get_pyannote_pipeline(hf_token: str):
    global _pyannote_pipeline

    if _pyannote_pipeline is not None:
        return _pyannote_pipeline

    from pyannote.audio import Pipeline
    import torch

    log("Loading pyannote pipeline...", 2)
    start_time = time.time()
    _pyannote_pipeline = Pipeline.from_pretrained(
        "pyannote/speaker-diarization-3.1",
        use_auth_token=hf_token,
    )
    elapsed = time.time() - start_time
    log(f"Pipeline loaded in {elapsed:.1f}s", 2)

    if torch.backends.mps.is_available():
        _pyannote_pipeline.to(torch.device("mps"))
        log("Pipeline on Apple Silicon (MPS)", 2)
    elif torch.cuda.is_available():
        _pyannote_pipeline.to(torch.device("cuda"))
        log("Pipeline on CUDA", 2)

    return _pyannote_pipeline


def run_speech_transcription(
    video_path: Path,
    audio_path: Path,
    model_size: str = DEFAULT_WHISPER_MODEL,
    language: Optional[str] = None,
    num_speakers: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    Phase 4: Full speech transcription with speaker diarization.
    Returns list of segments: {start, end, text, speaker}
    """
    # Extract audio
    log("Extracting audio...", 1)
    if not extract_audio(video_path, audio_path):
        log("ERROR: Audio extraction failed", 1)
        return []

    size = audio_path.stat().st_size
    log(f"Audio: {format_size(size)}", 2)

    # Whisper
    log("Transcribing with Whisper...", 1)
    model = get_whisper_model(model_size)

    opts = {"word_timestamps": True, "verbose": False, "beam_size": 5}
    if language:
        opts["language"] = language

    start_time = time.time()
    whisper_result = model.transcribe(str(audio_path), **opts)
    elapsed = time.time() - start_time

    segments = whisper_result.get("segments", [])
    detected_lang = whisper_result.get("language", "unknown")
    log(f"Transcribed: {len(segments)} segments, language: {detected_lang}, "
        f"time: {format_duration(elapsed)}", 2)

    # Diarization
    hf_token = get_hf_token()
    if hf_token:
        log("Running speaker diarization...", 1)
        try:
            pipeline = get_pyannote_pipeline(hf_token)

            diar_opts = {}
            if num_speakers:
                diar_opts["num_speakers"] = num_speakers

            start_time = time.time()
            diarization = pipeline(str(audio_path), **diar_opts)
            elapsed = time.time() - start_time

            speaker_timeline = []
            for turn, _, speaker in diarization.itertracks(yield_label=True):
                speaker_timeline.append({
                    "start": turn.start,
                    "end": turn.end,
                    "speaker": speaker,
                })

            speakers_found = set(s["speaker"] for s in speaker_timeline)
            log(f"Diarization: {len(speakers_found)} speakers, "
                f"time: {format_duration(elapsed)}", 2)

            for seg in segments:
                best_speaker = ""
                best_overlap = 0
                for st in speaker_timeline:
                    overlap = min(seg["end"], st["end"]) - max(seg["start"], st["start"])
                    if overlap > best_overlap:
                        best_overlap = overlap
                        best_speaker = st["speaker"]
                seg["speaker"] = best_speaker

        except Exception as e:
            log(f"WARNING: Diarization failed: {e}", 1)
            for seg in segments:
                seg["speaker"] = ""
    else:
        log("Skipping diarization (no HF_TOKEN)", 1)
        for seg in segments:
            seg["speaker"] = ""

    result = []
    for seg in segments:
        result.append({
            "start": round(seg["start"], 2),
            "end": round(seg["end"], 2),
            "text": seg["text"].strip(),
            "speaker": seg.get("speaker", ""),
        })

    return result


# ============================================================================
# Phase 5: Audio Levels Check
# ============================================================================

def check_audio_levels(video_path: Path) -> Dict[str, Any]:
    """Check audio levels: LUFS, clipping, silence detection."""
    levels = {
        "integrated_lufs": None,
        "true_peak_dbtp": None,
        "lra": None,
        "clipping_detected": False,
        "silences": [],
    }

    # Loudness
    log("Measuring loudness (LUFS)...", 2)
    cmd = [
        "ffmpeg", "-hide_banner", "-i", str(video_path),
        "-af", "loudnorm=print_format=json",
        "-f", "null", "-"
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)

    json_match = re.search(r'\{[^{}]*"input_i"[^{}]*\}', result.stderr, re.DOTALL)
    if json_match:
        try:
            loudness_data = json.loads(json_match.group())
            levels["integrated_lufs"] = float(loudness_data.get("input_i", 0))
            levels["true_peak_dbtp"] = float(loudness_data.get("input_tp", 0))
            levels["lra"] = float(loudness_data.get("input_lra", 0))

            if levels["true_peak_dbtp"] is not None and levels["true_peak_dbtp"] > -1.0:
                levels["clipping_detected"] = True
        except (json.JSONDecodeError, ValueError):
            pass

    # Silence detection
    log("Detecting silences...", 2)
    cmd = [
        "ffmpeg", "-hide_banner", "-i", str(video_path),
        "-af", f"silencedetect=noise={SILENCE_THRESHOLD_DB}dB:d={LONG_SILENCE_SEC}",
        "-f", "null", "-"
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)

    silence_starts = re.findall(r'silence_start: ([\d.]+)', result.stderr)
    silence_durations = re.findall(r'silence_duration: ([\d.]+)', result.stderr)
    silence_ends = re.findall(r'silence_end: ([\d.]+)', result.stderr)

    for i, start in enumerate(silence_starts):
        silence = {
            "start": float(start),
            "duration": float(silence_durations[i]) if i < len(silence_durations) else 0,
        }
        if i < len(silence_ends):
            silence["end"] = float(silence_ends[i])
        levels["silences"].append(silence)

    log(f"Loudness: {levels['integrated_lufs']} LUFS, "
        f"Peak: {levels['true_peak_dbtp']} dBTP, "
        f"Silences: {len(levels['silences'])}", 2)

    return levels


# ============================================================================
# Phase 6: Cross-check with Source Transcript
# ============================================================================

def load_source_transcript(transcript_path: Path) -> List[Dict[str, str]]:
    """Load the apply_names transcript for cross-checking."""
    entries = []
    try:
        with open(transcript_path, "r", encoding="utf-8") as f:
            current_speaker = ""
            current_time = ""
            current_text = ""

            for line in f:
                line = line.rstrip()
                match = re.match(r'\[(\d+:\d+:\d+)\]\s*(.+?):', line)
                if match:
                    if current_text.strip():
                        entries.append({
                            "timestamp": current_time,
                            "speaker": current_speaker,
                            "text": current_text.strip(),
                        })
                    current_time = match.group(1)
                    current_speaker = match.group(2).strip()
                    current_text = ""
                else:
                    current_text += " " + line

            if current_text.strip():
                entries.append({
                    "timestamp": current_time,
                    "speaker": current_speaker,
                    "text": current_text.strip(),
                })

    except Exception as e:
        log(f"WARNING: Could not load transcript: {e}", 1)

    return entries


def cross_check_transcript(
    ocr_appearances: List[Dict[str, Any]],
    speech_segments: List[Dict[str, Any]],
    source_transcript: List[Dict[str, str]],
) -> List[Dict[str, Any]]:
    """Cross-check OCR names vs speech vs source transcript."""
    issues = []

    # Common non-name words that appear in lower thirds
    NON_NAME_WORDS = {
        "ceo", "coo", "cfo", "cto", "cmo", "vp", "svp", "evp", "md",
        "founder", "co-founder", "cofounder", "partner", "director",
        "manager", "president", "chairman", "host", "guest", "editor",
        "dubai", "abu dhabi", "riyadh", "saudi", "uae", "qatar", "bahrain",
        "free zone", "freezone", "license", "visa", "company", "business",
        "episode", "part", "subscribe", "like", "share", "follow",
    }

    def looks_like_name(text: str) -> bool:
        """Check if text looks like a person's name vs a title/role."""
        t = text.strip()
        if len(t) <= 2 or len(t) >= 40:
            return False
        # All-caps text is usually a title/brand, not a name: CEO, ROASTOLOGY, EPISODE 5
        alpha_chars = [c for c in t if c.isalpha()]
        if alpha_chars and all(c.isupper() for c in alpha_chars):
            return False
        # Check against known non-names
        if t.lower() in NON_NAME_WORDS:
            return False
        # Names typically have at least one uppercase letter
        if not any(c.isupper() for c in t):
            return False
        return True

    ocr_names = []
    for app in ocr_appearances:
        text = app["text"]
        parts = re.split(r'[|,\n]', text)
        if parts:
            name_candidate = parts[0].strip()
            if looks_like_name(name_candidate):
                ocr_names.append({
                    "name": name_candidate,
                    "timestamp": app["start_time"],
                    "full_text": text,
                })

    source_speakers = set()
    if source_transcript:
        for entry in source_transcript:
            if entry["speaker"]:
                source_speakers.add(entry["speaker"])

    for ocr_name in ocr_names:
        matched = False
        for speaker in source_speakers:
            if (ocr_name["name"].lower() in speaker.lower() or
                    speaker.lower() in ocr_name["name"].lower()):
                matched = True
                break
        if not matched and source_speakers:
            issues.append({
                "timestamp": ocr_name["timestamp"],
                "type": "name_check",
                "detail": f"On-screen name '{ocr_name['name']}' — "
                          f"verify against speakers: {', '.join(source_speakers)}",
            })

    return issues


# ============================================================================
# Phase 7: Tempo & Structure Analysis
# ============================================================================

def analyze_tempo_and_structure(
    frames: List[Dict[str, Any]],
    speech_segments: List[Dict[str, Any]],
    video_info: Dict[str, Any],
    vision_results: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """Analyze video tempo, shot duration, speaker balance, and structure."""
    duration = video_info["duration"]

    analysis = {
        "total_duration": duration,
        "scene_count": sum(1 for f in frames if f.get("source") == "scene"),
        "avg_shot_duration": 0,
        "speaker_balance": {},
        "content_mix": {"talking_head": 0, "broll": 0, "graphic": 0, "other": 0},
        "has_intro": False,
        "has_outro": False,
        "has_end_screen": False,
    }

    # Average shot duration
    scene_frames = sorted(
        [f for f in frames if f.get("source") == "scene"],
        key=lambda f: f["timestamp"]
    )
    if len(scene_frames) > 1:
        durations = [
            scene_frames[i]["timestamp"] - scene_frames[i-1]["timestamp"]
            for i in range(1, len(scene_frames))
        ]
        if durations:
            analysis["avg_shot_duration"] = round(sum(durations) / len(durations), 1)

    # Speaker balance
    if speech_segments:
        speaker_time = defaultdict(float)
        for seg in speech_segments:
            speaker = seg.get("speaker", "unknown") or "unknown"
            speaker_time[speaker] += seg["end"] - seg["start"]

        total_speech = sum(speaker_time.values())
        if total_speech > 0:
            for speaker, t in speaker_time.items():
                analysis["speaker_balance"][speaker] = {
                    "seconds": round(t, 1),
                    "percent": round(t / total_speech * 100, 1),
                }

    # Content mix from vision
    for vr in vision_results:
        desc = vr.get("description", "").lower()
        if any(w in desc for w in ["talking head", "interview", "speaking", "person talking"]):
            analysis["content_mix"]["talking_head"] += 1
        elif any(w in desc for w in ["b-roll", "broll", "exterior", "location", "establishing"]):
            analysis["content_mix"]["broll"] += 1
        elif any(w in desc for w in ["graphic", "chart", "infographic", "text overlay", "title card"]):
            analysis["content_mix"]["graphic"] += 1
        else:
            analysis["content_mix"]["other"] += 1

    # Structure checks
    for frame in frames:
        desc = (frame.get("vision_description", "") or "").lower()
        t = frame["timestamp"]

        if t < 10:
            if any(w in desc for w in ["intro", "logo", "animation", "brand"]):
                analysis["has_intro"] = True
        if t > duration - 15:
            if any(w in desc for w in ["outro", "end", "subscribe", "end screen"]):
                analysis["has_outro"] = True
            if any(w in desc for w in ["subscribe", "bell", "next video", "end screen"]):
                analysis["has_end_screen"] = True

    return analysis


# ============================================================================
# Warning Generation
# ============================================================================

def generate_warnings(
    frames: List[Dict[str, Any]],
    ocr_appearances: List[Dict[str, Any]],
    speech_segments: List[Dict[str, Any]],
    audio_levels: Dict[str, Any],
    black_frames: List[Dict[str, Any]],
    jump_cuts: List[Dict[str, Any]],
    cross_check_issues: List[Dict[str, Any]],
    tempo: Dict[str, Any],
    did_vision: bool = False,
) -> List[Dict[str, Any]]:
    """Collect all warnings from all phases."""
    warnings = []

    for bf in black_frames:
        warnings.append({
            "timestamp": bf["timestamp"], "type": "black_frame",
            "severity": "warning", "detail": bf["detail"],
        })

    for jc in jump_cuts:
        warnings.append({
            "timestamp": jc["timestamp"], "type": "jump_cut",
            "severity": "info", "detail": jc["detail"],
        })

    for silence in audio_levels.get("silences", []):
        warnings.append({
            "timestamp": silence["start"], "type": "long_silence",
            "severity": "warning", "detail": f"Silence: {silence['duration']:.1f}s",
        })

    if audio_levels.get("clipping_detected"):
        warnings.append({
            "timestamp": 0, "type": "clipping", "severity": "warning",
            "detail": f"Audio clipping detected (peak: {audio_levels['true_peak_dbtp']} dBTP)",
        })

    lufs = audio_levels.get("integrated_lufs")
    if lufs is not None:
        if lufs < -20:
            warnings.append({
                "timestamp": 0, "type": "audio_quiet", "severity": "warning",
                "detail": f"Audio too quiet ({lufs:.1f} LUFS). YouTube target: -14 LUFS",
            })
        elif lufs > -10:
            warnings.append({
                "timestamp": 0, "type": "audio_loud", "severity": "info",
                "detail": f"Audio loud ({lufs:.1f} LUFS). YouTube target: -14 LUFS",
            })

    for issue in cross_check_issues:
        warnings.append({
            "timestamp": issue["timestamp"], "type": issue["type"],
            "severity": "info", "detail": issue["detail"],
        })

    # Structure warnings only when Vision LLM was used (otherwise always false)
    if did_vision:
        if not tempo.get("has_intro"):
            warnings.append({
                "timestamp": 0, "type": "no_intro", "severity": "info",
                "detail": "No intro/branding detected in first 10 seconds",
            })

        if not tempo.get("has_end_screen"):
            warnings.append({
                "timestamp": tempo.get("total_duration", 0), "type": "no_end_screen",
                "severity": "info", "detail": "No end screen / subscribe CTA detected",
            })

    warnings.sort(key=lambda w: w["timestamp"])
    return warnings


# ============================================================================
# Output: Excel Report
# ============================================================================

def save_xlsx_report(
    output_path: Path,
    video_info: Dict[str, Any],
    frames: List[Dict[str, Any]],
    ocr_appearances: List[Dict[str, Any]],
    speech_segments: List[Dict[str, Any]],
    audio_levels: Dict[str, Any],
    warnings: List[Dict[str, Any]],
    tempo: Dict[str, Any],
    processing_time: float,
    mode: str,
    video_name: str,
) -> Path:
    """Save comprehensive review report as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_header(ws, row, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # === Sheet 1: Summary ===
    ws = wb.active
    ws.title = "Summary"

    summary_data = [
        ("VIDEO REVIEW REPORT", ""),
        ("", ""),
        ("Video", video_name),
        ("Duration", format_duration(video_info["duration"])),
        ("Resolution", f"{video_info['width']}×{video_info['height']}"),
        ("Codec", f"{video_info['codec']} @ {video_info['fps']:.1f} fps"),
        ("File Size", format_size(video_info["size_bytes"])),
        ("", ""),
        ("ANALYSIS RESULTS", ""),
        ("Mode", mode),
        ("Frames Extracted", len(frames)),
        ("Scene Changes", sum(1 for f in frames if f.get("source") == "scene")),
        ("On-Screen Text", f"{len(ocr_appearances)} appearances"),
        ("Speech Segments", len(speech_segments)),
        ("Speakers", len(tempo.get("speaker_balance", {}))),
        ("", ""),
        ("AUDIO", ""),
        ("Loudness", f"{audio_levels.get('integrated_lufs', 'N/A')} LUFS (target: -14)"),
        ("True Peak", f"{audio_levels.get('true_peak_dbtp', 'N/A')} dBTP"),
        ("Loudness Range", f"{audio_levels.get('lra', 'N/A')} LU"),
        ("Clipping", "⚠️ YES" if audio_levels.get("clipping_detected") else "✅ No"),
        ("Long Silences", f"{len(audio_levels.get('silences', []))}"),
        ("", ""),
        ("TEMPO", ""),
        ("Avg Shot Duration", f"{tempo.get('avg_shot_duration', 0):.1f}s"),
        ("Intro Detected", "✅ Yes" if tempo.get("has_intro") else "❌ No"),
        ("End Screen", "✅ Yes" if tempo.get("has_end_screen") else "❌ No"),
        ("", ""),
        ("WARNINGS", f"{len(warnings)}"),
        ("Processing Time", format_duration(processing_time)),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for row, (label, value) in enumerate(summary_data, 1):
        cell_l = ws.cell(row=row, column=1, value=label)
        cell_v = ws.cell(row=row, column=2, value=str(value))
        if label in ("VIDEO REVIEW REPORT", "ANALYSIS RESULTS", "AUDIO", "TEMPO", "WARNINGS"):
            cell_l.font = Font(bold=True, size=12, color="2F5496")
        if label == "WARNINGS" and len(warnings) > 0:
            cell_v.fill = warning_fill

    if tempo.get("speaker_balance"):
        row = len(summary_data) + 2
        ws.cell(row=row, column=1, value="SPEAKER BALANCE").font = Font(
            bold=True, size=12, color="2F5496"
        )
        row += 1
        for speaker, data in tempo["speaker_balance"].items():
            ws.cell(row=row, column=1, value=speaker)
            ws.cell(row=row, column=2,
                    value=f"{format_duration(data['seconds'])} ({data['percent']}%)")
            row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    # === Sheet 2: Timeline ===
    ws2 = wb.create_sheet("Timeline")
    write_header(ws2, 1, ["Timecode", "Scene #", "OCR Text", "Scene Description",
                           "Speech", "Speaker", "Flags"])

    all_events = defaultdict(lambda: {
        "ocr": "", "description": "", "speech": "", "speaker": "", "flags": "",
        "scene_num": "",
    })

    scene_num = 0
    for frame in sorted(frames, key=lambda f: f["timestamp"]):
        t = frame["timestamp"]
        key = round(t, 1)
        if frame.get("source") == "scene":
            scene_num += 1
        all_events[key]["scene_num"] = scene_num
        if frame.get("ocr_text"):
            existing = all_events[key]["ocr"]
            if existing and frame["ocr_text"] not in existing:
                all_events[key]["ocr"] = existing + " | " + frame["ocr_text"]
            else:
                all_events[key]["ocr"] = frame["ocr_text"]
        if frame.get("vision_description"):
            all_events[key]["description"] = frame["vision_description"]

    for seg in speech_segments:
        key = round(seg["start"], 1)
        if all_events[key]["speech"]:
            all_events[key]["speech"] += " " + seg["text"]
        else:
            all_events[key]["speech"] = seg["text"]
        # Preserve speaker — don't overwrite with empty
        if seg.get("speaker"):
            all_events[key]["speaker"] = seg["speaker"]

    for w in warnings:
        key = round(w["timestamp"], 1)
        flag = f"⚠️ {w['type']}: {w['detail']}"
        if all_events[key]["flags"]:
            all_events[key]["flags"] += " | " + flag
        else:
            all_events[key]["flags"] = flag

    row = 2
    for t in sorted(all_events.keys()):
        e = all_events[t]
        ws2.cell(row=row, column=1, value=format_timestamp(t))
        ws2.cell(row=row, column=2, value=e.get("scene_num", ""))
        ws2.cell(row=row, column=3, value=e.get("ocr", ""))
        ws2.cell(row=row, column=4, value=e.get("description", ""))
        ws2.cell(row=row, column=5, value=e.get("speech", ""))
        ws2.cell(row=row, column=6, value=e.get("speaker", ""))
        ws2.cell(row=row, column=7, value=e.get("flags", ""))
        if e.get("flags"):
            for col in range(1, 8):
                ws2.cell(row=row, column=col).fill = warning_fill
        row += 1

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 35
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 50
    ws2.column_dimensions["F"].width = 15
    ws2.column_dimensions["G"].width = 40
    ws2.freeze_panes = "A2"

    # === Sheet 3: Warnings ===
    ws3 = wb.create_sheet("Warnings")
    write_header(ws3, 1, ["Timecode", "Type", "Severity", "Detail"])
    for i, w in enumerate(warnings, 2):
        ws3.cell(row=i, column=1, value=format_timestamp(w["timestamp"]))
        ws3.cell(row=i, column=2, value=w["type"])
        ws3.cell(row=i, column=3, value=w["severity"])
        ws3.cell(row=i, column=4, value=w["detail"])
        if w["severity"] == "warning":
            for col in range(1, 5):
                ws3.cell(row=i, column=col).fill = warning_fill
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 60
    ws3.freeze_panes = "A2"

    # === Sheet 4: On-Screen Text ===
    ws4 = wb.create_sheet("On-Screen Text")
    write_header(ws4, 1, ["Start", "End", "Duration", "Text", "Frame Count"])
    for i, app in enumerate(ocr_appearances, 2):
        ws4.cell(row=i, column=1, value=format_timestamp(app["start_time"]))
        ws4.cell(row=i, column=2, value=format_timestamp(app["end_time"]))
        ws4.cell(row=i, column=3, value=f"{app['duration']:.1f}s")
        ws4.cell(row=i, column=4, value=app["text"])
        ws4.cell(row=i, column=5, value=app["frame_count"])
    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 60
    ws4.column_dimensions["E"].width = 12
    ws4.freeze_panes = "A2"

    # === Sheet 5: Speech ===
    ws5 = wb.create_sheet("Speech")
    write_header(ws5, 1, ["Timecode", "End", "Speaker", "Text"])
    for i, seg in enumerate(speech_segments, 2):
        ws5.cell(row=i, column=1, value=format_timestamp(seg["start"]))
        ws5.cell(row=i, column=2, value=format_timestamp(seg["end"]))
        ws5.cell(row=i, column=3, value=seg.get("speaker", ""))
        ws5.cell(row=i, column=4, value=seg["text"])
    ws5.column_dimensions["A"].width = 10
    ws5.column_dimensions["B"].width = 10
    ws5.column_dimensions["C"].width = 15
    ws5.column_dimensions["D"].width = 80
    ws5.freeze_panes = "A2"

    wb.save(output_path)
    log(f"Saved: {output_path.name}", 1)
    return output_path


# ============================================================================
# Output: Markdown Report
# ============================================================================

def save_markdown_report(
    output_path: Path, video_info: Dict, frames: List, ocr_appearances: List,
    speech_segments: List, audio_levels: Dict, warnings: List, tempo: Dict,
    processing_time: float, mode: str, video_name: str,
) -> Path:
    """Save review report as Markdown."""
    lines = [
        f"# Video Review Report: {video_name}", "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Mode: {mode} | Processing time: {format_duration(processing_time)}", "",
        "## Video Info", "",
        "| Property | Value |", "|----------|-------|",
        f"| Duration | {format_duration(video_info['duration'])} |",
        f"| Resolution | {video_info['width']}×{video_info['height']} |",
        f"| Codec | {video_info['codec']} @ {video_info['fps']:.1f} fps |",
        f"| Size | {format_size(video_info['size_bytes'])} |", "",
        f"## Warnings ({len(warnings)})", "",
    ]

    if warnings:
        lines += ["| Timecode | Type | Detail |", "|----------|------|--------|"]
        for w in warnings:
            lines.append(f"| {format_timestamp(w['timestamp'])} | {w['type']} | {w['detail']} |")
    else:
        lines.append("✅ No warnings found.")
    lines.append("")

    lines += [
        "## Audio Levels", "",
        "| Metric | Value | Target |", "|--------|-------|--------|",
        f"| Loudness | {audio_levels.get('integrated_lufs', 'N/A')} LUFS | -14 LUFS |",
        f"| True Peak | {audio_levels.get('true_peak_dbtp', 'N/A')} dBTP | < -1 dBTP |",
        f"| Clipping | {'⚠️ YES' if audio_levels.get('clipping_detected') else '✅ No'} | No |", "",
        f"## On-Screen Text ({len(ocr_appearances)} appearances)", "",
    ]

    if ocr_appearances:
        lines += ["| Start | End | Duration | Text |", "|-------|-----|----------|------|"]
        for app in ocr_appearances:
            lines.append(f"| {format_timestamp(app['start_time'])} "
                         f"| {format_timestamp(app['end_time'])} "
                         f"| {app['duration']:.1f}s | {app['text']} |")
    lines.append("")

    lines += [
        "## Tempo & Structure", "",
        "| Metric | Value |", "|--------|-------|",
        f"| Scene Changes | {tempo.get('scene_count', 0)} |",
        f"| Avg Shot Duration | {tempo.get('avg_shot_duration', 0):.1f}s |",
        f"| Intro | {'✅' if tempo.get('has_intro') else '❌'} |",
        f"| End Screen | {'✅' if tempo.get('has_end_screen') else '❌'} |", "",
    ]

    if tempo.get("speaker_balance"):
        lines += ["### Speaker Balance", ""]
        for speaker, data in tempo["speaker_balance"].items():
            lines.append(f"- **{speaker}**: {format_duration(data['seconds'])} ({data['percent']}%)")
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")
    log(f"Saved: {output_path.name}", 1)
    return output_path


# ============================================================================
# Output: JSON
# ============================================================================

def save_json_report(
    output_path: Path, video_info: Dict, frames: List, ocr_appearances: List,
    speech_segments: List, audio_levels: Dict, warnings: List, tempo: Dict,
    processing_time: float, mode: str, video_name: str,
) -> Path:
    """Save review data as JSON for downstream scripts."""
    clean_frames = [{
        "index": f.get("index", 0),
        "timestamp": f["timestamp"],
        "source": f.get("source", ""),
        "ocr_text": f.get("ocr_text", ""),
        "has_text": f.get("has_text", False),
        "vision_description": f.get("vision_description", ""),
        "frame_file": f["path"].name if isinstance(f.get("path"), Path) else "",
    } for f in frames]

    report = {
        "version": VERSION,
        "generated": datetime.now().isoformat(),
        "mode": mode,
        "processing_time_sec": round(processing_time, 1),
        "video": {
            "name": video_name,
            "duration": video_info["duration"],
            "width": video_info["width"],
            "height": video_info["height"],
            "codec": video_info["codec"],
            "fps": video_info["fps"],
            "size_bytes": video_info["size_bytes"],
        },
        "frames": clean_frames,
        "ocr_appearances": ocr_appearances,
        "speech_segments": speech_segments,
        "audio_levels": audio_levels,
        "warnings": warnings,
        "tempo": tempo,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False, default=str)

    log(f"Saved: {output_path.name}", 1)
    return output_path


# ============================================================================
# Main Processing Pipeline
# ============================================================================

def process_video(
    video_path: Path,
    do_vision: bool = True,
    do_speech: bool = True,
    language: Optional[str] = None,
    num_speakers: Optional[int] = None,
    vision_model: str = DEFAULT_VISION_MODEL,
    vision_context: str = "",
    whisper_model: str = DEFAULT_WHISPER_MODEL,
    transcript_path: Optional[Path] = None,
    resume: bool = False,
) -> Dict[str, Any]:
    """Run the full review pipeline on a video."""

    video_name = video_path.stem
    output_dir = video_path.parent / f"{video_name}_review"
    output_dir.mkdir(parents=True, exist_ok=True)
    frames_dir = output_dir / "frames"
    logs_dir = output_dir / "logs"

    # Init log file
    init_log_file(logs_dir)

    # Get video info
    video_info = get_video_info(video_path)

    # Determine mode label and phase count
    parts = []
    if do_vision:
        parts.append("vision")
    if do_speech:
        parts.append("speech")
    if not parts:
        mode_label = "basic"
    elif len(parts) == 2:
        mode_label = "full"
    else:
        mode_label = parts[0]  # "vision" or "speech"
    phases = []
    phases.append("Frames")
    phases.append("OCR")
    if do_vision:
        phases.append("Vision LLM")
    if do_speech and video_info.get("has_audio"):
        phases.append("Speech")
    if video_info.get("has_audio"):
        phases.append("Audio")
    if transcript_path:
        phases.append("Cross-check")
    total_phases = len(phases)

    # ========================================
    # Work Estimate
    # ========================================
    est = estimate_work(video_info, mode_label, do_speech, do_vision, REGULAR_INTERVAL_SEC)

    BOX_W = 68  # inner width of the box

    def box_line(text: str) -> str:
        """Format a line inside the box with proper padding."""
        if len(text) > BOX_W:
            text = text[:BOX_W - 3] + "..."
        return f"║{text:<{BOX_W}}║"

    out("")
    out("╔" + "═" * BOX_W + "╗")
    out(box_line(f"  VIDEO REVIEW — {video_name}"))
    out("╠" + "═" * BOX_W + "╣")
    dur_str = format_duration(video_info["duration"])
    res_str = f"{video_info['width']}×{video_info['height']}"
    size_str = format_size(video_info["size_bytes"])
    codec_str = video_info["codec"]
    out(box_line(f"  Duration: {dur_str:>8}  │  Resolution: {res_str:<20}    "))
    out(box_line(f"  Size:     {size_str:>8}  │  Codec: {codec_str:<25}    "))
    vision_str = vision_model if do_vision else "off"
    out(box_line(f"  Mode:     {mode_label:>8}  │  Vision: {vision_str:<24}    "))
    out("╠" + "═" * BOX_W + "╣")
    out(box_line(f"  WORK ESTIMATE"))
    disk_str = f"{est['est_disk_mb']:.0f} MB"
    out(box_line(f"  Frames: ~{est['est_unique_frames']:<6}  │  Vision frames: ~{est['est_vision_frames']:<6}     "))
    out(box_line(f"  Est. time: ~{format_duration(est['est_time_sec']):<10}│  Est. disk: ~{disk_str:<10}     "))
    phases_str = " → ".join(phases)
    out(box_line(f"  Phases: {phases_str}"))
    out("╚" + "═" * BOX_W + "╝")
    out("")

    pipeline_start = time.time()
    phase_idx = 0
    phase_times: Dict[str, float] = {}

    # ========================================
    # Phase: Extract Frames
    # ========================================
    phase_idx += 1
    out(f"Phase {phase_idx}/{total_phases}: Extracting frames")
    phase_start = time.time()

    # Resume: check if frames already extracted
    cached_frames = load_intermediate(output_dir, "frames") if resume else None
    if cached_frames and frames_dir.exists():
        log("Resuming: using cached frames", 1)
        frames = []
        for cf in cached_frames:
            p = frames_dir / cf["frame_file"]
            if p.exists():
                frames.append({
                    "index": cf["index"],
                    "timestamp": cf["timestamp"],
                    "path": p,
                    "source": cf["source"],
                })
        log(f"Loaded {len(frames)} cached frames", 2)
    else:
        frames = extract_all_frames(video_path, frames_dir, video_info)
        # Save intermediate
        save_intermediate([{
            "index": f["index"], "timestamp": f["timestamp"],
            "source": f["source"], "frame_file": f["path"].name,
        } for f in frames], output_dir, "frames")

    scene_count = sum(1 for f in frames if f.get("source") == "scene")
    log(f"Scene changes: {scene_count}", 1)
    log(f"Total frames: {len(frames)}", 1)

    if not frames:
        log("ERROR: No frames extracted. Cannot continue.", 0)
        log("Check that the video file is valid and not empty.", 1)
        close_log_file()
        return {
            "video_name": video_name,
            "output_dir": str(output_dir),
            "warnings": 0,
            "processing_time": time.time() - pipeline_start,
            "error": "No frames extracted",
        }

    # Black frames and jump cuts
    log("Checking for black frames...", 1)
    black_frames = detect_black_frames(frames)
    if black_frames:
        log(f"Found {len(black_frames)} black frames", 2)

    log("Checking for jump cuts...", 1)
    jump_cuts = detect_jump_cuts(video_path, video_info["duration"])
    if jump_cuts:
        log(f"Found {len(jump_cuts)} potential jump cuts", 2)

    phase_elapsed = time.time() - phase_start
    phase_times["Frames"] = phase_elapsed
    log(f"✅ Done ({format_duration(phase_elapsed)})", 1)
    out("")

    # ========================================
    # Phase: OCR
    # ========================================
    phase_idx += 1
    out(f"Phase {phase_idx}/{total_phases}: OCR — scanning text on screen")
    phase_start = time.time()

    # Resume: check if OCR already done
    cached_ocr = load_intermediate(output_dir, "ocr") if resume else None
    if cached_ocr:
        log("Resuming: using cached OCR results", 1)
        # Restore OCR data onto frames
        ocr_by_ts = {item["timestamp"]: item for item in cached_ocr}
        for frame in frames:
            cached = ocr_by_ts.get(frame["timestamp"])
            if cached:
                frame["ocr_detections"] = cached.get("ocr_detections", [])
                frame["ocr_text"] = cached.get("ocr_text", "")
                frame["has_text"] = cached.get("has_text", False)
            else:
                frame["ocr_detections"] = []
                frame["ocr_text"] = ""
                frame["has_text"] = False
        text_frame_count = sum(1 for f in frames if f.get("has_text"))
        log(f"Loaded cached OCR: text on {text_frame_count} frames", 2)
    else:
        frames = run_ocr_all_frames(frames)
        text_frame_count = sum(1 for f in frames if f.get("has_text"))
        log(f"Text found on {text_frame_count} frames", 1)

        # Save OCR intermediate for resume
        save_intermediate([{
            "timestamp": f["timestamp"],
            "ocr_detections": f.get("ocr_detections", []),
            "ocr_text": f.get("ocr_text", ""),
            "has_text": f.get("has_text", False),
        } for f in frames], output_dir, "ocr")

    log("Grouping text appearances...", 1)
    ocr_appearances = group_text_appearances(frames)
    log(f"Grouped into {len(ocr_appearances)} unique appearances", 2)

    phase_elapsed = time.time() - phase_start
    phase_times["OCR"] = phase_elapsed
    log(f"✅ Done ({format_duration(phase_elapsed)})", 1)
    out("")

    # ========================================
    # Phase: Vision LLM (if enabled)
    # ========================================
    vision_results = []
    if do_vision:
        phase_idx += 1
        out(f"Phase {phase_idx}/{total_phases}: Vision LLM — analyzing scenes ({vision_model})")
        phase_start = time.time()

        vision_results = run_vision_analysis(frames, vision_model, vision_context=vision_context)

        phase_elapsed = time.time() - phase_start
        phase_times["Vision"] = phase_elapsed
        log(f"Analyzed {len(vision_results)} frames", 1)
        log(f"✅ Done ({format_duration(phase_elapsed)})", 1)
        out("")

    # ========================================
    # Phase: Speech (if enabled)
    # ========================================
    speech_segments = []
    if do_speech:
        if not video_info.get("has_audio"):
            log("⚠️  Skipping speech: video has no audio track", 1)
            out("")
        else:
            # Resume: check if speech already done
            cached_speech = load_intermediate(output_dir, "speech") if resume else None
            if cached_speech:
                phase_idx += 1
                out(f"Phase {phase_idx}/{total_phases}: Speech — resumed from cache")
                speech_segments = cached_speech
                speakers = set(s.get("speaker", "") for s in speech_segments if s.get("speaker"))
                log(f"Loaded: {len(speech_segments)} segments, {len(speakers)} speakers", 2)
                phase_times["Speech"] = 0
                out("")
            else:
                phase_idx += 1
                out(f"Phase {phase_idx}/{total_phases}: Speech transcription (Whisper {whisper_model})")
                phase_start = time.time()

                audio_path = output_dir / f"{video_name}_audio.wav"
                speech_segments = run_speech_transcription(
                    video_path, audio_path,
                    model_size=whisper_model,
                    language=language,
                    num_speakers=num_speakers,
                )

                speakers = set(s.get("speaker", "") for s in speech_segments if s.get("speaker"))
                log(f"Segments: {len(speech_segments)}, Speakers: {len(speakers)}", 1)

                # Save intermediate for resume
                save_intermediate(speech_segments, output_dir, "speech")

                # Clean up audio
                if audio_path.exists():
                    audio_path.unlink()

                phase_elapsed = time.time() - phase_start
                phase_times["Speech"] = phase_elapsed
                log(f"✅ Done ({format_duration(phase_elapsed)})", 1)
                out("")

    # ========================================
    # Phase: Audio Levels
    # ========================================
    audio_levels = {
        "integrated_lufs": None, "true_peak_dbtp": None, "lra": None,
        "clipping_detected": False, "silences": [],
    }
    if not video_info.get("has_audio"):
        log("⚠️  Skipping audio levels: video has no audio track", 1)
        out("")
    else:
        phase_idx += 1
        out(f"Phase {phase_idx}/{total_phases}: Audio levels check")
        phase_start = time.time()

        audio_levels = check_audio_levels(video_path)

        phase_elapsed = time.time() - phase_start
        phase_times["Audio"] = phase_elapsed
        log(f"✅ Done ({format_duration(phase_elapsed)})", 1)
        out("")

    # ========================================
    # Phase: Cross-check (if transcript provided)
    # ========================================
    cross_check_issues = []
    if transcript_path:
        phase_idx += 1
        out(f"Phase {phase_idx}/{total_phases}: Cross-check with source transcript")
        phase_start = time.time()

        source_transcript = load_source_transcript(transcript_path)
        log(f"Loaded {len(source_transcript)} entries from source", 2)

        cross_check_issues = cross_check_transcript(
            ocr_appearances, speech_segments, source_transcript
        )
        log(f"Cross-check issues: {len(cross_check_issues)}", 2)

        phase_elapsed = time.time() - phase_start
        phase_times["Cross-check"] = phase_elapsed
        log(f"✅ Done ({format_duration(phase_elapsed)})", 1)
        out("")

    # ========================================
    # Analyze & Report
    # ========================================
    tempo = analyze_tempo_and_structure(frames, speech_segments, video_info, vision_results)

    warnings = generate_warnings(
        frames, ocr_appearances, speech_segments,
        audio_levels, black_frames, jump_cuts,
        cross_check_issues, tempo,
        did_vision=do_vision,
    )

    processing_time = time.time() - pipeline_start

    out("Saving reports...")

    xlsx_path = output_dir / f"{video_name}_review.xlsx"
    save_xlsx_report(
        xlsx_path, video_info, frames, ocr_appearances,
        speech_segments, audio_levels, warnings, tempo,
        processing_time, mode_label, video_name,
    )

    md_path = output_dir / f"{video_name}_review.md"
    save_markdown_report(
        md_path, video_info, frames, ocr_appearances,
        speech_segments, audio_levels, warnings, tempo,
        processing_time, mode_label, video_name,
    )

    json_path = output_dir / f"{video_name}_review.json"
    save_json_report(
        json_path, video_info, frames, ocr_appearances,
        speech_segments, audio_levels, warnings, tempo,
        processing_time, mode_label, video_name,
    )

    # ========================================
    # Final Summary
    # ========================================
    speakers = set(s.get("speaker", "") for s in speech_segments if s.get("speaker"))

    out("")
    out("═" * 70)
    out("RESULTS")
    out("═" * 70)
    out(f"  Scenes:           {scene_count}")
    out(f"  On-screen text:   {len(ocr_appearances)} appearances")
    if do_speech:
        out(f"  Speakers:         {len(speakers)}"
              + (f" ({', '.join(sorted(speakers))})" if speakers else ""))
    out(f"  ⚠️  Warnings:     {len(warnings)}")

    for w in warnings:
        out(f"     • {format_timestamp(w['timestamp'])}  {w['type']}: {w['detail']}")

    out(f"")
    if phase_times:
        out(f"  Phase timing:")
        for pname, ptime in phase_times.items():
            out(f"     {pname:<15} {format_duration(ptime):>10}")
        out(f"     {'─' * 26}")
        out(f"     {'TOTAL':<15} {format_duration(processing_time):>10}")
        out(f"")

    out(f"  Output: {output_dir}/")
    out(f"  Excel:  {xlsx_path.name}")
    out("═" * 70)

    close_log_file()

    return {
        "video_name": video_name,
        "output_dir": str(output_dir),
        "warnings": len(warnings),
        "processing_time": processing_time,
    }


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Video review pipeline — analyze finished videos before publishing",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Version: {VERSION}

Examples:
  Preflight check only:
    %(prog)s --preflight

  Quick review (OCR + Speech + Audio, no Vision LLM):
    %(prog)s video.mp4 --quick

  Full review (OCR + Vision LLM + Speech + Audio):
    %(prog)s video.mp4

  Visual only (no speech transcription):
    %(prog)s video.mp4 --skip-speech

  Full review with cross-check:
    %(prog)s video.mp4 --transcript path/to/apply_names.txt

  Resume after interruption:
    %(prog)s video.mp4 --resume

Before first run:
  1. pip install easyocr Pillow openpyxl numpy openai-whisper torch pyannote.audio soundfile
  2. OLLAMA_MAX_VRAM=20g ollama serve   (Terminal 1)
  3. ollama pull minicpm-v
  4. export HF_TOKEN="hf_xxx"
        """
    )

    parser.add_argument("input", nargs="?", help="Video file to analyze")

    # Mode
    parser.add_argument("--quick", action="store_true",
                        help="Quick mode: skip Vision LLM")
    parser.add_argument("--skip-speech", action="store_true",
                        help="Skip speech transcription (visual-only review)")
    parser.add_argument("--skip-vision", action="store_true",
                        help="Skip Vision LLM analysis")

    # Speech
    parser.add_argument("-l", "--language", default=None,
                        help="Language code (en, ru, ar)")
    parser.add_argument("-n", "--num-speakers", type=int, default=None,
                        help="Number of speakers")
    parser.add_argument("-m", "--whisper-model", default=DEFAULT_WHISPER_MODEL,
                        choices=["tiny", "base", "small", "medium", "large", "large-v2", "large-v3"],
                        help=f"Whisper model (default: {DEFAULT_WHISPER_MODEL})")

    # Vision
    parser.add_argument("--vision-model", default=DEFAULT_VISION_MODEL,
                        help=f"Ollama vision model (default: {DEFAULT_VISION_MODEL})")
    parser.add_argument("--vision-context", default="",
                        help="Channel context for vision prompt (e.g. 'business in UAE')")

    # Cross-check
    parser.add_argument("--transcript", type=str, default=None,
                        help="Path to source transcript for cross-checking")

    # Frame extraction
    parser.add_argument("--interval", type=float, default=REGULAR_INTERVAL_SEC,
                        help=f"Frame interval in seconds (default: {REGULAR_INTERVAL_SEC})")
    parser.add_argument("--scene-threshold", type=float, default=SCENE_THRESHOLD,
                        help=f"Scene detection threshold 0-1 (default: {SCENE_THRESHOLD})")

    # OCR
    parser.add_argument("--ocr-languages", type=str, default=",".join(OCR_LANGUAGES),
                        help=f"OCR languages, comma-separated (default: {','.join(OCR_LANGUAGES)})")

    # Utility
    parser.add_argument("--resume", action="store_true",
                        help="Resume from last run (skip completed phases)")
    parser.add_argument("--preflight", action="store_true",
                        help="Run preflight checks only")
    parser.add_argument("--skip-preflight", action="store_true",
                        help="Skip preflight checks")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show work estimate without processing")
    parser.add_argument("--version", action="version",
                        version=f"%(prog)s {VERSION}")

    args = parser.parse_args()

    # Apply config overrides
    mod = sys.modules[__name__]
    mod.REGULAR_INTERVAL_SEC = args.interval
    mod.SCENE_THRESHOLD = args.scene_threshold
    mod.OCR_LANGUAGES = args.ocr_languages.split(",")

    # Determine what to run
    do_vision = not (args.quick or args.skip_vision)
    do_speech = not args.skip_speech

    # Input path
    input_path: Optional[Path] = None
    if args.input:
        input_path = Path(args.input).resolve()

    # Preflight only
    if args.preflight:
        result = run_preflight_checks(input_path, do_vision=do_vision, do_speech=do_speech)
        sys.exit(0 if result.ok else 1)

    # No input
    if not args.input:
        parser.print_help()
        print("\nError: No input specified. Use --preflight to check setup.")
        sys.exit(1)

    # Validate input
    if not input_path.exists():
        print(f"ERROR: File not found: {input_path}")
        sys.exit(1)

    if input_path.suffix.lower() not in VIDEO_EXTS:
        print(f"ERROR: Not a video file: {input_path.suffix}")
        print(f"       Supported: {', '.join(sorted(VIDEO_EXTS))}")
        sys.exit(1)

    # Run preflight
    if not args.skip_preflight:
        result = run_preflight_checks(input_path, do_vision=do_vision, do_speech=do_speech)
        if not result.ok:
            sys.exit(1)
        print("")

    # Dry run
    if args.dry_run:
        vinfo = get_video_info(input_path)
        est = estimate_work(vinfo, "full" if do_vision else "quick",
                            do_speech, do_vision, REGULAR_INTERVAL_SEC)
        print("")
        log("DRY RUN — work estimate:")
        log(f"Video: {input_path.name}", 1)
        log(f"Duration: {format_duration(vinfo['duration'])}", 1)
        log(f"Resolution: {vinfo['width']}×{vinfo['height']}", 1)
        log(f"Vision: {'on' if do_vision else 'off'}, Speech: {'on' if do_speech else 'off'}", 1)
        log(f"Est. frames: ~{est['est_unique_frames']}", 1)
        log(f"Est. vision frames: ~{est['est_vision_frames']}", 1)
        log(f"Est. disk: ~{est['est_disk_mb']:.0f} MB", 1)
        log(f"Est. time: ~{format_duration(est['est_time_sec'])}", 1)
        log(f"  Frames: ~{format_duration(est['breakdown']['frames'])}", 2)
        log(f"  OCR: ~{format_duration(est['breakdown']['ocr'])}", 2)
        if do_vision:
            log(f"  Vision LLM: ~{format_duration(est['breakdown']['vision'])}", 2)
        if do_speech:
            log(f"  Speech: ~{format_duration(est['breakdown']['speech'])}", 2)
        log(f"  Audio: ~{format_duration(est['breakdown']['audio'])}", 2)
        log("(--dry-run: no processing performed)")
        sys.exit(0)

    # Transcript path
    transcript_path = None
    if args.transcript:
        transcript_path = Path(args.transcript).resolve()
        if not transcript_path.exists():
            print(f"WARNING: Transcript not found: {transcript_path}")
            transcript_path = None

    # Process
    try:
        result = process_video(
            video_path=input_path,
            do_vision=do_vision,
            do_speech=do_speech,
            language=args.language,
            num_speakers=args.num_speakers,
            vision_model=args.vision_model,
            vision_context=args.vision_context,
            whisper_model=args.whisper_model,
            transcript_path=transcript_path,
            resume=args.resume,
        )
        sys.exit(0)

    except KeyboardInterrupt:
        print("\n")
        log("INTERRUPTED by user (Ctrl+C)")
        close_log_file()
        sys.exit(1)
    except Exception as e:
        log(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        close_log_file()
        sys.exit(1)


if __name__ == "__main__":
    main()
