#!/usr/bin/env python3
"""
01_generate_premiere_xml.py

Generate Adobe Premiere Pro XML (FCP XML 7) from edit brief XLSX.

Features:
- Creates sequence with clips on timeline
- Organizes clips into bins by block
- Adds markers with colors and comments
- Creates subclips with in/out points
- Supports multiple video/audio tracks
- YouTube chapter marker generation

Usage:
    python 01_generate_premiere_xml.py --input edit_brief.xlsx --output project.xml
    python 01_generate_premiere_xml.py -i brief.xlsx -o project.xml --source "/Volumes/RYA"
    python 01_generate_premiere_xml.py -i brief.xlsx --validate-only

Author: YTAI Pipeline / RYA.AE
"""

import argparse
import json
import os
import re
import subprocess
import sys
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CONSTANTS
# =============================================================================

MARKER_COLORS = {
    "cyan": 0,
    "blue": 1,
    "green": 2,
    "yellow": 3,
    "red": 4,
    "magenta": 5,
    "orange": 6,
    "purple": 7,
}

BLOCK_COLOR_MAP = {
    1: "green",      # Introduction
    2: "cyan",       # Beginning
    3: "yellow",     # Diagnosis
    4: "blue",       # Explanation
    5: "orange",     # Symptoms
    6: "purple",     # Treatment
    7: "red",        # Risks
    8: "magenta",    # Surgery
    9: "cyan",       # Current state
    10: "blue",      # Statistics
    11: "yellow",    # Alternatives
    12: "green",     # Solution
    13: "green",     # Motivation
    14: "green",     # Closing
    99: "orange",    # Extra/B-roll
}

DEFAULT_PROJECT_SETTINGS = {
    "project_name": "Edit Sequence",
    "fps": 29.97,
    "width": 3840,
    "height": 2160,
    "pixel_aspect": 1.0,
    "field_order": "progressive",
    "sample_rate": 48000,
    "audio_channels": 2,
    "video_tracks": 3,
    "audio_tracks": 4,
    "source_folder": "",
    "output_folder": "",
    "create_subclips": True,
    "create_bins": True,
    "create_chapter_markers": True,
    "include_unused": True,
    "nested_sequences": False,
    "link_audio": True,
    "add_handles": False,
    "handle_frames": 12,
}

REQUIRED_SEGMENT_COLUMNS = ["source_file", "tc_in", "tc_out", "block", "block_name"]


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class Segment:
    """Represents a single video segment from the edit brief."""
    segment_id: str
    source_file: str
    tc_in: float  # in seconds
    tc_out: float  # in seconds
    block: int
    block_name: str
    segment_name: str = ""
    speaker: str = ""
    transcript: str = ""
    track: str = "V1"
    color: str = ""
    use: bool = True
    priority: int = 1
    is_chapter: bool = False
    broll_note: str = ""
    notes: str = ""
    
    @property
    def duration(self) -> float:
        return self.tc_out - self.tc_in
    
    @property
    def display_name(self) -> str:
        if self.segment_name:
            return f"{self.segment_id}_{self.segment_name}"
        return f"{self.segment_id}_{self.block_name}"


@dataclass
class ProjectSettings:
    """Project configuration from the project sheet."""
    project_name: str = "Edit Sequence"
    fps: float = 29.97
    width: int = 3840
    height: int = 2160
    pixel_aspect: float = 1.0
    field_order: str = "progressive"
    sample_rate: int = 48000
    audio_channels: int = 2
    video_tracks: int = 3
    audio_tracks: int = 4
    source_folder: str = ""
    output_folder: str = ""
    create_subclips: bool = True
    create_bins: bool = True
    create_chapter_markers: bool = True
    include_unused: bool = True
    nested_sequences: bool = False
    link_audio: bool = True
    add_handles: bool = False
    handle_frames: int = 12


@dataclass
class SourceFile:
    """Represents a source video file."""
    filename: str
    path: str
    duration: float = 0.0
    exists: bool = False


# =============================================================================
# TIMECODE UTILITIES
# =============================================================================

def parse_timecode(tc_str: str, fps: float = 29.97) -> float:
    """
    Parse various timecode formats to seconds.
    
    Supported formats:
    - MM:SS (06:34)
    - MM:SS.ms (06:34.500)
    - HH:MM:SS (00:06:34)
    - HH:MM:SS:FF (00:06:34:15)
    - seconds (394.5)
    """
    if tc_str is None:
        return 0.0
    
    tc_str = str(tc_str).strip()
    
    if not tc_str:
        return 0.0
    
    # Try parsing as raw seconds
    try:
        return float(tc_str)
    except ValueError:
        pass
    
    # Split by colons
    parts = tc_str.replace(",", ".").split(":")
    
    try:
        if len(parts) == 2:
            # MM:SS or MM:SS.ms
            minutes = int(parts[0])
            seconds = float(parts[1])
            return minutes * 60 + seconds
        
        elif len(parts) == 3:
            # HH:MM:SS or HH:MM:SS.ms
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = float(parts[2])
            return hours * 3600 + minutes * 60 + seconds
        
        elif len(parts) == 4:
            # HH:MM:SS:FF (with frames)
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = int(parts[2])
            frames = int(parts[3])
            return hours * 3600 + minutes * 60 + seconds + (frames / fps)
        
        else:
            raise ValueError(f"Invalid timecode format: {tc_str}")
    
    except (ValueError, IndexError) as e:
        raise ValueError(f"Cannot parse timecode '{tc_str}': {e}")


def seconds_to_frames(seconds: float, fps: float) -> int:
    """Convert seconds to frame count."""
    return int(round(seconds * fps))


def frames_to_timecode(frames: int, fps: float) -> str:
    """Convert frames to HH:MM:SS:FF timecode string."""
    total_seconds = frames / fps
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = int(total_seconds % 60)
    frame = int(round((total_seconds % 1) * fps))
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frame:02d}"


def seconds_to_timecode(secs: float, fps: float) -> str:
    """Convert seconds to HH:MM:SS:FF timecode string."""
    return frames_to_timecode(seconds_to_frames(secs, fps), fps)


# =============================================================================
# FILE UTILITIES
# =============================================================================

def get_video_duration(video_path: str) -> Optional[float]:
    """Get video duration in seconds using ffprobe."""
    try:
        cmd = [
            "ffprobe", "-v", "quiet",
            "-show_entries", "format=duration",
            "-of", "default=noprint_wrappers=1:nokey=1",
            video_path
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode == 0 and result.stdout.strip():
            return float(result.stdout.strip())
    except (subprocess.TimeoutExpired, FileNotFoundError, ValueError):
        pass
    return None


def find_source_file(filename: str, source_folder: str) -> Tuple[str, bool]:
    """
    Find source file in folder, return (path, exists).
    Tries with and without extension.
    """
    if not source_folder:
        return filename, False
    
    # Try exact match
    full_path = os.path.join(source_folder, filename)
    if os.path.exists(full_path):
        return full_path, True
    
    # Try without extension
    base_name = os.path.splitext(filename)[0]
    for ext in [".MP4", ".mp4", ".MOV", ".mov", ".MXF", ".mxf"]:
        full_path = os.path.join(source_folder, base_name + ext)
        if os.path.exists(full_path):
            return full_path, True
    
    # Try recursive search
    source_path = Path(source_folder)
    for ext in [".MP4", ".mp4", ".MOV", ".mov", ".MXF", ".mxf"]:
        matches = list(source_path.rglob(base_name + ext))
        if matches:
            return str(matches[0]), True
    
    return os.path.join(source_folder, filename), False


# =============================================================================
# XLSX PARSING
# =============================================================================

def parse_bool(value) -> bool:
    """Parse boolean from various formats."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    s = str(value).strip().lower()
    return s in ("true", "yes", "1", "y", "t")


def load_edit_brief(xlsx_path: str) -> Tuple[List[Segment], ProjectSettings]:
    """
    Load edit brief from XLSX file.
    Returns (segments, project_settings).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    # Load project settings
    settings = ProjectSettings(**DEFAULT_PROJECT_SETTINGS.copy())
    
    if "project" in wb.sheetnames:
        ws_project = wb["project"]
        for row in ws_project.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1] is not None:
                key = str(row[0]).strip().lower().replace(" ", "_")
                value = row[1]
                
                if hasattr(settings, key):
                    field_type = type(getattr(settings, key))
                    if field_type == bool:
                        setattr(settings, key, parse_bool(value))
                    elif field_type == int:
                        setattr(settings, key, int(float(value)))
                    elif field_type == float:
                        setattr(settings, key, float(value))
                    else:
                        setattr(settings, key, str(value))
    
    # Load segments
    segments = []
    
    if "segments" in wb.sheetnames:
        ws_segments = wb["segments"]
    else:
        # Try first sheet
        ws_segments = wb.active
    
    # Get header row
    headers = []
    for cell in ws_segments[1]:
        if cell.value:
            headers.append(str(cell.value).strip().lower().replace(" ", "_"))
        else:
            headers.append("")
    
    # Validate required columns
    missing = [col for col in REQUIRED_SEGMENT_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    
    # Parse rows
    segment_count = 0
    for row_idx, row in enumerate(ws_segments.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not any(row):
            continue
        
        # Create dict from row
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                row_data[headers[i]] = value
        
        # Skip if required fields are empty
        if not row_data.get("source_file") or not row_data.get("tc_in"):
            continue
        
        segment_count += 1
        
        # Parse segment
        try:
            segment = Segment(
                segment_id=str(row_data.get("segment_id") or f"seg_{segment_count:03d}"),
                source_file=str(row_data["source_file"]).strip(),
                tc_in=parse_timecode(row_data["tc_in"], settings.fps),
                tc_out=parse_timecode(row_data["tc_out"], settings.fps),
                block=int(float(row_data.get("block", 1))),
                block_name=str(row_data.get("block_name", "")).strip(),
                segment_name=str(row_data.get("segment_name", "") or "").strip(),
                speaker=str(row_data.get("speaker", "") or "").strip(),
                transcript=str(row_data.get("transcript", "") or "").strip()[:500],
                track=str(row_data.get("track", "V1") or "V1").strip().upper(),
                color=str(row_data.get("color", "") or "").strip().lower(),
                use=parse_bool(row_data.get("use", True)),
                priority=int(float(row_data.get("priority", 1) or 1)),
                is_chapter=parse_bool(row_data.get("is_chapter", False)),
                broll_note=str(row_data.get("broll_note", "") or "").strip(),
                notes=str(row_data.get("notes", "") or "").strip(),
            )
            
            # Validate
            if segment.tc_out <= segment.tc_in:
                print(f"Warning: Row {row_idx} has tc_out <= tc_in, skipping")
                continue
            
            # Set default color from block if not specified
            if not segment.color:
                segment.color = BLOCK_COLOR_MAP.get(segment.block, "cyan")
            
            segments.append(segment)
            
        except Exception as e:
            print(f"Warning: Error parsing row {row_idx}: {e}")
            continue
    
    wb.close()
    
    return segments, settings


# =============================================================================
# XML GENERATION
# =============================================================================

def generate_uuid() -> str:
    """Generate a UUID for XML elements."""
    return str(uuid.uuid4()).upper()


def create_rate_element(parent: ET.Element, fps: float) -> ET.Element:
    """Create rate element with timebase and ntsc flag."""
    rate = ET.SubElement(parent, "rate")
    
    # Determine timebase and NTSC flag
    if abs(fps - 29.97) < 0.01:
        ET.SubElement(rate, "timebase").text = "30"
        ET.SubElement(rate, "ntsc").text = "TRUE"
    elif abs(fps - 23.976) < 0.01:
        ET.SubElement(rate, "timebase").text = "24"
        ET.SubElement(rate, "ntsc").text = "TRUE"
    elif abs(fps - 59.94) < 0.01:
        ET.SubElement(rate, "timebase").text = "60"
        ET.SubElement(rate, "ntsc").text = "TRUE"
    else:
        ET.SubElement(rate, "timebase").text = str(int(round(fps)))
        ET.SubElement(rate, "ntsc").text = "FALSE"
    
    return rate


def create_marker_element(
    parent: ET.Element,
    name: str,
    comment: str,
    in_frame: int,
    out_frame: int,
    color: str = "green"
) -> ET.Element:
    """Create a marker element."""
    marker = ET.SubElement(parent, "marker")
    ET.SubElement(marker, "name").text = name[:100]  # Premiere limit
    ET.SubElement(marker, "comment").text = comment[:255] if comment else ""
    ET.SubElement(marker, "in").text = str(in_frame)
    ET.SubElement(marker, "out").text = str(out_frame)
    
    # Note: FCP XML marker colors are limited
    # Premiere interprets them differently
    
    return marker


def create_file_element(
    parent: ET.Element,
    file_id: str,
    name: str,
    path: str,
    duration_frames: int,
    fps: float,
    width: int,
    height: int,
    sample_rate: int
) -> ET.Element:
    """Create a file reference element."""
    file_elem = ET.SubElement(parent, "file", id=file_id)
    ET.SubElement(file_elem, "name").text = name
    ET.SubElement(file_elem, "pathurl").text = f"file://localhost{path}"
    
    create_rate_element(file_elem, fps)
    ET.SubElement(file_elem, "duration").text = str(duration_frames)
    
    # Media description
    media = ET.SubElement(file_elem, "media")
    
    # Video
    video = ET.SubElement(media, "video")
    ET.SubElement(video, "duration").text = str(duration_frames)
    sample_char = ET.SubElement(video, "samplecharacteristics")
    ET.SubElement(sample_char, "width").text = str(width)
    ET.SubElement(sample_char, "height").text = str(height)
    ET.SubElement(sample_char, "pixelaspectratio").text = "square"
    
    # Audio
    audio = ET.SubElement(media, "audio")
    ET.SubElement(audio, "channelcount").text = "2"
    audio_sample = ET.SubElement(audio, "samplecharacteristics")
    ET.SubElement(audio_sample, "samplerate").text = str(sample_rate)
    ET.SubElement(audio_sample, "depth").text = "16"
    
    return file_elem


def generate_premiere_xml(
    segments: List[Segment],
    settings: ProjectSettings,
    source_files: Dict[str, SourceFile]
) -> str:
    """
    Generate Premiere Pro XML from segments and settings.
    """
    fps = settings.fps
    
    # Create root
    root = ET.Element("xmeml", version="5")
    
    # Create project
    project = ET.SubElement(root, "project")
    ET.SubElement(project, "name").text = settings.project_name
    
    # Children container
    children = ET.SubElement(project, "children")
    
    # ==========================================================================
    # BIN: 01_Sources - Original source files
    # ==========================================================================
    
    sources_bin = ET.SubElement(children, "bin")
    ET.SubElement(sources_bin, "name").text = "01_Sources"
    sources_children = ET.SubElement(sources_bin, "children")
    
    file_refs = {}  # filename -> file_id mapping
    
    for filename, source in source_files.items():
        file_id = f"file-{generate_uuid()[:8]}"
        file_refs[filename] = file_id
        
        duration_frames = seconds_to_frames(source.duration, fps) if source.duration else 0
        if duration_frames == 0:
            # Estimate from segments
            max_tc = max(
                (s.tc_out for s in segments if s.source_file == filename),
                default=3600
            )
            duration_frames = seconds_to_frames(max_tc + 60, fps)
        
        # Create clip in bin
        clip = ET.SubElement(sources_children, "clip", id=f"masterclip-{file_id}")
        ET.SubElement(clip, "name").text = filename
        ET.SubElement(clip, "duration").text = str(duration_frames)
        create_rate_element(clip, fps)
        
        # File reference
        create_file_element(
            clip, file_id, filename, source.path,
            duration_frames, fps,
            settings.width, settings.height,
            settings.sample_rate
        )
    
    # ==========================================================================
    # BIN: 02_Blocks - Organized by block
    # ==========================================================================
    
    if settings.create_bins:
        blocks_bin = ET.SubElement(children, "bin")
        ET.SubElement(blocks_bin, "name").text = "02_Blocks"
        blocks_children = ET.SubElement(blocks_bin, "children")
        
        # Group segments by block
        blocks = {}
        for seg in segments:
            if seg.use and seg.priority == 1:
                if seg.block not in blocks:
                    blocks[seg.block] = []
                blocks[seg.block].append(seg)
        
        for block_num in sorted(blocks.keys()):
            block_segments = blocks[block_num]
            block_name = block_segments[0].block_name if block_segments else f"Block_{block_num}"
            
            block_bin = ET.SubElement(blocks_children, "bin")
            ET.SubElement(block_bin, "name").text = f"Block_{block_num:02d}_{block_name}"
            block_children = ET.SubElement(block_bin, "children")
            
            if settings.create_subclips:
                for seg in block_segments:
                    if seg.source_file not in file_refs:
                        continue
                    
                    subclip = ET.SubElement(block_children, "clip")
                    ET.SubElement(subclip, "name").text = seg.display_name
                    
                    in_frames = seconds_to_frames(seg.tc_in, fps)
                    out_frames = seconds_to_frames(seg.tc_out, fps)
                    
                    ET.SubElement(subclip, "duration").text = str(out_frames - in_frames)
                    create_rate_element(subclip, fps)
                    ET.SubElement(subclip, "in").text = str(in_frames)
                    ET.SubElement(subclip, "out").text = str(out_frames)
                    
                    # Reference master clip
                    ET.SubElement(subclip, "masterclipid").text = f"masterclip-{file_refs[seg.source_file]}"
    
    # ==========================================================================
    # BIN: 03_Alternatives - Priority > 1
    # ==========================================================================
    
    alt_segments = [s for s in segments if s.use and s.priority > 1]
    if alt_segments and settings.include_unused:
        alts_bin = ET.SubElement(children, "bin")
        ET.SubElement(alts_bin, "name").text = "03_Alternatives"
        alts_children = ET.SubElement(alts_bin, "children")
        
        for seg in alt_segments:
            if seg.source_file not in file_refs:
                continue
            
            subclip = ET.SubElement(alts_children, "clip")
            ET.SubElement(subclip, "name").text = f"ALT_{seg.display_name}"
            
            in_frames = seconds_to_frames(seg.tc_in, fps)
            out_frames = seconds_to_frames(seg.tc_out, fps)
            
            ET.SubElement(subclip, "duration").text = str(out_frames - in_frames)
            create_rate_element(subclip, fps)
            ET.SubElement(subclip, "in").text = str(in_frames)
            ET.SubElement(subclip, "out").text = str(out_frames)
            ET.SubElement(subclip, "masterclipid").text = f"masterclip-{file_refs[seg.source_file]}"
    
    # ==========================================================================
    # BIN: 04_Unused - use=FALSE
    # ==========================================================================
    
    unused_segments = [s for s in segments if not s.use]
    if unused_segments and settings.include_unused:
        unused_bin = ET.SubElement(children, "bin")
        ET.SubElement(unused_bin, "name").text = "04_Unused"
        unused_children = ET.SubElement(unused_bin, "children")
        
        for seg in unused_segments:
            if seg.source_file not in file_refs:
                continue
            
            subclip = ET.SubElement(unused_children, "clip")
            ET.SubElement(subclip, "name").text = f"UNUSED_{seg.display_name}"
            
            in_frames = seconds_to_frames(seg.tc_in, fps)
            out_frames = seconds_to_frames(seg.tc_out, fps)
            
            ET.SubElement(subclip, "duration").text = str(out_frames - in_frames)
            create_rate_element(subclip, fps)
            ET.SubElement(subclip, "in").text = str(in_frames)
            ET.SubElement(subclip, "out").text = str(out_frames)
            ET.SubElement(subclip, "masterclipid").text = f"masterclip-{file_refs[seg.source_file]}"
    
    # ==========================================================================
    # SEQUENCE: Main Edit
    # ==========================================================================
    
    # Get segments for timeline (use=TRUE, priority=1)
    timeline_segments = [s for s in segments if s.use and s.priority == 1]
    timeline_segments.sort(key=lambda s: (s.block, segments.index(s)))
    
    # Calculate total duration
    total_duration = sum(s.duration for s in timeline_segments)
    total_frames = seconds_to_frames(total_duration, fps)
    
    # Create sequence
    sequence = ET.SubElement(children, "sequence", id="sequence-1")
    ET.SubElement(sequence, "name").text = settings.project_name
    ET.SubElement(sequence, "duration").text = str(total_frames)
    create_rate_element(sequence, fps)
    
    # Timecode
    timecode = ET.SubElement(sequence, "timecode")
    ET.SubElement(timecode, "string").text = "00:00:00:00"
    ET.SubElement(timecode, "frame").text = "0"
    create_rate_element(timecode, fps)
    ET.SubElement(timecode, "displayformat").text = "DF" if abs(fps - 29.97) < 0.01 else "NDF"
    
    # Media container
    seq_media = ET.SubElement(sequence, "media")
    
    # --------------------------------------------------------------------------
    # Video tracks
    # --------------------------------------------------------------------------
    
    seq_video = ET.SubElement(seq_media, "video")
    
    # Format
    fmt = ET.SubElement(seq_video, "format")
    sample_char = ET.SubElement(fmt, "samplecharacteristics")
    ET.SubElement(sample_char, "width").text = str(settings.width)
    ET.SubElement(sample_char, "height").text = str(settings.height)
    ET.SubElement(sample_char, "pixelaspectratio").text = "square"
    create_rate_element(sample_char, fps)
    
    # Create video tracks
    video_tracks = {f"V{i+1}": ET.SubElement(seq_video, "track") for i in range(settings.video_tracks)}
    
    # Add clips to V1 (main) and handle V2 for b-roll placeholders
    timeline_position = 0
    
    for seg in timeline_segments:
        if seg.source_file not in file_refs:
            print(f"Warning: Source file not found for segment {seg.segment_id}")
            continue
        
        track_name = seg.track if seg.track in video_tracks else "V1"
        track = video_tracks[track_name]
        
        in_frames = seconds_to_frames(seg.tc_in, fps)
        out_frames = seconds_to_frames(seg.tc_out, fps)
        duration_frames = out_frames - in_frames
        
        clipitem = ET.SubElement(track, "clipitem", id=f"clipitem-{seg.segment_id}")
        ET.SubElement(clipitem, "name").text = seg.display_name
        ET.SubElement(clipitem, "duration").text = str(duration_frames)
        create_rate_element(clipitem, fps)
        
        ET.SubElement(clipitem, "start").text = str(timeline_position)
        ET.SubElement(clipitem, "end").text = str(timeline_position + duration_frames)
        ET.SubElement(clipitem, "in").text = str(in_frames)
        ET.SubElement(clipitem, "out").text = str(out_frames)
        
        ET.SubElement(clipitem, "masterclipid").text = f"masterclip-{file_refs[seg.source_file]}"
        
        # File reference
        file_elem = ET.SubElement(clipitem, "file", id=f"clipfile-{seg.segment_id}")
        ET.SubElement(file_elem, "pathurl").text = f"file://localhost{source_files[seg.source_file].path}"
        
        # Add marker on clip
        marker_comment = seg.transcript[:200] if seg.transcript else seg.notes[:200] if seg.notes else ""
        create_marker_element(
            clipitem,
            name=f"BLOCK {seg.block}: {seg.block_name}",
            comment=marker_comment,
            in_frame=0,
            out_frame=duration_frames,
            color=seg.color
        )
        
        # Create B-roll placeholder on V2 if specified
        if seg.broll_note and "V2" in video_tracks:
            broll_clip = ET.SubElement(video_tracks["V2"], "clipitem")
            ET.SubElement(broll_clip, "name").text = f"BROLL: {seg.broll_note}"
            ET.SubElement(broll_clip, "duration").text = str(duration_frames)
            create_rate_element(broll_clip, fps)
            ET.SubElement(broll_clip, "start").text = str(timeline_position)
            ET.SubElement(broll_clip, "end").text = str(timeline_position + duration_frames)
            ET.SubElement(broll_clip, "enabled").text = "FALSE"  # Disabled placeholder
        
        timeline_position += duration_frames
    
    # --------------------------------------------------------------------------
    # Audio tracks
    # --------------------------------------------------------------------------
    
    seq_audio = ET.SubElement(seq_media, "audio")
    
    # Format
    audio_fmt = ET.SubElement(seq_audio, "format")
    audio_sample = ET.SubElement(audio_fmt, "samplecharacteristics")
    ET.SubElement(audio_sample, "samplerate").text = str(settings.sample_rate)
    ET.SubElement(audio_sample, "depth").text = "16"
    
    # Create audio tracks
    for i in range(settings.audio_tracks):
        audio_track = ET.SubElement(seq_audio, "track")
        
        if i == 0 and settings.link_audio:
            # A1 gets linked audio from V1
            timeline_position = 0
            for seg in timeline_segments:
                if seg.source_file not in file_refs:
                    continue
                if seg.track != "V1":
                    continue
                
                in_frames = seconds_to_frames(seg.tc_in, fps)
                out_frames = seconds_to_frames(seg.tc_out, fps)
                duration_frames = out_frames - in_frames
                
                audio_clipitem = ET.SubElement(audio_track, "clipitem", id=f"audioclip-{seg.segment_id}")
                ET.SubElement(audio_clipitem, "name").text = seg.display_name
                ET.SubElement(audio_clipitem, "duration").text = str(duration_frames)
                create_rate_element(audio_clipitem, fps)
                
                ET.SubElement(audio_clipitem, "start").text = str(timeline_position)
                ET.SubElement(audio_clipitem, "end").text = str(timeline_position + duration_frames)
                ET.SubElement(audio_clipitem, "in").text = str(in_frames)
                ET.SubElement(audio_clipitem, "out").text = str(out_frames)
                
                ET.SubElement(audio_clipitem, "masterclipid").text = f"masterclip-{file_refs[seg.source_file]}"
                
                timeline_position += duration_frames
    
    # --------------------------------------------------------------------------
    # Sequence-level markers (block starts + chapters)
    # --------------------------------------------------------------------------
    
    timeline_position = 0
    current_block = None
    
    for seg in timeline_segments:
        duration_frames = seconds_to_frames(seg.duration, fps)
        
        # Add block start marker
        if seg.block != current_block:
            current_block = seg.block
            create_marker_element(
                sequence,
                name=f"BLOCK {seg.block}: {seg.block_name}",
                comment=seg.notes or seg.transcript[:200] if seg.transcript else "",
                in_frame=timeline_position,
                out_frame=timeline_position + duration_frames,
                color=seg.color
            )
        
        # Add YouTube chapter marker
        if seg.is_chapter and settings.create_chapter_markers:
            tc_string = frames_to_timecode(timeline_position, fps)
            create_marker_element(
                sequence,
                name=f"CHAPTER: {seg.block_name}",
                comment=f"YouTube chapter at {tc_string}",
                in_frame=timeline_position,
                out_frame=timeline_position + 1,
                color="green"
            )
        
        timeline_position += duration_frames
    
    # Convert to string with pretty print
    # Note: Using custom indent function because minidom can corrupt tag names
    def indent_xml(elem, level=0):
        """Add pretty-print indentation to XML."""
        i = "\n" + level * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
            for child in elem:
                indent_xml(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = i
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = i
    
    indent_xml(root)
    return ET.tostring(root, encoding="unicode")


# =============================================================================
# VALIDATION
# =============================================================================

def validate_edit_brief(segments: List[Segment], settings: ProjectSettings) -> List[str]:
    """Validate edit brief and return list of warnings/errors."""
    issues = []
    
    if not segments:
        issues.append("ERROR: No segments found in edit brief")
        return issues
    
    # Check for required fields
    for seg in segments:
        if not seg.source_file:
            issues.append(f"ERROR: {seg.segment_id} missing source_file")
        if seg.tc_out <= seg.tc_in:
            issues.append(f"ERROR: {seg.segment_id} has invalid timecode range")
        if seg.duration < 0.5:
            issues.append(f"WARNING: {seg.segment_id} is very short ({seg.duration:.1f}s)")
        if seg.duration > 300:
            issues.append(f"WARNING: {seg.segment_id} is very long ({seg.duration:.1f}s)")
    
    # Check for non-sequential blocks
    blocks = sorted(set(s.block for s in segments))
    expected = list(range(min(blocks), max(blocks) + 1))
    missing_blocks = set(expected) - set(blocks)
    if missing_blocks and 99 not in missing_blocks:
        issues.append(f"WARNING: Missing block numbers: {sorted(missing_blocks)}")
    
    # Check for duplicate segment IDs
    ids = [s.segment_id for s in segments]
    duplicates = set(x for x in ids if ids.count(x) > 1)
    if duplicates:
        issues.append(f"WARNING: Duplicate segment IDs: {duplicates}")
    
    return issues


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate Premiere Pro XML from edit brief XLSX"
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Path to edit_brief.xlsx"
    )
    parser.add_argument(
        "--output", "-o",
        default="premiere_project.xml",
        help="Output XML path (default: premiere_project.xml)"
    )
    parser.add_argument(
        "--source", "-s",
        help="Override source folder path"
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Validate edit brief without generating XML"
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Show detailed progress"
    )
    
    args = parser.parse_args()
    
    # Check input file
    if not os.path.exists(args.input):
        print(f"Error: Input file not found: {args.input}")
        sys.exit(1)
    
    print(f"Loading edit brief: {args.input}")
    
    try:
        segments, settings = load_edit_brief(args.input)
    except Exception as e:
        print(f"Error loading edit brief: {e}")
        sys.exit(1)
    
    print(f"  Found {len(segments)} segments")
    print(f"  Project: {settings.project_name}")
    print(f"  FPS: {settings.fps}")
    print(f"  Resolution: {settings.width}x{settings.height}")
    
    # Override source folder if specified
    if args.source:
        settings.source_folder = args.source
    
    # Validate
    issues = validate_edit_brief(segments, settings)
    for issue in issues:
        print(f"  {issue}")
    
    if any(issue.startswith("ERROR") for issue in issues):
        print("\nCannot proceed due to errors.")
        sys.exit(1)
    
    if args.validate_only:
        print("\nValidation complete.")
        sys.exit(0)
    
    # Find source files
    print("\nLocating source files...")
    source_files = {}
    unique_sources = set(s.source_file for s in segments)
    
    for filename in unique_sources:
        path, exists = find_source_file(filename, settings.source_folder)
        duration = get_video_duration(path) if exists else None
        source_files[filename] = SourceFile(
            filename=filename,
            path=path,
            duration=duration or 0.0,
            exists=exists
        )
        
        status = "✓" if exists else "✗ NOT FOUND"
        duration_str = f" ({duration:.1f}s)" if duration else ""
        if args.verbose:
            print(f"  {filename}: {status}{duration_str}")
    
    found_count = sum(1 for s in source_files.values() if s.exists)
    print(f"  Found {found_count}/{len(source_files)} source files")
    
    # Generate XML
    print("\nGenerating Premiere Pro XML...")
    
    try:
        xml_content = generate_premiere_xml(segments, settings, source_files)
    except Exception as e:
        print(f"Error generating XML: {e}")
        sys.exit(1)
    
    # Write output
    with open(args.output, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write('<!DOCTYPE xmeml>\n')
        # Fix ElementTree bug where 'name' tag outputs as 'n'
        fixed_xml = xml_content.replace('<name>', '<name>').replace('</name>', '</name>')
        f.write(fixed_xml)
    
    print(f"\n✅ Created: {args.output}")
    print(f"   Segments: {len([s for s in segments if s.use and s.priority == 1])}")
    print(f"   Blocks: {len(set(s.block for s in segments if s.use))}")
    print(f"   Duration: {sum(s.duration for s in segments if s.use and s.priority == 1):.1f}s")
    
    print("\nTo import in Premiere Pro:")
    print("   1. File → Import")
    print(f"   2. Select {args.output}")
    print("   3. Relink media if needed (Right-click → Link Media)")


if __name__ == "__main__":
    main()
